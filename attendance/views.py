from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.db.models import Q, Sum, Count, Avg
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt, ensure_csrf_cookie
from django.core.paginator import Paginator
from datetime import datetime, timedelta, date, time
import json
import csv

from .models import (
    AttendanceRecord, WorkSchedule, EmployeeSchedule,
    OvertimeRequest, AttendanceCorrection, BiometricDevice,
    BiometricTemplate, BreakRecord, AttendancePolicy
)
from .forms import (
    AttendanceRecordForm, WorkScheduleForm,
    EmployeeScheduleForm, OvertimeRequestForm, AttendanceCorrectionForm,
    BiometricDeviceForm, BreakRecordForm, AttendanceReportForm
)
from .biometric_utils import get_biometric_service, FingerprintTemplate
from hr.models import Employee
from core.utils import get_current_employee


@login_required
def attendance_dashboard(request):
    """Main attendance dashboard"""
    today = timezone.now().date()

    # Get current user's employee record
    employee = get_current_employee(request)

    # Today's attendance summary
    total_employees = Employee.objects.filter(employment_status='active').count()
    present_today = AttendanceRecord.objects.filter(
        date=today,
        status__in=['present', 'late', 'early_leave']
    ).count()
    absent_today = AttendanceRecord.objects.filter(
        date=today,
        status='absent'
    ).count()
    on_leave_today = AttendanceRecord.objects.filter(
        date=today,
        status='leave'
    ).count()

    # Recent attendance records
    recent_records = AttendanceRecord.objects.select_related('employee').order_by('-date', '-clock_in')[:10]

    # Pending requests
    pending_overtime = OvertimeRequest.objects.filter(status='pending').count()
    pending_corrections = AttendanceCorrection.objects.filter(status='pending').count()

    # Employee's own attendance
    employee_attendance = None
    if employee:
        employee_attendance = AttendanceRecord.objects.filter(
            employee=employee,
            date=today
        ).first()

    context = {
        'today': today,
        'total_employees': total_employees,
        'present_today': present_today,
        'absent_today': absent_today,
        'on_leave_today': on_leave_today,
        'recent_records': recent_records,
        'pending_overtime': pending_overtime,
        'pending_corrections': pending_corrections,
        'employee_attendance': employee_attendance,
        'employee': employee,
    }

    return render(request, 'attendance/dashboard.html', context)


@login_required
def attendance_list(request):
    """List attendance records with filters"""
    records = AttendanceRecord.objects.select_related('employee', 'schedule', 'current_project').order_by('-date', 'employee')

    # Filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    employee_id = request.GET.get('employee')
    status = request.GET.get('status')
    project_id = request.GET.get('project')

    if date_from:
        records = records.filter(date__gte=date_from)
    if date_to:
        records = records.filter(date__lte=date_to)
    if employee_id:
        records = records.filter(employee_id=employee_id)
    if status:
        records = records.filter(status=status)
    if project_id:
        records = records.filter(current_project_id=project_id)

    # Calculate statistics
    from django.db.models import Count, Sum
    total_records = records.count()
    present_count = records.filter(status='present').count()
    late_count = records.filter(status='late').count()
    absent_count = records.filter(status='absent').count()
    total_overtime = records.aggregate(total=Sum('overtime_hours'))['total'] or 0

    # Pagination
    per_page = request.GET.get('per_page', 25)
    try:
        per_page = int(per_page)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get employees for filter dropdown
    employees = Employee.objects.filter(employment_status='active').order_by('first_name', 'last_name')

    # Get active projects for filter dropdown and edit modal
    from project.models import Project
    projects_list = Project.objects.filter(
        status__in=['open', 'in_progress', 'on_hold']
    ).order_by('project_name')

    # For edit modal (JSON format)
    projects = list(projects_list.values('id', 'project_name', 'project_code'))

    context = {
        'page_obj': page_obj,
        'employees': employees,
        'projects': json.dumps(projects),
        'projects_list': projects_list,  # For filter dropdown
        'status_choices': AttendanceRecord.ATTENDANCE_STATUS,
        'filters': {
            'date_from': date_from,
            'date_to': date_to,
            'employee': employee_id,
            'status': status,
            'project': project_id,
        },
        'stats': {
            'total': total_records,
            'present': present_count,
            'late': late_count,
            'absent': absent_count,
            'overtime': total_overtime,
        },
        'per_page': per_page,
    }

    return render(request, 'attendance/attendance_list.html', context)


@login_required
def manual_attendance(request):
    """Manual attendance entry"""

    if request.method == 'POST':
        form = AttendanceRecordForm(request.POST)
        if form.is_valid():
            attendance = form.save(commit=False)
            attendance.is_manual_entry = True
            attendance.approved_by = get_current_employee(request)
            attendance.approved_at = timezone.now()

            # Get employee's schedule if not set
            if not attendance.schedule:
                attendance.schedule = get_employee_schedule(attendance.employee, attendance.date)

            attendance.save()
            messages.success(request, _('Attendance record created successfully'))
            return redirect('attendance:attendance_list')
    else:
        form = AttendanceRecordForm()

    return render(request, 'attendance/manual_attendance.html', {'form': form})


@login_required
@ensure_csrf_cookie
def mark_attendance(request):
    """Mark attendance for employees - list view with quick marking"""

    today = timezone.now().date()

    # Handle POST request for marking attendance
    if request.method == 'POST':
        # Check if it's an AJAX request (for API-style response)
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                  request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

        employee_id = request.POST.get('employee_id')
        status = request.POST.get('status')
        notes = request.POST.get('notes', '')
        custom_clock_in = request.POST.get('clock_in', '').strip()
        custom_clock_out = request.POST.get('clock_out', '').strip()
        project_id = request.POST.get('project_id', '').strip()
        custom_date = request.POST.get('date', '').strip()

        # Use custom date if provided, otherwise use today
        attendance_date = today
        if custom_date:
            try:
                attendance_date = datetime.strptime(custom_date, '%Y-%m-%d').date()
            except ValueError:
                pass  # Invalid date format, will use today

        try:
            employee = Employee.objects.get(id=employee_id)

            # Get employee's schedule
            schedule = get_employee_schedule(employee, attendance_date)

            # Set clock times based on custom input or defaults
            now = timezone.now()
            clock_in_time = None
            clock_out_time = None

            # Parse custom times if provided
            if custom_clock_in:
                try:
                    clock_in_parsed = datetime.strptime(custom_clock_in, '%H:%M').time()
                    clock_in_time = datetime.combine(attendance_date, clock_in_parsed)
                    clock_in_time = timezone.make_aware(clock_in_time)
                except ValueError:
                    pass  # Invalid time format, will use defaults

            if custom_clock_out:
                try:
                    clock_out_parsed = datetime.strptime(custom_clock_out, '%H:%M').time()
                    clock_out_time = datetime.combine(attendance_date, clock_out_parsed)
                    clock_out_time = timezone.make_aware(clock_out_time)
                except ValueError:
                    pass  # Invalid time format, will use defaults

            # Use defaults if custom times not provided or invalid
            if not clock_in_time and status in ['present', 'late', 'half_day', 'early_leave']:
                # For present/late/half_day, set clock in time
                if schedule:
                    day_schedule = schedule.get_day_schedule(attendance_date.weekday())
                    if day_schedule[0]:  # Has scheduled start time
                        # Set clock in to scheduled start time
                        clock_in_time = datetime.combine(attendance_date, day_schedule[0])
                        clock_in_time = timezone.make_aware(clock_in_time)

                        # For late status, add some minutes to make it late
                        if status == 'late':
                            clock_in_time = clock_in_time + timedelta(minutes=30)

                else:
                    # No schedule, use default times
                    clock_in_time = datetime.combine(attendance_date, time(8, 0))
                    clock_in_time = timezone.make_aware(clock_in_time)

            # Set clock out defaults if not provided via custom input
            if not clock_out_time and status in ['present', 'late', 'half_day', 'early_leave']:
                if schedule:
                    day_schedule = schedule.get_day_schedule(attendance_date.weekday())
                    # Set clock out time for completed days
                    if day_schedule[1]:  # Has scheduled end time
                        clock_out_time = datetime.combine(attendance_date, day_schedule[1])
                        clock_out_time = timezone.make_aware(clock_out_time)

                        # For early leave, subtract some minutes
                        if status == 'early_leave':
                            clock_out_time = clock_out_time - timedelta(minutes=30)
                        # For half day, set to midday
                        elif status == 'half_day':
                            clock_out_time = datetime.combine(attendance_date, time(12, 0))
                            clock_out_time = timezone.make_aware(clock_out_time)
                else:
                    # No schedule, use default times
                    if status == 'half_day':
                        clock_out_time = datetime.combine(attendance_date, time(12, 0))
                    else:
                        clock_out_time = datetime.combine(attendance_date, time(17, 0))
                    clock_out_time = timezone.make_aware(clock_out_time)

            # Get project if provided
            current_project = None
            if project_id:
                try:
                    from project.models import Project
                    current_project = Project.objects.get(id=project_id)
                except Exception:
                    pass

            attendance, created = AttendanceRecord.objects.get_or_create(
                employee=employee,
                date=attendance_date,
                defaults={
                    'schedule': schedule,
                    'status': status,
                    'clock_in': clock_in_time,
                    'clock_out': clock_out_time,
                    'current_project': current_project,
                    'is_manual_entry': True,
                    'manual_entry_reason': notes or 'Marked via attendance dashboard',
                    'approved_by': get_current_employee(request),
                    'approved_at': timezone.now()
                }
            )

            if not created:
                # Update existing record
                attendance.status = status
                # Always update clock times (even if None to clear them)
                attendance.clock_in = clock_in_time
                attendance.clock_out = clock_out_time
                # Always update project (even if None to clear it)
                attendance.current_project = current_project
                # Always update notes - save what user entered (even if empty)
                attendance.manual_entry_reason = notes or ''
                attendance.approved_by = get_current_employee(request)
                attendance.approved_at = timezone.now()
                attendance.save()

            message = f'Attendance marked as {status} for {employee.full_name}'

            if is_ajax:
                return JsonResponse({
                    'success': True,
                    'message': message,
                    'clock_in': attendance.clock_in.strftime('%H:%M:%S') if attendance.clock_in else None,
                    'clock_out': attendance.clock_out.strftime('%H:%M:%S') if attendance.clock_out else None,
                })
            else:
                messages.success(request, message)
                return redirect('attendance:mark_attendance')

        except Employee.DoesNotExist:
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'error': 'Employee not found'
                }, status=404)
            else:
                messages.error(request, 'Employee not found')
                return redirect('attendance:mark_attendance')

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()

            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'error': str(e)
                }, status=500)
            else:
                messages.error(request, f'Error: {str(e)}')
                return redirect('attendance:mark_attendance')

    # Get all active employees
    employees = Employee.objects.filter(employment_status='active').select_related('department', 'position').order_by('first_name', 'last_name')

    # Get today's attendance records
    today_attendance = AttendanceRecord.objects.filter(date=today).select_related('employee', 'current_project')
    attendance_dict = {att.employee_id: att for att in today_attendance}

    # Combine employee data with attendance status
    employee_attendance_list = []
    for employee in employees:
        attendance = attendance_dict.get(employee.id)
        employee_attendance_list.append({
            'employee': employee,
            'attendance': attendance,
            'has_attendance': attendance is not None,
            'status': attendance.status if attendance else None,
            'clock_in': attendance.clock_in if attendance else None,
            'clock_out': attendance.clock_out if attendance else None,
            'project': attendance.current_project if attendance else None,
        })

    # Filter by search query
    search_query = request.GET.get('search', '')
    if search_query:
        employee_attendance_list = [
            item for item in employee_attendance_list
            if search_query.lower() in item['employee'].full_name.lower() or
               search_query.lower() in (item['employee'].employee_id or '').lower()
        ]

    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter:
        if status_filter == 'unmarked':
            employee_attendance_list = [item for item in employee_attendance_list if not item['has_attendance']]
        else:
            employee_attendance_list = [item for item in employee_attendance_list if item['status'] == status_filter]

    # Filter by department
    department_filter = request.GET.get('department', '')
    if department_filter:
        employee_attendance_list = [
            item for item in employee_attendance_list
            if item['employee'].department_id == int(department_filter)
        ]

    # Pagination
    per_page = request.GET.get('per_page', 25)
    try:
        per_page = int(per_page)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(employee_attendance_list, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get departments for filter
    from hr.models import Department
    departments = Department.objects.all().order_by('name')

    # Get projects for optional assignment
    try:
        from project.models import Project
        from django.db import connection

        # Get active projects (not completed or cancelled)
        projects = list(Project.objects.exclude(
            status__in=['completed', 'cancelled']
        ).order_by('project_name'))

    except Exception as e:
        projects = []

    # Statistics
    total_employees = len(employee_attendance_list)
    marked_count = sum(1 for item in employee_attendance_list if item['has_attendance'])
    unmarked_count = total_employees - marked_count
    present_count = sum(1 for item in employee_attendance_list if item['status'] == 'present')
    absent_count = sum(1 for item in employee_attendance_list if item['status'] == 'absent')
    late_count = sum(1 for item in employee_attendance_list if item['status'] == 'late')

    context = {
        'page_obj': page_obj,
        'today': today,
        'search_query': search_query,
        'status_filter': status_filter,
        'department_filter': department_filter,
        'departments': departments,
        'projects': projects,
        'status_choices': AttendanceRecord.ATTENDANCE_STATUS,
        'stats': {
            'total': total_employees,
            'marked': marked_count,
            'unmarked': unmarked_count,
            'present': present_count,
            'absent': absent_count,
            'late': late_count,
        },
        'per_page': per_page,
    }

    return render(request, 'attendance/mark_attendance.html', context)


@login_required
def delete_attendance(request, employee_id):
    """Delete attendance record for today"""
    if request.method != 'POST':
        return redirect('attendance:mark_attendance')

    today = timezone.now().date()

    try:
        employee = Employee.objects.get(id=employee_id)
        attendance = AttendanceRecord.objects.filter(
            employee=employee,
            date=today
        ).first()

        if attendance:
            employee_name = employee.full_name
            attendance.delete()
            messages.success(request, f'Attendance record deleted for {employee_name}')
        else:
            messages.warning(request, 'No attendance record found for today')

    except Employee.DoesNotExist:
        messages.error(request, 'Employee not found')
    except Exception as e:
        messages.error(request, f'Error deleting attendance: {str(e)}')

    return redirect('attendance:mark_attendance')


@login_required
def delete_attendance_record(request, record_id):
    """Delete a specific attendance record by ID"""
    if request.method != 'POST':
        return redirect('attendance:attendance_list')

    try:
        attendance = AttendanceRecord.objects.get(id=record_id)
        employee_name = attendance.employee.full_name
        attendance_date = attendance.date
        attendance.delete()
        messages.success(request, f'Attendance record deleted for {employee_name} on {attendance_date}')

    except AttendanceRecord.DoesNotExist:
        messages.error(request, 'Attendance record not found')
    except Exception as e:
        messages.error(request, f'Error deleting attendance: {str(e)}')

    return redirect('attendance:attendance_list')


@login_required
def get_employees(request):
    """API endpoint to get all employees for autocomplete"""
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        employees = Employee.objects.filter(
            employment_status='active'
        ).select_related('department', 'position').order_by('first_name', 'last_name')

        employee_data = [
            {
                'id': emp.id,
                'full_name': emp.full_name,
                'employee_id': emp.employee_id or '',
                'department': emp.department.name if emp.department else '',
                'position': emp.position.name if emp.position else ''
            }
            for emp in employees
        ]

        return JsonResponse({'employees': employee_data})

    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def schedule_list(request):
    """List work schedules"""
    schedules = WorkSchedule.objects.filter(is_active=True).order_by('name')
    return render(request, 'attendance/schedule_list.html', {'schedules': schedules})


@login_required
def schedule_create(request):
    """Create work schedule"""
    if request.method == 'POST':
        form = WorkScheduleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, _('Schedule created successfully'))
            return redirect('attendance:schedule_list')
    else:
        form = WorkScheduleForm()

    return render(request, 'attendance/schedule_form.html', {'form': form, 'title': _('Create Schedule')})


@login_required
def schedule_edit(request, pk):
    """Edit work schedule"""
    schedule = get_object_or_404(WorkSchedule, pk=pk)

    if request.method == 'POST':
        form = WorkScheduleForm(request.POST, instance=schedule)
        if form.is_valid():
            form.save()
            messages.success(request, _('Schedule updated successfully'))
            return redirect('attendance:schedule_list')
    else:
        form = WorkScheduleForm(instance=schedule)

    return render(request, 'attendance/schedule_form.html', {'form': form, 'title': _('Edit Schedule')})


@login_required
def employee_schedule_assign(request):
    """Assign schedule to employee"""
    if request.method == 'POST':
        form = EmployeeScheduleForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, _('Schedule assigned successfully'))
            return redirect('attendance:schedule_list')
    else:
        form = EmployeeScheduleForm()

    return render(request, 'attendance/employee_schedule_form.html', {'form': form})


@login_required
def overtime_request_list(request):
    """List overtime requests"""
    from django.db.models import Sum, Count
    from datetime import datetime

    # Base queryset for statistics (unfiltered)
    base_queryset = OvertimeRequest.objects.select_related('employee', 'approved_by')

    # Calculate statistics from base queryset (not affected by filters)
    current_month = datetime.now().month
    current_year = datetime.now().year

    total_requests = base_queryset.count()
    pending_requests = base_queryset.filter(status='pending').count()
    approved_requests = base_queryset.filter(status='approved').count()
    rejected_requests = base_queryset.filter(status='rejected').count()

    # Approved hours for current month only
    approved_stats = base_queryset.filter(
        status='approved',
        date__month=current_month,
        date__year=current_year
    ).aggregate(
        total_hours=Sum('hours'),
        total_amount=Sum('amount')
    )

    # Filtered queryset for table display
    requests_list = base_queryset.order_by('-date', '-created_at')

    # Apply filters
    status = request.GET.get('status')
    employee_id = request.GET.get('employee')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if status:
        requests_list = requests_list.filter(status=status)
    if employee_id:
        requests_list = requests_list.filter(employee_id=employee_id)
    if date_from:
        requests_list = requests_list.filter(date__gte=date_from)
    if date_to:
        requests_list = requests_list.filter(date__lte=date_to)

    # Pagination
    per_page = request.GET.get('per_page', 25)
    try:
        per_page = int(per_page)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(requests_list, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get employees for filter dropdown
    employees = Employee.objects.filter(employment_status='active').order_by('first_name', 'last_name')

    context = {
        'page_obj': page_obj,
        'requests': page_obj,  # For template compatibility
        'employees': employees,
        'filters': {
            'status': status,
            'employee': employee_id,
            'date_from': date_from,
            'date_to': date_to,
        },
        'stats': {
            'total': total_requests,
            'pending_requests': pending_requests,
            'approved_requests': approved_requests,
            'rejected_requests': rejected_requests,
            'approved_hours': approved_stats['total_hours'] or 0,
            'approved_amount': approved_stats['total_amount'] or 0,
        },
        'per_page': per_page,
    }

    return render(request, 'attendance/overtime_list.html', context)


@login_required
def overtime_request_create(request):
    """Create overtime request"""
    if request.method == 'POST':
        form = OvertimeRequestForm(request.POST)
        if form.is_valid():
            overtime = form.save(commit=False)
            overtime.requested_by = get_current_employee(request)

            # Check if saving as draft or submitting
            action = request.POST.get('action', 'submit')
            if action == 'save_draft':
                overtime.status = 'draft'
                messages.success(request, _('Overtime request saved as draft'))
            else:
                overtime.status = 'pending'
                messages.success(request, _('Overtime request submitted successfully'))

            overtime.save()
            return redirect('attendance:overtime_list')
    else:
        form = OvertimeRequestForm()

    return render(request, 'attendance/overtime_form.html', {'form': form})


@login_required
def overtime_request_update(request, pk):
    """Update overtime request"""
    overtime = get_object_or_404(OvertimeRequest, pk=pk)

    # Only allow editing draft or rejected requests
    if overtime.status not in ['draft', 'rejected']:
        messages.error(request, _('Cannot edit overtime request with status: {}').format(overtime.get_status_display()))
        return redirect('attendance:overtime_list')

    if request.method == 'POST':
        form = OvertimeRequestForm(request.POST, instance=overtime)
        if form.is_valid():
            overtime = form.save(commit=False)

            # Check if saving as draft or submitting
            action = request.POST.get('action', 'submit')
            if action == 'save_draft':
                overtime.status = 'draft'
                messages.success(request, _('Overtime request updated and saved as draft'))
            else:
                overtime.status = 'pending'
                messages.success(request, _('Overtime request updated and submitted successfully'))

            overtime.save()
            return redirect('attendance:overtime_list')
    else:
        form = OvertimeRequestForm(instance=overtime)

    return render(request, 'attendance/overtime_form.html', {
        'form': form,
        'overtime': overtime,
        'is_update': True
    })


@login_required
def overtime_request_approve(request, pk):
    """Approve/reject overtime request"""
    overtime = get_object_or_404(OvertimeRequest, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'approve':
            overtime.status = 'approved'
            overtime.approved_by = get_current_employee(request)
            overtime.approved_at = timezone.now()
            messages.success(request, _('Overtime request approved'))
        elif action == 'reject':
            overtime.status = 'rejected'
            overtime.rejection_reason = request.POST.get('rejection_reason', '')
            messages.success(request, _('Overtime request rejected'))

        overtime.save()
        return redirect('attendance:overtime_list')

    return render(request, 'attendance/overtime_approve.html', {'overtime': overtime})


@login_required
def overtime_request_delete(request, pk):
    """Delete overtime request"""
    if request.method != 'POST':
        return redirect('attendance:overtime_list')

    try:
        overtime = OvertimeRequest.objects.get(pk=pk)
        employee_name = overtime.employee.full_name
        overtime_date = overtime.date
        overtime.delete()
        messages.success(request, _(f'Overtime request deleted for {employee_name} on {overtime_date}'))

    except OvertimeRequest.DoesNotExist:
        messages.error(request, _('Overtime request not found'))
    except Exception as e:
        messages.error(request, _(f'Error deleting overtime request: {str(e)}'))

    return redirect('attendance:overtime_list')


@login_required
def correction_request_list(request):
    """List attendance correction requests"""
    from django.db.models import Count

    corrections_list = AttendanceCorrection.objects.select_related(
        'requested_by', 'attendance', 'approved_by'
    ).order_by('-created_at')

    # Filters
    status = request.GET.get('status')
    employee_id = request.GET.get('employee')
    correction_type = request.GET.get('correction_type')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if status:
        corrections_list = corrections_list.filter(status=status)
    if employee_id:
        corrections_list = corrections_list.filter(requested_by_id=employee_id)
    if correction_type:
        corrections_list = corrections_list.filter(correction_type=correction_type)
    if date_from:
        corrections_list = corrections_list.filter(attendance__date__gte=date_from)
    if date_to:
        corrections_list = corrections_list.filter(attendance__date__lte=date_to)

    # Calculate statistics
    total_requests = corrections_list.count()
    pending_requests = corrections_list.filter(status='pending').count()
    approved_requests = corrections_list.filter(status='approved').count()
    rejected_requests = corrections_list.filter(status='rejected').count()
    draft_requests = corrections_list.filter(status='draft').count()

    # Pagination
    per_page = request.GET.get('per_page', 25)
    try:
        per_page = int(per_page)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(corrections_list, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get employees for filter dropdown
    employees = Employee.objects.filter(employment_status='active').order_by('first_name', 'last_name')

    context = {
        'page_obj': page_obj,
        'corrections': page_obj,  # For template compatibility
        'employees': employees,
        'filters': {
            'status': status,
            'employee': employee_id,
            'correction_type': correction_type,
            'date_from': date_from,
            'date_to': date_to,
        },
        'stats': {
            'total': total_requests,
            'pending_requests': pending_requests,
            'approved_requests': approved_requests,
            'rejected_requests': rejected_requests,
            'draft_requests': draft_requests,
        },
        'per_page': per_page,
    }

    return render(request, 'attendance/correction_list.html', context)


@login_required
def correction_request_create(request):
    """Create attendance correction request"""
    if request.method == 'POST':
        form = AttendanceCorrectionForm(request.POST, request.FILES)
        if form.is_valid():
            correction = form.save(commit=False)
            correction.requested_by = get_current_employee(request)
            correction.save()
            messages.success(request, _('Correction request submitted successfully'))
            return redirect('attendance:correction_list')
    else:
        form = AttendanceCorrectionForm()

    return render(request, 'attendance/correction_form.html', {'form': form})


@login_required
def correction_request_approve(request, pk):
    """Approve/reject correction request"""
    correction = get_object_or_404(AttendanceCorrection, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'approve':
            # Apply the correction
            attendance = correction.attendance

            if correction.correction_type in ['clock_in', 'both'] and correction.new_clock_in:
                attendance.clock_in = correction.new_clock_in

            if correction.correction_type in ['clock_out', 'both'] and correction.new_clock_out:
                attendance.clock_out = correction.new_clock_out

            if correction.correction_type == 'status' and correction.new_status:
                attendance.status = correction.new_status

            attendance.save()

            # Update correction status
            correction.status = 'approved'
            correction.approved_by = get_current_employee(request)
            correction.approved_at = timezone.now()
            messages.success(request, _('Correction approved and applied'))

        elif action == 'reject':
            correction.status = 'rejected'
            correction.rejection_reason = request.POST.get('rejection_reason', '')
            messages.success(request, _('Correction request rejected'))

        correction.save()
        return redirect('attendance:correction_list')

    return render(request, 'attendance/correction_approve.html', {'correction': correction})


@login_required
def biometric_device_list(request):
    """List biometric devices"""
    devices = BiometricDevice.objects.all().order_by('location', 'name')
    return render(request, 'attendance/device_list.html', {'devices': devices})


@login_required
def biometric_device_create(request):
    """Add new biometric device"""
    if request.method == 'POST':
        form = BiometricDeviceForm(request.POST)
        if form.is_valid():
            device = form.save()

            # Register device with biometric service
            service = get_biometric_service()
            service.add_device(
                device.device_id,
                device.ip_address,
                device.port,
                device.device_type
            )

            messages.success(request, _('Device added successfully'))
            return redirect('attendance:device_list')
    else:
        form = BiometricDeviceForm()

    return render(request, 'attendance/device_form.html', {'form': form})


@login_required
def biometric_device_sync(request, pk):
    """Sync attendance from biometric device"""
    device = get_object_or_404(BiometricDevice, pk=pk)

    try:
        service = get_biometric_service()
        logs = service.sync_attendance(device.device_id)

        # Process attendance logs
        processed = 0
        for log in logs:
            try:
                employee = Employee.objects.get(employee_id=log['user_id'])
                attendance_date = log['timestamp'].date()

                attendance, created = AttendanceRecord.objects.get_or_create(
                    employee=employee,
                    date=attendance_date,
                    defaults={'schedule': get_employee_schedule(employee, attendance_date)}
                )

                if log['type'] == 'clock_in':
                    attendance.clock_in = log['timestamp']
                    attendance.clock_in_device = device
                elif log['type'] == 'clock_out':
                    attendance.clock_out = log['timestamp']
                    attendance.clock_out_device = device

                attendance.save()
                processed += 1

            except Employee.DoesNotExist:
                continue

        # Update device last sync
        device.last_sync = timezone.now()
        device.save()

        messages.success(request, f'Synced {processed} attendance records from {device.name}')

    except Exception as e:
        messages.error(request, f'Failed to sync device: {str(e)}')

    return redirect('attendance:device_list')


@login_required
def attendance_report(request):
    """Generate attendance reports with filters and table display"""
    today = timezone.now().date()

    # Get filter parameters from GET request
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    employee_id = request.GET.get('employee', '')
    status = request.GET.get('status', '')
    department_id = request.GET.get('department', '')
    search_query = request.GET.get('search', '')

    # Set default date range if not provided (current month)
    if not date_from:
        date_from = today.replace(day=1)
    else:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()

    if not date_to:
        date_to = today
    else:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Query attendance records with filters
    records = AttendanceRecord.objects.filter(
        date__gte=date_from,
        date__lte=date_to
    ).select_related('employee', 'employee__department', 'employee__position', 'schedule', 'current_project').order_by('-date', 'employee__first_name')

    # Apply filters
    if employee_id:
        records = records.filter(employee_id=employee_id)

    if status:
        records = records.filter(status=status)

    if department_id:
        records = records.filter(employee__department_id=department_id)

    if search_query:
        records = records.filter(
            Q(employee__first_name__icontains=search_query) |
            Q(employee__last_name__icontains=search_query) |
            Q(employee__employee_id__icontains=search_query)
        )

    # Calculate statistics
    total_records = records.count()
    present_count = records.filter(status='present').count()
    late_count = records.filter(status='late').count()
    absent_count = records.filter(status='absent').count()
    leave_count = records.filter(status='leave').count()
    total_overtime = records.aggregate(total=Sum('overtime_hours'))['total'] or 0

    # Get today's stats for cards
    today_records = AttendanceRecord.objects.filter(date=today)
    present_today = today_records.filter(status__in=['present', 'late', 'early_leave']).count()
    late_today = today_records.filter(status='late').count()
    absent_today = today_records.filter(status='absent').count()
    total_employees = Employee.objects.filter(employment_status='active').count()

    # Pagination
    per_page = request.GET.get('per_page', 25)
    try:
        per_page = int(per_page)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get employees and departments for filter dropdowns
    employees = Employee.objects.filter(employment_status='active').order_by('first_name', 'last_name')
    from hr.models import Department
    departments = Department.objects.all().order_by('name')

    context = {
        'page_obj': page_obj,
        'employees': employees,
        'departments': departments,
        'status_choices': AttendanceRecord.ATTENDANCE_STATUS,
        'filters': {
            'date_from': date_from.strftime('%Y-%m-%d'),
            'date_to': date_to.strftime('%Y-%m-%d'),
            'employee': employee_id,
            'status': status,
            'department': department_id,
            'search': search_query,
        },
        'stats': {
            'total': total_records,
            'present': present_count,
            'late': late_count,
            'absent': absent_count,
            'leave': leave_count,
            'overtime': total_overtime,
            'total_employees': total_employees,
            'present_today': present_today,
            'late_today': late_today,
            'absent_today': absent_today,
        },
        'per_page': per_page,
    }

    return render(request, 'attendance/reports.html', context)


@login_required
def attendance_report_export(request):
    """Export attendance report in various formats"""
    # Get filter parameters from GET request
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    employee_id = request.GET.get('employee', '')
    status = request.GET.get('status', '')
    department_id = request.GET.get('department', '')
    search_query = request.GET.get('search', '')
    export_format = request.GET.get('export_format', 'csv')

    today = timezone.now().date()

    # Set default date range if not provided (current month)
    if not date_from:
        date_from = today.replace(day=1)
    else:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()

    if not date_to:
        date_to = today
    else:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Query attendance records with filters
    records = AttendanceRecord.objects.filter(
        date__gte=date_from,
        date__lte=date_to
    ).select_related('employee', 'employee__department', 'employee__position', 'schedule', 'current_project').order_by('-date', 'employee__first_name')

    # Apply filters
    if employee_id:
        records = records.filter(employee_id=employee_id)

    if status:
        records = records.filter(status=status)

    if department_id:
        records = records.filter(employee__department_id=department_id)

    if search_query:
        records = records.filter(
            Q(employee__first_name__icontains=search_query) |
            Q(employee__last_name__icontains=search_query) |
            Q(employee__employee_id__icontains=search_query)
        )

    # Generate report based on format
    if export_format == 'csv':
        return export_attendance_csv(records, date_from, date_to)
    elif export_format == 'xlsx':
        return export_attendance_excel(records, date_from, date_to)
    else:  # PDF
        return export_attendance_pdf(records, date_from, date_to)


def export_attendance_csv(records, start_date, end_date):
    """Export attendance records as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{start_date}_{end_date}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Date', 'Employee ID', 'Employee Name', 'Department', 'Position',
        'Clock In', 'Clock Out', 'Total Hours', 'Overtime Hours',
        'Status', 'Late Minutes', 'Early Leave Minutes', 'Project'
    ])

    for record in records:
        writer.writerow([
            record.date,
            record.employee.employee_id or '',
            record.employee.full_name,
            record.employee.department.name if record.employee.department else '',
            record.employee.position.name if record.employee.position else '',
            record.clock_in.strftime('%H:%M:%S') if record.clock_in else '',
            record.clock_out.strftime('%H:%M:%S') if record.clock_out else '',
            record.total_hours or 0,
            record.overtime_hours or 0,
            record.get_status_display(),
            record.late_minutes or 0,
            record.early_leave_minutes or 0,
            record.current_project.project_name if record.current_project else '',
        ])

    return response


def export_attendance_excel(records, start_date, end_date):
    """Export attendance records as Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        # Headers
        headers = [
            'Date', 'Employee ID', 'Employee Name', 'Department', 'Position',
            'Clock In', 'Clock Out', 'Total Hours', 'Overtime Hours',
            'Status', 'Late Minutes', 'Early Leave Minutes', 'Project'
        ]

        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Data rows
        for row_num, record in enumerate(records, 2):
            ws.cell(row=row_num, column=1, value=record.date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=2, value=record.employee.employee_id or '')
            ws.cell(row=row_num, column=3, value=record.employee.full_name)
            ws.cell(row=row_num, column=4, value=record.employee.department.name if record.employee.department else '')
            ws.cell(row=row_num, column=5, value=record.employee.position.name if record.employee.position else '')
            ws.cell(row=row_num, column=6, value=record.clock_in.strftime('%H:%M:%S') if record.clock_in else '')
            ws.cell(row=row_num, column=7, value=record.clock_out.strftime('%H:%M:%S') if record.clock_out else '')
            ws.cell(row=row_num, column=8, value=record.total_hours or 0)
            ws.cell(row=row_num, column=9, value=record.overtime_hours or 0)
            ws.cell(row=row_num, column=10, value=record.get_status_display())
            ws.cell(row=row_num, column=11, value=record.late_minutes or 0)
            ws.cell(row=row_num, column=12, value=record.early_leave_minutes or 0)
            ws.cell(row=row_num, column=13, value=record.current_project.project_name if record.current_project else '')

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="attendance_report_{start_date}_{end_date}.xlsx"'

        wb.save(response)
        return response

    except ImportError:
        # Fallback to CSV if openpyxl not available
        return export_attendance_csv(records, start_date, end_date)


def export_attendance_pdf(records, start_date, end_date):
    """Export attendance records as PDF"""
    # Implementation would require reportlab or weasyprint
    # For now, return CSV as fallback
    return export_attendance_csv(records, start_date, end_date)


@login_required
def overtime_report(request):
    """Generate overtime reports with filters and table display"""
    today = timezone.now().date()

    # Get filter parameters from GET request
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    employee_id = request.GET.get('employee', '')
    status = request.GET.get('status', '')
    department_id = request.GET.get('department', '')
    project_id = request.GET.get('project', '')
    search_query = request.GET.get('search', '')

    # Set default date range if not provided (current month)
    if not date_from:
        date_from = today.replace(day=1)
    else:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()

    if not date_to:
        date_to = today
    else:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Query overtime requests with filters
    records = OvertimeRequest.objects.filter(
        date__gte=date_from,
        date__lte=date_to
    ).select_related('employee', 'employee__department', 'employee__position', 'project', 'approved_by')

    # Apply filters
    if employee_id:
        records = records.filter(employee_id=employee_id)

    if status:
        records = records.filter(status=status)

    if department_id:
        records = records.filter(employee__department_id=department_id)

    if project_id:
        records = records.filter(project_id=project_id)

    if search_query:
        records = records.filter(
            Q(employee__first_name__icontains=search_query) |
            Q(employee__last_name__icontains=search_query) |
            Q(employee__employee_id__icontains=search_query)
        )

    # Sorting
    sort_by = request.GET.get('sort', 'date')
    order = request.GET.get('order', 'desc')

    valid_sort_fields = ['date', 'employee__first_name', 'employee__last_name', 'hours', 'status', 'created_at']

    if sort_by in valid_sort_fields:
        if order == 'desc':
            sort_by = f'-{sort_by}'
        records = records.order_by(sort_by)
    else:
        records = records.order_by('-date')

    # Calculate statistics
    total_records = records.count()
    pending_count = records.filter(status='pending').count()
    approved_count = records.filter(status='approved').count()
    rejected_count = records.filter(status='rejected').count()
    cancelled_count = records.filter(status='cancelled').count()

    # Calculate total hours and amount
    total_hours = records.filter(status='approved').aggregate(total=Sum('hours'))['total'] or 0
    total_amount = records.filter(status='approved').aggregate(total=Sum('amount'))['total'] or 0

    # Pagination
    per_page = request.GET.get('per_page', 25)
    try:
        per_page = int(per_page)
    except (ValueError, TypeError):
        per_page = 25

    paginator = Paginator(records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get employees, departments and projects for filter dropdowns
    employees = Employee.objects.filter(employment_status='active').order_by('first_name', 'last_name')
    from hr.models import Department
    departments = Department.objects.all().order_by('name')
    from project.models import Project
    projects = Project.objects.all().order_by('project_name')

    context = {
        'page_obj': page_obj,
        'employees': employees,
        'departments': departments,
        'projects': projects,
        'status_choices': OvertimeRequest.STATUS_CHOICES,
        'filters': {
            'date_from': date_from.strftime('%Y-%m-%d'),
            'date_to': date_to.strftime('%Y-%m-%d'),
            'employee': employee_id,
            'status': status,
            'department': department_id,
            'project': project_id,
            'search': search_query,
        },
        'stats': {
            'total': total_records,
            'pending': pending_count,
            'approved': approved_count,
            'rejected': rejected_count,
            'cancelled': cancelled_count,
            'total_hours': total_hours,
            'total_amount': total_amount,
        },
        'per_page': per_page,
        'sort': request.GET.get('sort', 'date'),
        'order': request.GET.get('order', 'desc'),
    }

    return render(request, 'attendance/overtime_report.html', context)


@login_required
def overtime_report_export(request):
    """Export overtime report in various formats"""
    # Get filter parameters from GET request
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    employee_id = request.GET.get('employee', '')
    status = request.GET.get('status', '')
    department_id = request.GET.get('department', '')
    project_id = request.GET.get('project', '')
    search_query = request.GET.get('search', '')
    export_format = request.GET.get('export_format', 'csv')

    today = timezone.now().date()

    # Set default date range if not provided (current month)
    if not date_from:
        date_from = today.replace(day=1)
    else:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()

    if not date_to:
        date_to = today
    else:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()

    # Query overtime requests with filters
    records = OvertimeRequest.objects.filter(
        date__gte=date_from,
        date__lte=date_to
    ).select_related('employee', 'employee__department', 'employee__position', 'project', 'approved_by').order_by('-date', 'employee__first_name')

    # Apply filters
    if employee_id:
        records = records.filter(employee_id=employee_id)

    if status:
        records = records.filter(status=status)

    if department_id:
        records = records.filter(employee__department_id=department_id)

    if project_id:
        records = records.filter(project_id=project_id)

    if search_query:
        records = records.filter(
            Q(employee__first_name__icontains=search_query) |
            Q(employee__last_name__icontains=search_query) |
            Q(employee__employee_id__icontains=search_query)
        )

    # Generate report based on format
    if export_format == 'csv':
        return export_overtime_csv(records, date_from, date_to)
    elif export_format == 'xlsx':
        return export_overtime_excel(records, date_from, date_to)
    else:  # PDF
        return export_overtime_pdf(records, date_from, date_to)


def export_overtime_csv(records, start_date, end_date):
    """Export overtime records as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="overtime_report_{start_date}_to_{end_date}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Date', 'Employee ID', 'Employee Name', 'Department', 'Project',
        'Start Time', 'End Time', 'Hours', 'Rate', 'Amount', 'Status',
        'Reason', 'Approved By', 'Approved At'
    ])

    for record in records:
        writer.writerow([
            record.date,
            record.employee.employee_id if hasattr(record.employee, 'employee_id') else '',
            record.employee.full_name,
            record.employee.department.name if record.employee.department else '',
            record.project.project_name if record.project else '',
            record.start_time,
            record.end_time,
            record.hours,
            record.overtime_rate,
            record.amount if record.amount else '',
            record.get_status_display(),
            record.reason,
            record.approved_by.full_name if record.approved_by else '',
            record.approved_at.strftime('%Y-%m-%d %H:%M') if record.approved_at else '',
        ])

    return response


def export_overtime_excel(records, start_date, end_date):
    """Export overtime records as Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Overtime Report'

        # Header styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        headers = [
            'Date', 'Employee ID', 'Employee Name', 'Department', 'Project',
            'Start Time', 'End Time', 'Hours', 'Rate', 'Amount', 'Status',
            'Reason', 'Approved By', 'Approved At'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Data rows
        for row_num, record in enumerate(records, 2):
            ws.cell(row=row_num, column=1, value=record.date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=2, value=record.employee.employee_id if hasattr(record.employee, 'employee_id') else '')
            ws.cell(row=row_num, column=3, value=record.employee.full_name)
            ws.cell(row=row_num, column=4, value=record.employee.department.name if record.employee.department else '')
            ws.cell(row=row_num, column=5, value=record.project.project_name if record.project else '')
            ws.cell(row=row_num, column=6, value=record.start_time.strftime('%H:%M'))
            ws.cell(row=row_num, column=7, value=record.end_time.strftime('%H:%M'))
            ws.cell(row=row_num, column=8, value=float(record.hours))
            ws.cell(row=row_num, column=9, value=float(record.overtime_rate))
            ws.cell(row=row_num, column=10, value=float(record.amount) if record.amount else '')
            ws.cell(row=row_num, column=11, value=record.get_status_display())
            ws.cell(row=row_num, column=12, value=record.reason)
            ws.cell(row=row_num, column=13, value=record.approved_by.full_name if record.approved_by else '')
            ws.cell(row=row_num, column=14, value=record.approved_at.strftime('%Y-%m-%d %H:%M') if record.approved_at else '')

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="overtime_report_{start_date}_to_{end_date}.xlsx"'
        wb.save(response)
        return response
    except ImportError:
        # If openpyxl is not available, fall back to CSV
        return export_overtime_csv(records, start_date, end_date)


def export_overtime_pdf(records, start_date, end_date):
    """Export overtime records as PDF"""
    # Implementation would require reportlab or weasyprint
    # For now, return CSV as fallback
    return export_overtime_csv(records, start_date, end_date)


# Helper functions
def get_employee_schedule(employee, date):
    """Get employee's active schedule for a specific date"""
    schedule = EmployeeSchedule.objects.filter(
        employee=employee,
        effective_from__lte=date,
        is_active=True
    ).filter(
        Q(effective_to__gte=date) | Q(effective_to__isnull=True)
    ).select_related('schedule').first()

    return schedule.schedule if schedule else None


# API endpoints for AJAX operations
