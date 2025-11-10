"""
Step 1: Excel extraction to JSON and Images
Based on OCR/extract_excel_to_json.py but optimized for Django integration
"""

import ast
from datetime import datetime
import os
import pandas as pd
import json
from pathlib import Path
from PIL import Image
import io
import zipfile
import re
import shutil
import openpyxl
import tempfile
import logging
import pytesseract
import cv2
import numpy as np

from django.conf import settings

logger = logging.getLogger(__name__)

def extract_basic_document_info(image_path):
    """Extract basic document information like document number from ID/Passport image."""
    try:
        # Load and preprocess image
        img = cv2.imread(image_path)
        if img is None:
            return {}
        
        # Convert to grayscale and enhance contrast
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
        enhanced = clahe.apply(gray)
        
        # Apply threshold
        _, thresh = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
        
        # Scale up for better OCR
        scale_factor = 2
        height, width = thresh.shape
        resized = cv2.resize(thresh, (width * scale_factor, height * scale_factor), interpolation=cv2.INTER_CUBIC)
        
        # Convert back to PIL Image for OCR
        pil_image = Image.fromarray(resized)
        
        # Extract text with basic OCR
        config = '--oem 3 --psm 6'
        text = pytesseract.image_to_string(pil_image, config=config)
        
        # Look for document number patterns
        document_info = {}
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip().upper()
            # Look for common document number patterns
            # KHID: Usually 9 digits
            khid_match = re.search(r'\b\d{9}\b', line)
            if khid_match:
                document_info['document_number'] = khid_match.group()
                document_info['document_type'] = 'KHID'
                break
            
            # Passport: Usually alphanumeric, 6-10 chars
            passport_match = re.search(r'\b[A-Z]\d{7,8}\b|\b[A-Z]{1,2}\d{6,8}\b', line)
            if passport_match:
                document_info['document_number'] = passport_match.group()
                document_info['document_type'] = 'PASSPORT'
                break
            
            # Look for any alphanumeric sequence that might be a document number
            general_match = re.search(r'\b[A-Z0-9]{8,12}\b', line)
            if general_match and not any(word in line for word in ['NAME', 'DATE', 'SEX', 'NATIONALITY']):
                document_info['document_number'] = general_match.group()
                document_info['document_type'] = 'UNKNOWN'
        
        return document_info

    except Exception as e:
        return {}

def clean_filename(text, max_length=50):
    """Clean text to be safe for filename."""
    if not text:
        return "unnamed"
    text = re.sub(r'[^\w\s-]', '', str(text))
    text = text.strip().replace(' ', '_')
    return text[:max_length]

def extract_images_from_excel(excel_path):
    """Extract all images from Excel file (ZIP structure)."""
    images = []
    
    try:
        # Check if the file is a valid ZIP file first
        if not zipfile.is_zipfile(excel_path):
            return images
            
        with zipfile.ZipFile(excel_path, 'r') as zip_file:
            # Find all media files (both 'media/' and 'xl/media/' patterns)
            media_files = [f for f in zip_file.filelist 
                          if ('media/' in f.filename or any(f.filename.lower().endswith(ext) for ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff'])) 
                          and not f.is_dir()]
            media_files.sort(key=lambda x: x.filename)
            
            for idx, file_info in enumerate(media_files):
                # Read image data
                image_data = zip_file.read(file_info.filename)
                
                # Get image info
                img = Image.open(io.BytesIO(image_data))
                extension = f".{img.format.lower()}" if img.format else '.png'
                
                images.append({
                    "data": image_data,
                    "width": img.width,
                    "height": img.height,
                    "format": img.format.lower() if img.format else 'png',
                    "extension": extension,
                    "index": idx
                })

    except Exception as e:
        pass
    
    return images

def detect_image_positions(excel_path, num_images=0):
    """Detect which cells contain images by checking for #VALUE! or formula errors.
    
    Args:
        excel_path: Path to Excel file
        num_images: Number of images found in the Excel ZIP (for fallback logic)
    """
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
    except Exception as e:
        raise ValueError(f"Invalid Excel file: {e}")
    
    try:
        ws = wb['Workers'] if 'Workers' in wb.sheetnames else wb.active
    except Exception as e:
        ws = wb.active
    
    # Get column indices
    cols = {}
    for cell in ws[1]:
        if cell.value:
            cols[cell.value] = cell.column
    
    # Track image positions
    image_positions = []
    
    # Get all potential image columns dynamically from the worksheet
    image_columns_candidates = ['PHOTO', 'IDKH/PASSPORT', 'WORKINGPERMIT', 'VISA', 'ID', 'PASSPORT', 'DOCUMENT']
    actual_image_columns = []
    
    for col_name in image_columns_candidates:
        if col_name in cols:
            actual_image_columns.append((col_name, cols[col_name]))
    
    # Sort by column number to maintain Excel order
    actual_image_columns.sort(key=lambda x: x[1])
    
    # Check each row for image placeholders
    for row_idx in range(2, ws.max_row + 1):
        row_images = []
        
        # Check all potential image columns
        for col_name, col_idx in actual_image_columns:
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Check if cell has #VALUE! or similar (indicates image)
            # Also check for other common image indicators
            if cell.value:
                cell_str = str(cell.value).strip()
                # Check various image indicators
                is_image_indicator = (
                    cell_str == '#VALUE!' or 
                    '=IMAGE' in cell_str.upper() or
                    cell_str == '=IMAGE()' or
                    cell_str.startswith('=EMBED') or
                    cell_str.startswith('=_xlfn.IMAGE') or
                    # Sometimes Excel shows as empty string when image is present
                    (cell_str == '' and col_name in ['PHOTO', 'IDKH/PASSPORT', 'WORKINGPERMIT', 'VISA'])
                )

                if is_image_indicator:
                    row_images.append((col_name, col_idx))
        
        if row_images:
            name_cell = ws.cell(row=row_idx, column=cols.get('NAME', 2))
            # Sort by column number to ensure correct order
            row_images.sort(key=lambda x: x[1])
            image_positions.append({
                'row': row_idx,
                'name': name_cell.value,
                'columns': [col[0] for col in row_images]  # Just column names in order
            })
    
    pass
    
    # FALLBACK: If no image positions detected but we know there are images in the ZIP
    # Create a fallback mapping assuming images are distributed to valid rows
    if not image_positions:
        # Count valid data rows (rows with NAME)
        valid_rows = []
        for row_idx in range(2, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=cols.get('NAME', 2))
            if name_cell.value and str(name_cell.value).strip():
                valid_rows.append(row_idx)

        # Smart fallback: Detect pattern from number of images vs rows
        # Common patterns:
        # - 2 images per row: IDKH/PASSPORT + PHOTO
        # - 3 images per row: IDKH/PASSPORT + PHOTO + VISA/WORKINGPERMIT
        # - 4 images per row: All columns

        images_per_row = num_images / len(valid_rows) if valid_rows else 0

        for row_idx in valid_rows:
            name_cell = ws.cell(row=row_idx, column=cols.get('NAME', 2))
            
            # Determine columns based on ratio
            if images_per_row >= 3.5:
                # 4 images per row - all columns
                fallback_columns = ['IDKH/PASSPORT', 'WORKINGPERMIT', 'VISA', 'PHOTO']
            elif images_per_row >= 2.5:
                # 3 images per row
                fallback_columns = ['IDKH/PASSPORT', 'WORKINGPERMIT', 'VISA']
            elif images_per_row >= 1.5:
                # 2 images per row - most common scenario
                fallback_columns = ['IDKH/PASSPORT', 'PHOTO']
            else:
                # 1 image per row - assume passport/ID
                fallback_columns = ['IDKH/PASSPORT']
            
            # Only include columns that actually exist in the Excel
            fallback_columns = [col for col in fallback_columns if col in cols]
            
            if fallback_columns:
                image_positions.append({
                    'row': row_idx,
                    'name': name_cell.value,
                    'columns': fallback_columns
                })
    
    wb.close()
    return image_positions

def extract_excel_to_json_and_images(excel_file_path, progress_tracker=None):
    """
    Step 1: Extract Excel to JSON + Images (NO MRZ processing yet)
    Returns: (json_path, images_dir, record_count, image_count)
    """
    media_root = getattr(settings, 'MEDIA_ROOT', '')
    
    # Create timestamped filenames
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    json_filename = f"worker_eforms_{timestamp}.json"
    
    # Create directories
    json_dir = os.path.join(media_root, 'worker_import_json')
    images_dir = os.path.join(media_root, 'worker_import_photos')
    
    os.makedirs(json_dir, exist_ok=True)
    os.makedirs(images_dir, exist_ok=True)
    
    json_path = os.path.join(json_dir, json_filename)
    
    # Update progress
    if progress_tracker:
        progress_tracker.update(
            stage='extracting_excel',
            stage_message='Extracting images from Excel file...'
        )
    
    # Verify Excel file exists
    excel_path = Path(excel_file_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file '{excel_path}' not found!")
    
    # Check file extension
    if not excel_path.suffix.lower() in ['.xlsx', '.xls']:
        raise ValueError(f"Invalid file format. Expected .xlsx or .xls, got {excel_path.suffix}")
    
    # Extract images first
    images = extract_images_from_excel(excel_path)
    
    # Update progress
    if progress_tracker:
        progress_tracker.update(
            stage='detecting_positions',
            stage_message='Detecting image positions in Excel...'
        )
    
    # Detect where images should go (pass number of images for fallback logic)
    image_positions = detect_image_positions(excel_path, num_images=len(images))
    
    # FALLBACK: If we have images but no detected positions, create fallback mapping
    if len(images) > 0 and len(image_positions) == 0:
        # Read Excel to get valid rows
        try:
            df_preview = pd.read_excel(excel_path, sheet_name='Workers')
        except:
            df_preview = pd.read_excel(excel_path)
        
        # Create fallback mapping for valid rows
        for idx, row in df_preview.iterrows():
            if pd.notna(row.get('NAME')) and str(row.get('NAME')).strip():
                row_num = idx + 2  # Excel row (1-indexed + header)
                image_positions.append({
                    'row': row_num,
                    'name': row.get('NAME'),
                    'columns': ['IDKH/PASSPORT']  # Default to passport column
                })
    
    # Read Excel data
    try:
        df = pd.read_excel(excel_path, sheet_name='Workers')
    except Exception as e:
        # Fallback to default sheet if 'Workers' sheet doesn't exist
        try:
            df = pd.read_excel(excel_path)
        except Exception as e2:
            raise ValueError(f"Cannot read Excel file: {e2}")

    # Preserve original Excel column order
    original_column_order = list(df.columns)
    
    # Update progress
    if progress_tracker:
        progress_tracker.update(
            stage='processing_records',
            stage_message='Processing Excel records...'
        )
    
    # Calculate total expected images vs available images
    total_expected_positions = sum(len(pos['columns']) for pos in image_positions)
    total_available_images = len(images)
    
    # Create smart distribution plan
    if total_available_images < total_expected_positions:
        # Not enough unique images - but Excel might reuse images
        # Strategy: Distribute all available images, then reuse the last image for remaining positions
        priority_order = ['IDKH/PASSPORT', 'PHOTO', 'WORKINGPERMIT', 'VISA']
        
        # Create a flat list of all positions with priorities
        all_positions = []
        for pos in image_positions:
            for col_name in pos['columns']:
                priority = priority_order.index(col_name) if col_name in priority_order else 999
                all_positions.append({
                    'row': pos['row'],
                    'name': pos['name'],
                    'column': col_name,
                    'priority': priority
                })
        
        # Sort by priority, then by row
        all_positions.sort(key=lambda x: (x['priority'], x['row']))
        
        # Simple row-by-row mapping: assign images in Excel reading order
        # Row 1: IDKH/PASSPORT then PHOTO, Row 2: IDKH/PASSPORT then PHOTO, etc.
        distribution_plan = []
        
        # Group by row and sort within each row by column position
        position_by_row = {}
        for pos in all_positions:
            if pos['row'] not in position_by_row:
                position_by_row[pos['row']] = []
            position_by_row[pos['row']].append(pos)
        
        # Sort each row's positions by column type order (IDKH/PASSPORT first, then PHOTO)
        for row_num in position_by_row:
            position_by_row[row_num].sort(key=lambda x: priority_order.index(x['column']))
        
        # Assign images sequentially through rows
        image_index = 0
        for row_num in sorted(position_by_row.keys()):
            for pos in position_by_row[row_num]:
                if image_index < total_available_images:
                    # Use next available image
                    current_image_index = image_index
                    reused = False
                    image_index += 1
                else:
                    # Need to reuse - for photos, reuse the first photo image
                    if pos['column'] == 'PHOTO':
                        # Find the first PHOTO assignment and reuse that image
                        first_photo = next((p for p in distribution_plan if p['column'] == 'PHOTO'), None)
                        if first_photo:
                            current_image_index = first_photo['image_index']
                        else:
                            current_image_index = 1  # Assume image index 1 is a photo
                    else:
                        # For ID documents, this shouldn't happen with our priority system
                        current_image_index = image_index - 1
                    reused = True
                
                distribution_plan.append({
                    **pos,
                    'image_index': current_image_index,
                    'reused': reused
                })
    else:
        # Simple case - enough images for all positions
        distribution_plan = None
    
    # Process records (NO MRZ processing yet)
    records = []
    image_index = 0
    
    for idx, row in df.iterrows():
        # Skip rows with empty NAME field
        if pd.isna(row.get('NAME')) or row.get('NAME') == '' or str(row.get('NAME')).strip() == '':
            continue
            
        # Check if row has valid data (at least NO field)
        if pd.notna(row.get('NO')):
            record = {}
            
            # Convert all fields
            for col in df.columns:
                value = row[col]
                
                if pd.isna(value):
                    record[col] = None
                elif isinstance(value, pd.Timestamp):
                    record[col] = value.strftime('%Y-%m-%d')
                elif hasattr(value, 'item'):  # numpy type
                    record[col] = value.item()
                # Skip #VALUE! or =IMAGE() placeholders
                elif str(value) in ['#VALUE!', '=IMAGE()']:
                    record[col] = None
                else:
                    record[col] = value
            
            # Find image mapping for this row
            row_num = idx + 2  # Excel row number (1-indexed + header)
            row_image_info = next((p for p in image_positions if p['row'] == row_num), None)
            
            if row_image_info:
                worker_id = clean_filename(record.get('NAME')) if record.get('NAME') else f"worker_{int(record.get('NO', idx + 1))}"
                
                # Use smart distribution plan if available
                if distribution_plan:
                    # Get assignments for this worker from the distribution plan
                    worker_assignments = [plan for plan in distribution_plan if plan['row'] == row_num]
                    
                    # Process each assignment
                    for assignment in worker_assignments:
                        col_name = assignment['column']
                        img_idx = assignment['image_index']
                        is_reused = assignment['reused']
                        
                        if img_idx < len(images):
                            img = images[img_idx]
                            
                            # Create filename - include reuse indicator if needed
                            base_name = f"{worker_id}_{col_name.lower().replace('/', '_')}"
                            if is_reused:
                                base_name += "_reused"
                            filename = f"{base_name}{img['extension']}"
                            
                            # Save image (even if reused, save with different filename)
                            filepath = os.path.join(images_dir, filename)
                            try:
                                with open(filepath, 'wb') as f:
                                    f.write(img.get('data', b''))
                                
                                # Extract basic document info for ID/PASSPORT images
                                basic_doc_info = {}
                                if col_name == 'IDKH/PASSPORT':
                                    basic_doc_info = extract_basic_document_info(filepath)
                                
                                # Add image info to the correct column
                                record[col_name] = {
                                    "filename": filename,
                                    "path": filepath.replace('\\', '/'),
                                    "width": img.get('width', 0),
                                    "height": img.get('height', 0),
                                    "format": img.get('format', 'jpg'),
                                    "mrz_processed": False,
                                    "source_image_index": img_idx,
                                    "reused": is_reused,
                                    **basic_doc_info  # Add document number and type if found
                                }
                            except Exception as e:
                                pass
                else:
                    # Simple case - sequential assignment
                    actual_columns = row_image_info['columns']
                    
                    # Distribute images to the determined columns
                    for col_idx, col_name in enumerate(actual_columns):
                        if image_index < len(images):
                            img = images[image_index]
                            
                            # Create filename using NAME only (add number if multiple images for same person)
                            if len(actual_columns) > 1:
                                # Multiple images for this person, add index
                                filename = f"{worker_id}_{col_idx + 1}{img['extension']}"
                            else:
                                # Single image for this person
                                filename = f"{worker_id}{img['extension']}"
                            
                            # Save image
                            filepath = os.path.join(images_dir, filename)
                            try:
                                with open(filepath, 'wb') as f:
                                    f.write(img.get('data', b''))
                                
                                # Extract basic document info for ID/PASSPORT images
                                basic_doc_info = {}
                                if col_name == 'IDKH/PASSPORT':
                                    basic_doc_info = extract_basic_document_info(filepath)
                                
                                # Add image info to the correct column (NO MRZ yet)
                                record[col_name] = {
                                    "filename": filename,
                                    "path": filepath.replace('\\', '/'),
                                    "width": img.get('width', 0),
                                    "height": img.get('height', 0),
                                    "format": img.get('format', 'jpg'),
                                    "mrz_processed": False,  # Mark as not yet processed
                                    **basic_doc_info  # Add document number and type if found
                                }
                            except Exception as e:
                                # Continue processing even if one image fails
                                pass
                            
                            image_index += 1
            
            records.append(record)
    
    # Create JSON output with metadata including original column order
    output_data = {
        'records': records,
        'metadata': {
            'original_column_order': original_column_order,
            'total_records': len(records),
            'total_images': total_expected_positions if distribution_plan else image_index,
            'total_unique_images': total_available_images,
            'images_reused': total_expected_positions > total_available_images,
            'extraction_timestamp': datetime.now().isoformat()
        }
    }
    
    # Save to JSON (without MRZ data yet)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, indent=2, ensure_ascii=False)
    
    # Update progress
    if progress_tracker:
        progress_tracker.update(
            stage='completed_extraction',
            stage_message=f'Step 1 complete: Extracted {len(records)} records with {image_index} images'
        )
    
    # Calculate final image count
    total_images_assigned = total_expected_positions if distribution_plan else image_index

    return json_path, images_dir, len(records), total_images_assigned