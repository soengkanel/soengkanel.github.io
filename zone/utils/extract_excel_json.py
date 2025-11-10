"""
Extract Excel data and images to JSON format
Maps images to their correct columns (PHOTO, IDKH/PASSPORT, WORKINGPERMIT, VISA)
"""

import ast
from datetime import datetime
import os
from django.conf import settings
import pandas as pd
import json
from pathlib import Path
from PIL import Image
import pytesseract
import easyocr
import io
import zipfile
import re
import shutil
import openpyxl
import cv2
import numpy as np

from utils.mrz_parser import detect_mrz_type, parse_khid_mrz, parse_passport_mrz

json_filename = f"worker_eforms_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

# json_path = Path(json_filename)

json_path = os.path.join(json_filename, settings.BASE_DIR, 'media', 'worker_import_json')

os.makedirs(json_path, exist_ok=True)

def clean_filename(text, max_length=50):
    """Clean text to be safe for filename."""
    if not text:
        return "unnamed"
    text = re.sub(r'[^\w\s-]', '', str(text))
    text = text.strip().replace(' ', '_')
    return text[:max_length]

def extract_images_from_excel(excel_path):
    """Extract all images from Excel file (ZIP structure)."""
    images = []
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as zip_file:
            # Find all media files
            media_files = [f for f in zip_file.filelist if 'media/' in f.filename and not f.is_dir()]
            media_files.sort(key=lambda x: x.filename)
            
            print(f"Found {len(media_files)} images in Excel file")
            
            for idx, file_info in enumerate(media_files):
                # Read image data
                image_data = zip_file.read(file_info.filename)
                
                # Get image info
                img = Image.open(io.BytesIO(image_data))
                extension = f".{img.format.lower()}" if img.format else '.png'
                
                images.append({
                    "data": image_data,
                    "width": img.width,
                    "height": img.height,
                    "format": img.format.lower() if img.format else 'png',
                    "extension": extension,
                    "index": idx
                })
                
                print(f"  Image {idx + 1}: {img.width}x{img.height}px {extension}")
                
    except Exception as e:
        print(f"Error extracting images: {e}")
    
    return images

def detect_image_positions(excel_path):
    """Detect which cells contain images by checking for #VALUE! or formula errors."""
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb.active
    
    # Get column indices
    cols = {}
    for cell in ws[1]:
        if cell.value:
            cols[cell.value] = cell.column
    
    # Track image positions
    image_positions = []
    
    # Define image columns in their Excel order (left to right)
    image_columns_ordered = [
        ('IDKH/PASSPORT', 14),
        ('WORKINGPERMIT', 15), 
        ('VISA', 16),
        ('PHOTO', 17)
    ]
    
    # Check each row for image placeholders
    for row_idx in range(2, ws.max_row + 1):
        row_images = []
        
        # Check image columns IN ORDER (important for correct mapping)
        for col_name, expected_col in image_columns_ordered:
            if col_name in cols:
                cell = ws.cell(row=row_idx, column=cols[col_name])
                # Check if cell has #VALUE! or similar (indicates image)
                if cell.value and (str(cell.value) == '#VALUE!' or 
                                 '=IMAGE' in str(cell.value).upper() or
                                 cell.value == '=IMAGE()'):
                    row_images.append((col_name, cols[col_name]))
        
        if row_images:
            name_cell = ws.cell(row=row_idx, column=cols.get('NAME', 2))
            # Sort by column number to ensure correct order
            row_images.sort(key=lambda x: x[1])
            image_positions.append({
                'row': row_idx,
                'name': name_cell.value,
                'columns': [col[0] for col in row_images]  # Just column names in order
            })
    
    wb.close()
    return image_positions


# Initialize EasyOCR reader
# EasyOCR supports multiple languages and provides better accuracy
try:
    # Fix PIL compatibility issue
    import PIL
    if not hasattr(PIL.Image, 'ANTIALIAS'):
        PIL.Image.ANTIALIAS = PIL.Image.LANCZOS
    
    # Initialize with English and other languages that might be in passports
    reader = easyocr.Reader(['en'], gpu=False)  # Set gpu=True if you have CUDA
    print("EasyOCR initialized successfully")
except Exception as e:
    print(f"Warning: Could not initialize EasyOCR: {e}")
    reader = None

def extractMrz():
    print("\n")
    print("PROCESSING MRZ FROM PASSPORT IMAGES:")
    print("="*60)

    # Point to the correct tessdata directory
    #os.environ['TESSDATA_PREFIX'] = 'C:\\Program Files\\Tesseract-OCR\\tessdata'
    #pytesseract.pytesseract.tesseract_cmd = r'C:\\Program Files\\Tesseract-OCR\\tesseract.exe'

    # file_path = 'worker_eforms.json'
    file_path = json_path
    # output_path = 'worker_eforms_with_mrz.json'

    filename = f"worker_eforms_with_mrz_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    output_path = filename

    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    transformed_data = []
    
    # Process each record
    for i, record in enumerate(data):
        print(f"\nProcessing record {i+1}: {record.get('NAME', 'Unknown')}")
        transformed_record = record.copy()
        
        # Process ALL IDKH/PASSPORT images (could be multiple per person)
        passport_images_processed = []
        
        # Check for IDKH/PASSPORT field
        if record.get("IDKH/PASSPORT"):
            # Handle both single image and multiple images
            idkh_data = record["IDKH/PASSPORT"]
            
            # If it's a single image (has 'path' key)
            if isinstance(idkh_data, dict) and idkh_data.get("path"):
                images_to_process = [idkh_data]
            # If it's multiple images (list)
            elif isinstance(idkh_data, list):
                images_to_process = idkh_data
            else:
                images_to_process = []
            
            # Process each passport/ID image
            for img_idx, img_data in enumerate(images_to_process):
                if img_data.get("path"):
                    try:
                        print(f"  Processing passport/ID image {img_idx + 1}: {img_data['filename']}")
                        
                        # Try EasyOCR first if available
                        if reader:
                            try:
                                print(f"    Using EasyOCR for text extraction...")
                                
                                # EasyOCR can read directly from file path
                                result = reader.readtext(img_data["path"], detail=1)
                                
                                # Combine all text from EasyOCR
                                easyocr_text = '\n'.join([text[1] for text in result])
                                
                                # Also get bounding boxes and confidence scores
                                ocr_details = []
                                for bbox, text, conf in result:
                                    # Convert numpy arrays to lists for JSON serialization
                                    bbox_list = [[float(x), float(y)] for x, y in bbox] if bbox is not None else None
                                    ocr_details.append({
                                        "text": text,
                                        "confidence": float(conf),
                                        "bbox": bbox_list
                                    })
                                
                                # Extract MRZ lines from EasyOCR results
                                all_lines = easyocr_text.strip().split('\n')
                                mrz_lines = []
                                
                                # First, try to find clear MRZ patterns
                                for line in all_lines:
                                    line = line.strip()
                                    # Check for MRZ patterns
                                    if (('<' in line and len(line) > 20) or
                                        ('P<' in line.upper()) or
                                        ('ID<' in line.upper()) or
                                        (len(line) > 35 and line.count('<') > 3) or
                                        (len(line) > 30 and any(c in line for c in ['<', '«']) and any(c.isdigit() for c in line))):
                                        mrz_lines.append(line)
                                
                                # If no MRZ found, look more aggressively in OCR details
                                if not mrz_lines:
                                    # Look for lines that might be MRZ but poorly OCR'd
                                    for detail in ocr_details:
                                        text = detail['text'].strip()
                                        # Look for patterns typical of MRZ
                                        if (len(text) > 25 and 
                                            (('<' in text) or 
                                             ('P<' in text.upper()) or
                                             (text.count(' ') < 3 and any(c.isdigit() for c in text) and any(c.isupper() for c in text)) or
                                             (any(char in text for char in ['<', '«', 'T', 'A']) and len([c for c in text if c.isdigit()]) > 5))):
                                            mrz_lines.append(text)
                                
                                # If still no MRZ found, try enhanced preprocessing with Tesseract
                                if not mrz_lines:
                                    print(f"    No MRZ found with EasyOCR, trying enhanced preprocessing...")
                                    try:
                                        # Load image and preprocess
                                        img_pil = Image.open(img_data["path"])
                                        processed_img = preprocess_image_for_mrz(img_pil)
                                        processed_pil = Image.fromarray(processed_img)
                                        
                                        # Try OCR with different settings
                                        configs = [
                                            '--oem 3 --psm 6',
                                            '--oem 3 --psm 7',
                                            '-c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789< --oem 3 --psm 7'
                                        ]
                                        
                                        for config in configs:
                                            text = pytesseract.image_to_string(processed_pil, config=config)
                                            lines = [line.strip() for line in text.split('\n') if line.strip() and len(line.strip()) > 20]
                                            
                                            for line in lines:
                                                # Check for MRZ patterns
                                                if (('<' in line and len(line) > 25) or
                                                    ('P<' in line.upper()) or
                                                    (len(line) > 35 and any(c.isdigit() for c in line) and any(c.isupper() for c in line))):
                                                    mrz_lines.append(line)
                                            
                                            if mrz_lines:
                                                print(f"    Found {len(mrz_lines)} MRZ lines with enhanced preprocessing")
                                                break
                                                
                                    except Exception as e:
                                        print(f"    Enhanced preprocessing failed: {str(e)[:100]}")
                                
                                # Clean and format MRZ output
                                if mrz_lines:
                                    # Clean each MRZ line to remove invalid characters
                                    cleaned_lines = []
                                    for line in mrz_lines:
                                        # Keep only valid MRZ characters: A-Z, 0-9, <
                                        cleaned_line = ''.join(c for c in line.upper() if c.isalnum() or c == '<')
                                        if len(cleaned_line) > 20:  # Valid MRZ lines are typically long
                                            cleaned_lines.append(cleaned_line)
                                    
                                    if cleaned_lines:
                                        passport_mrz = '\n'.join(cleaned_lines)
                                        print(f"    Found {len(cleaned_lines)} cleaned MRZ lines with EasyOCR")
                                    else:
                                        passport_mrz = None
                                        print(f"    No valid MRZ lines after cleaning")
                                else:
                                    passport_mrz = None
                                    print(f"    No MRZ lines detected, but extracted {len(result)} text regions")
                                
                            except Exception as e:
                                # Fallback to Tesseract OCR
                                print(f"    EasyOCR failed: {str(e)[:100]}")
                                print(f"    Falling back to Tesseract OCR...")
                                img = Image.open(img_data["path"])
                                
                                # Try OCR with English language
                                custom_config = r'--oem 3 --psm 6'
                                text = pytesseract.image_to_string(img, lang='eng', config=custom_config)
                                
                                # Extract MRZ lines
                                all_lines = text.strip().split('\n')
                                mrz_lines = []
                                for line in all_lines:
                                    line = line.strip()
                                    cleaned_line = line.replace('�', '<').replace('«', '<').replace('K', '<').replace('C', '<')
                                    if (('<' in line and len(line) > 10) or
                                        ('<' in cleaned_line and len(cleaned_line) > 10) or
                                        (len(line) > 25 and any(c in line for c in ['<', 'K', 'C', 'E']) and any(c.isdigit() for c in line)) or
                                        ('VNM' in line.upper() or 'VIETNAM' in line.upper()) or
                                        ('VN' in line and len(line) > 30) or
                                        (line.startswith('P') and len(line) > 30)):
                                        mrz_lines.append(line)
                                
                                # Clean and format MRZ output
                                if mrz_lines:
                                    # Clean each MRZ line to remove invalid characters
                                    cleaned_lines = []
                                    for line in mrz_lines:
                                        # Keep only valid MRZ characters: A-Z, 0-9, <
                                        cleaned_line = ''.join(c for c in line.upper() if c.isalnum() or c == '<')
                                        if len(cleaned_line) > 20:  # Valid MRZ lines are typically long
                                            cleaned_lines.append(cleaned_line)
                                    
                                    if cleaned_lines:
                                        passport_mrz = '\n'.join(cleaned_lines)
                                        print(f"    Found {len(cleaned_lines)} cleaned MRZ lines")
                                    else:
                                        passport_mrz = None
                                        print(f"    No valid MRZ lines after cleaning")
                                else:
                                    passport_mrz = None
                        else:
                            # No FastMRZ, use basic OCR
                            img = Image.open(img_data["path"])
                            custom_config = r'--oem 3 --psm 6'
                            text = pytesseract.image_to_string(img, lang='eng', config=custom_config)
                            
                            # Extract MRZ lines
                            all_lines = text.strip().split('\n')
                            mrz_lines = [line.strip() for line in all_lines if '<' in line and len(line.strip()) > 20]
                            
                            # Clean and format MRZ output
                            if mrz_lines:
                                # Clean each MRZ line to remove invalid characters
                                cleaned_lines = []
                                for line in mrz_lines:
                                    # Keep only valid MRZ characters: A-Z, 0-9, <
                                    cleaned_line = ''.join(c for c in line.upper() if c.isalnum() or c == '<')
                                    if len(cleaned_line) > 20:  # Valid MRZ lines are typically long
                                        cleaned_lines.append(cleaned_line)
                                
                                if cleaned_lines:
                                    passport_mrz = '\n'.join(cleaned_lines)
                                    print(f"    Found {len(cleaned_lines)} cleaned MRZ lines")
                                else:
                                    passport_mrz = None
                                    print(f"    No valid MRZ lines after cleaning")
                            else:
                                passport_mrz = None
                        
                        print("    OCR processing complete!")

                        # Add MRZ data to the image data
                        img_data_with_mrz = img_data.copy()
                        img_data_with_mrz["mrz"] = passport_mrz
                        img_data_with_mrz["document"] = {}
                        document_mrz = {}
                        
                        # Process MRZ if lines were found
                        if 'mrz_lines' in locals() and mrz_lines:
                            doucType = detect_mrz_type(mrz_lines)
                            
                            if doucType == "KHID":
                                document_mrz = parse_khid_mrz(mrz_lines)
                                print("KHID:", parse_khid_mrz(mrz_lines))
                            elif doucType == "PASSPORT":
                                document_mrz = parse_passport_mrz(mrz_lines)
                                print("PASSPORT: ", parse_passport_mrz(mrz_lines))

                        img_data_with_mrz["document"] = document_mrz 

                        passport_images_processed.append(img_data_with_mrz)
                        
                    except Exception as e:
                        print(f"    Error processing image: {str(e)[:200]}")
                        print("    Adding image without MRZ data...")
                        passport_images_processed.append(img_data)
        
        # Update the record with processed passport images
        if passport_images_processed:
            # If originally was single image, keep as single
            if len(passport_images_processed) == 1 and isinstance(record.get("IDKH/PASSPORT"), dict):
                transformed_record['IDKH/PASSPORT'] = passport_images_processed[0]
            else:
                transformed_record['IDKH/PASSPORT'] = passport_images_processed
        elif not record.get("IDKH/PASSPORT"):
            print(f"  No passport/ID images found for {record.get('NAME', 'Unknown')}")
        
        transformed_data.append(transformed_record)

    # Save results
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(transformed_data, f, indent=2, ensure_ascii=False)
    
    print(f"\n{'='*60}")
    print(f"MRZ PROCESSING COMPLETE!")
    print(f"{'='*60}")
    print(f"Results saved to: {output_path}")
    print(f"Total records processed: {len(transformed_data)}")
    
    return transformed_data

def preprocess_image_for_mrz(image):
    """Enhanced image preprocessing specifically for MRZ extraction"""
    # Convert PIL Image to OpenCV format (if needed)
    if isinstance(image, Image.Image):  # PIL Image
        image = np.array(image)
        image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)  # PIL uses RGB; OpenCV uses BGR
    
    # Convert to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Increase contrast with CLAHE
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    enhanced = clahe.apply(gray)
    
    # Apply threshold
    _, thresh = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # Scale up the image for better OCR
    scale_factor = 3
    height, width = thresh.shape
    resized = cv2.resize(thresh, (width * scale_factor, height * scale_factor), interpolation=cv2.INTER_CUBIC)
    
    return resized


def reading_excel(excel_file):
    """
    Main extraction function.
    Properly maps images to their respective columns.
    """
    media_root = getattr(settings, 'MEDIA_ROOT', '')
    
    # Convert to Path objects
    excel_path = Path(excel_file)
    # images_dir = Path(images_folder)
    images_dir = os.path.join(media_root, 'worker_pworker_import_photoshotos'),
    
    # Verify Excel file exists
    if not excel_path.exists():
        print(f"Error: Excel file '{excel_path}' not found!")
        return None
    
    # Clean and create images directory
    if images_dir.exists():
        shutil.rmtree(images_dir)
    images_dir.mkdir(parents=True, exist_ok=True)
    
    # Extract images first
    images = extract_images_from_excel(excel_path)

    # Detect where images should go

    image_positions = detect_image_positions(excel_path)
    
    # Read Excel data
    df = pd.read_excel(excel_path)
    
    # Process records
    records = []
    image_index = 0
    
    for idx, row in df.iterrows():
        # Skip rows with empty NAME field
        if pd.isna(row.get('NAME')) or row.get('NAME') == '' or str(row.get('NAME')).strip() == '':
            continue
            
        # Check if row has valid data (at least NO field)
        if pd.notna(row.get('NO')):
            record = {}
            
            # Convert all fields
            for col in df.columns:
                value = row[col]
                
                if pd.isna(value):
                    record[col] = None
                elif isinstance(value, pd.Timestamp):
                    record[col] = value.strftime('%Y-%m-%d')
                elif hasattr(value, 'item'):  # numpy type
                    record[col] = value.item()
                # Skip #VALUE! or =IMAGE() placeholders
                elif str(value) in ['#VALUE!', '=IMAGE()']:
                    record[col] = None
                else:
                    record[col] = value
            
            # Find image mapping for this row
            row_num = idx + 2  # Excel row number (1-indexed + header)
            row_image_info = next((p for p in image_positions if p['row'] == row_num), None)
            
            if row_image_info and image_index < len(images):
                worker_id = clean_filename(record.get('NAME')) if record.get('NAME') else f"worker_{int(record.get('NO', idx + 1))}"
                
                # Distribute images to the correct columns
                for col_idx, col_name in enumerate(row_image_info['columns']):
                    if image_index < len(images):
                        img = images[image_index]
                        
                        # Create filename using NAME only (add number if multiple images for same person)
                        if len(row_image_info['columns']) > 1:
                            # Multiple images for this person, add index
                            filename = f"{worker_id}_{col_idx + 1}{img['extension']}"
                        else:
                            # Single image for this person
                            filename = f"{worker_id}{img['extension']}"
                        
                        # Save image
                        filepath = images_dir / filename
                        with open(filepath, 'wb') as f:
                            f.write(img['data'])
                        
                        # Add image info to the correct column
                        record[col_name] = {
                            "filename": filename,
                            "path": str(filepath).replace('\\', '/'),
                            "width": img['width'],
                            "height": img['height'],
                            "format": img['format']
                        }
                        
                        print(f"  Row {row_num}: {worker_id} -> {col_name}: {filename}")
                        image_index += 1
            
            records.append(record)
    
    # Save to JSON
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(records, f, indent=2, ensure_ascii=False)
    
    for i, record in enumerate(records[:7]):  # Show first 7
        name = record.get('NAME', f"Worker {record.get('NO', i+1)}")
        # Check which columns have images
        img_cols = []
        for col in ['PHOTO', 'IDKH/PASSPORT', 'WORKINGPERMIT', 'VISA']:
            if record.get(col) and isinstance(record[col], dict):
                img_cols.append(col)
        
        if img_cols:
            print(f"  {i+1}. {name}: {', '.join(img_cols)}")
        else:
            print(f"  {i+1}. {name}: No images")
    
    # Call Extract MRZ from Json File
    extractMrz()
    return records
