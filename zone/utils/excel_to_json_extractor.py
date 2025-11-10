"""
Excel to JSON extractor with proper image mapping
Based on OCR/extract_excel_to_json.py but integrated for Django
"""

import ast
from datetime import datetime
import os
from django.conf import settings
import pandas as pd
import json
from pathlib import Path
from PIL import Image
import pytesseract
import easyocr
import io
import zipfile
import re
import shutil
import openpyxl
import cv2
import numpy as np
import tempfile
import logging

from .mrz_parser import detect_mrz_type, parse_khid_mrz, parse_passport_mrz

logger = logging.getLogger(__name__)

def clean_filename(text, max_length=50):
    """Clean text to be safe for filename."""
    if not text:
        return "unnamed"
    text = re.sub(r'[^\w\s-]', '', str(text))
    text = text.strip().replace(' ', '_')
    return text[:max_length]

def extract_images_from_excel(excel_path):
    """Extract all images from Excel file (ZIP structure)."""
    images = []
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as zip_file:
            # Find all media files
            media_files = [f for f in zip_file.filelist if 'media/' in f.filename and not f.is_dir()]
            media_files.sort(key=lambda x: x.filename)
            
            for idx, file_info in enumerate(media_files):
                # Read image data
                image_data = zip_file.read(file_info.filename)
                
                # Get image info
                img = Image.open(io.BytesIO(image_data))
                extension = f".{img.format.lower()}" if img.format else '.png'
                
                images.append({
                    "data": image_data,
                    "width": img.width,
                    "height": img.height,
                    "format": img.format.lower() if img.format else 'png',
                    "extension": extension,
                    "index": idx
                })

    except Exception as e:
        pass
    
    return images

def detect_image_positions(excel_path):
    """Detect which cells contain images by checking for #VALUE! or formula errors."""
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    ws = wb['Workers'] if 'Workers' in wb.sheetnames else wb.active
    
    # Get column indices
    cols = {}
    for cell in ws[1]:
        if cell.value:
            cols[cell.value] = cell.column
    
    # Track image positions
    image_positions = []
    
    # Define image columns in their Excel order (left to right)
    image_columns_ordered = [
        ('IDKH/PASSPORT', 14),
        ('WORKINGPERMIT', 15), 
        ('VISA', 16),
        ('PHOTO', 17)
    ]
    
    # Check each row for image placeholders
    for row_idx in range(2, ws.max_row + 1):
        row_images = []
        
        # Check image columns IN ORDER (important for correct mapping)
        for col_name, expected_col in image_columns_ordered:
            if col_name in cols:
                cell = ws.cell(row=row_idx, column=cols[col_name])
                # Check if cell has #VALUE! or similar (indicates image)
                if cell.value and (str(cell.value) == '#VALUE!' or 
                                 '=IMAGE' in str(cell.value).upper() or
                                 cell.value == '=IMAGE()'):
                    row_images.append((col_name, cols[col_name]))
        
        if row_images:
            name_cell = ws.cell(row=row_idx, column=cols.get('NAME', 2))
            # Sort by column number to ensure correct order
            row_images.sort(key=lambda x: x[1])
            image_positions.append({
                'row': row_idx,
                'name': name_cell.value,
                'columns': [col[0] for col in row_images]  # Just column names in order
            })
    
    wb.close()
    return image_positions

def preprocess_image_for_mrz(image):
    """Enhanced image preprocessing specifically for MRZ extraction"""
    # Convert PIL Image to OpenCV format (if needed)
    if isinstance(image, Image.Image):  # PIL Image
        image = np.array(image)
        image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)  # PIL uses RGB; OpenCV uses BGR
    
    # Convert to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Increase contrast with CLAHE
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    enhanced = clahe.apply(gray)
    
    # Apply threshold
    _, thresh = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # Scale up the image for better OCR
    scale_factor = 3
    height, width = thresh.shape
    resized = cv2.resize(thresh, (width * scale_factor, height * scale_factor), interpolation=cv2.INTER_CUBIC)
    
    return resized

def extract_excel_to_json(excel_file_path, progress_tracker=None):
    """
    Main extraction function that converts Excel to JSON with proper image mapping.
    Returns the path to the generated JSON file.
    """
    media_root = getattr(settings, 'MEDIA_ROOT', '')
    
    # Create timestamped filenames
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    json_filename = f"worker_eforms_{timestamp}.json"
    
    # Create directories
    json_dir = os.path.join(media_root, 'worker_import_json')
    images_dir = os.path.join(media_root, 'worker_import_photos')
    
    os.makedirs(json_dir, exist_ok=True)
    os.makedirs(images_dir, exist_ok=True)
    
    json_path = os.path.join(json_dir, json_filename)
    
    # Update progress
    if progress_tracker:
        progress_tracker.update(
            stage='extracting_excel',
            stage_message='Extracting images from Excel file...'
        )
    
    # Verify Excel file exists
    excel_path = Path(excel_file_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file '{excel_path}' not found!")
    
    # Extract images first
    images = extract_images_from_excel(excel_path)
    
    # Update progress
    if progress_tracker:
        progress_tracker.update(
            stage='detecting_positions',
            stage_message='Detecting image positions in Excel...'
        )
    
    # Detect where images should go
    image_positions = detect_image_positions(excel_path)

    # Read Excel data
    df = pd.read_excel(excel_path, sheet_name='Workers')
    
    # Update progress
    if progress_tracker:
        progress_tracker.update(
            stage='processing_records',
            stage_message='Processing Excel records...'
        )
    
    # Process records
    records = []
    image_index = 0
    
    for idx, row in df.iterrows():
        # Skip rows with empty NAME field
        if pd.isna(row.get('NAME')) or row.get('NAME') == '' or str(row.get('NAME')).strip() == '':
            continue
            
        # Check if row has valid data (at least NO field)
        if pd.notna(row.get('NO')):
            record = {}
            
            # Convert all fields
            for col in df.columns:
                value = row[col]
                
                if pd.isna(value):
                    record[col] = None
                elif isinstance(value, pd.Timestamp):
                    record[col] = value.strftime('%Y-%m-%d')
                elif hasattr(value, 'item'):  # numpy type
                    record[col] = value.item()
                # Skip #VALUE! or =IMAGE() placeholders
                elif str(value) in ['#VALUE!', '=IMAGE()']:
                    record[col] = None
                else:
                    record[col] = value
            
            # Find image mapping for this row
            row_num = idx + 2  # Excel row number (1-indexed + header)
            row_image_info = next((p for p in image_positions if p['row'] == row_num), None)
            
            if row_image_info and image_index < len(images):
                worker_id = clean_filename(record.get('NAME')) if record.get('NAME') else f"worker_{int(record.get('NO', idx + 1))}"
                
                # Distribute images to the correct columns
                for col_idx, col_name in enumerate(row_image_info['columns']):
                    if image_index < len(images):
                        img = images[image_index]
                        
                        # Create filename using NAME only (add number if multiple images for same person)
                        if len(row_image_info['columns']) > 1:
                            # Multiple images for this person, add index
                            filename = f"{worker_id}_{col_idx + 1}{img['extension']}"
                        else:
                            # Single image for this person
                            filename = f"{worker_id}{img['extension']}"
                        
                        # Save image
                        filepath = os.path.join(images_dir, filename)
                        with open(filepath, 'wb') as f:
                            f.write(img['data'])

                        # Add image info to the correct column
                        record[col_name] = {
                            "filename": filename,
                            "path": filepath.replace('\\\\', '/'),
                            "width": img['width'],
                            "height": img['height'],
                            "format": img['format']
                        }

                        image_index += 1
            
            records.append(record)
    
    # Save to JSON
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(records, f, indent=2, ensure_ascii=False)
    
    # Update progress
    if progress_tracker:
        progress_tracker.update(
            stage='completed_extraction',
            stage_message=f'Extracted {len(records)} records with {image_index} images'
        )

    return json_path, len(records), image_index