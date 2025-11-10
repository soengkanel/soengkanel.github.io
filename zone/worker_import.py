"""
Worker Import Module
Clean implementation leveraging extract_xlsx_final.py logic
"""
import zipfile
import os
import json
import openpyxl
import shutil
from datetime import datetime
import xml.etree.ElementTree as ET
import re
from typing import Dict, List, Tuple, Optional, Any
import logging
from django.core.files.base import ContentFile
from django.db import transaction

logger = logging.getLogger(__name__)


class ExcelWorkerImporter:
    """
    Excel Worker Importer using exact cell-to-image mapping
    Based on extract_xlsx_final.py logic
    """
    
    # Column mapping based on actual Excel structure
    COLUMN_MAPPING = {
        1: 'no',           # A - NO
        2: 'photo',        # B - PHOTO
        3: 'name',         # C - NAME
        4: 'sex',          # D - SEX
        5: 'dob',          # E - DOB
        6: 'nationality',  # F - NATIONALITY
        7: 'building',     # G - BUILDING
        8: 'position',     # H - POSITION
        9: 'joined_date',  # I - JOINEDDATE
        10: 'resigned_date', # J - RESIGNEDDATE
        11: 'passport_file', # K - IDKH/PASSPORT
        12: 'passport_no',   # L - IDKHPASSPORT NO
        13: 'passport_dates', # M - IDKH/PASSPORT_DATE
        14: 'workpermit_file', # N - WORKINGPERMIT
        15: 'workpermit_dates', # O - WORKINGPERMIT_DATE
        16: 'visa_file',     # P - VISA
        17: 'visa_dates',    # Q - VISA_DATE
        # No more columns after Q
    }
    
    NATIONALITY_MAPPING = {
        # Southeast Asian countries
        'VIETNAMESE': 'VN',
        'VIETNAM': 'VN',
        'CAMBODIAN': 'KH',
        'CAMBODIA': 'KH',
        'INDONESIAN': 'ID',
        'INDONESIA': 'ID',
        'THAI': 'TH',
        'THAILAND': 'TH',
        'MYANMAR': 'MM',
        'BURMESE': 'MM',
        'PHILIPPINES': 'PH',
        'FILIPINO': 'PH',
        'PHILIPPINE': 'PH',
        'LAOS': 'LA',
        'LAO': 'LA',
        'LAOTIAN': 'LA',
        'MALAYSIAN': 'MY',
        'MALAYSIA': 'MY',
        'SINGAPOREAN': 'SG',
        'SINGAPORE': 'SG',
        'BRUNEI': 'BN',
        'BRUNEIAN': 'BN',
        'TIMOR-LESTE': 'TL',
        'EAST TIMORESE': 'TL',

        # East Asian countries
        'CHINESE': 'CN',
        'CHINA': 'CN',
        'JAPANESE': 'JP',
        'JAPAN': 'JP',
        'SOUTH KOREAN': 'KR',
        'KOREAN': 'KR',
        'NORTH KOREAN': 'KP',
        'MONGOLIAN': 'MN',
        'MONGOLIA': 'MN',
        'TAIWANESE': 'TW',
        'TAIWAN': 'TW',
        'HONG KONG': 'HK',
        'MACAU': 'MO',

        # South Asian countries
        'INDIAN': 'IN',
        'INDIA': 'IN',
        'PAKISTANI': 'PK',
        'PAKISTAN': 'PK',
        'BANGLADESHI': 'BD',
        'BANGLADESH': 'BD',
        'SRI LANKAN': 'LK',
        'SRI LANKA': 'LK',
        'NEPALESE': 'NP',
        'NEPAL': 'NP',
        'BHUTANESE': 'BT',
        'BHUTAN': 'BT',
        'MALDIVIAN': 'MV',
        'MALDIVES': 'MV',
        'AFGHAN': 'AF',
        'AFGHANISTAN': 'AF',

        # Middle Eastern countries
        'SAUDI': 'SA',
        'SAUDI ARABIAN': 'SA',
        'EMIRATI': 'AE',
        'UAE': 'AE',
        'IRANIAN': 'IR',
        'IRAN': 'IR',
        'IRAQI': 'IQ',
        'IRAQ': 'IQ',
        'TURKISH': 'TR',
        'TURKEY': 'TR',
        'ISRAELI': 'IL',
        'ISRAEL': 'IL',
        'PALESTINIAN': 'PS',
        'PALESTINE': 'PS',
        'JORDANIAN': 'JO',
        'JORDAN': 'JO',
        'LEBANESE': 'LB',
        'LEBANON': 'LB',
        'SYRIAN': 'SY',
        'SYRIA': 'SY',
        'KUWAITI': 'KW',
        'KUWAIT': 'KW',
        'QATARI': 'QA',
        'QATAR': 'QA',
        'BAHRAINI': 'BH',
        'BAHRAIN': 'BH',
        'OMANI': 'OM',
        'OMAN': 'OM',
        'YEMENI': 'YE',
        'YEMEN': 'YE',

        # European countries
        'BRITISH': 'GB',
        'UK': 'GB',
        'ENGLISH': 'GB',
        'SCOTTISH': 'GB',
        'WELSH': 'GB',
        'IRISH': 'IE',
        'IRELAND': 'IE',
        'FRENCH': 'FR',
        'FRANCE': 'FR',
        'GERMAN': 'DE',
        'GERMANY': 'DE',
        'ITALIAN': 'IT',
        'ITALY': 'IT',
        'SPANISH': 'ES',
        'SPAIN': 'ES',
        'PORTUGUESE': 'PT',
        'PORTUGAL': 'PT',
        'DUTCH': 'NL',
        'NETHERLANDS': 'NL',
        'BELGIAN': 'BE',
        'BELGIUM': 'BE',
        'SWISS': 'CH',
        'SWITZERLAND': 'CH',
        'AUSTRIAN': 'AT',
        'AUSTRIA': 'AT',
        'SWEDISH': 'SE',
        'SWEDEN': 'SE',
        'NORWEGIAN': 'NO',
        'NORWAY': 'NO',
        'DANISH': 'DK',
        'DENMARK': 'DK',
        'FINNISH': 'FI',
        'FINLAND': 'FI',
        'RUSSIAN': 'RU',
        'RUSSIA': 'RU',
        'POLISH': 'PL',
        'POLAND': 'PL',
        'CZECH': 'CZ',
        'CZECH REPUBLIC': 'CZ',
        'SLOVAK': 'SK',
        'SLOVAKIA': 'SK',
        'HUNGARIAN': 'HU',
        'HUNGARY': 'HU',
        'ROMANIAN': 'RO',
        'ROMANIA': 'RO',
        'BULGARIAN': 'BG',
        'BULGARIA': 'BG',
        'CROATIAN': 'HR',
        'CROATIA': 'HR',
        'SERBIAN': 'RS',
        'SERBIA': 'RS',
        'GREEK': 'GR',
        'GREECE': 'GR',
        'UKRAINIAN': 'UA',
        'UKRAINE': 'UA',

        # African countries
        'SOUTH AFRICAN': 'ZA',
        'SOUTH AFRICA': 'ZA',
        'NIGERIAN': 'NG',
        'NIGERIA': 'NG',
        'EGYPTIAN': 'EG',
        'EGYPT': 'EG',
        'KENYAN': 'KE',
        'KENYA': 'KE',
        'ETHIOPIAN': 'ET',
        'ETHIOPIA': 'ET',
        'GHANAIAN': 'GH',
        'GHANA': 'GH',
        'MOROCCAN': 'MA',
        'MOROCCO': 'MA',
        'ALGERIAN': 'DZ',
        'ALGERIA': 'DZ',
        'TUNISIAN': 'TN',
        'TUNISIA': 'TN',
        'ZIMBABWEAN': 'ZW',
        'ZIMBABWE': 'ZW',

        # American countries
        'AMERICAN': 'US',
        'USA': 'US',
        'US': 'US',
        'CANADIAN': 'CA',
        'CANADA': 'CA',
        'MEXICAN': 'MX',
        'MEXICO': 'MX',
        'BRAZILIAN': 'BR',
        'BRAZIL': 'BR',
        'ARGENTINIAN': 'AR',
        'ARGENTINA': 'AR',
        'CHILEAN': 'CL',
        'CHILE': 'CL',
        'COLOMBIAN': 'CO',
        'COLOMBIA': 'CO',
        'PERUVIAN': 'PE',
        'PERU': 'PE',
        'VENEZUELAN': 'VE',
        'VENEZUELA': 'VE',

        # Oceania
        'AUSTRALIAN': 'AU',
        'AUSTRALIA': 'AU',
        'NEW ZEALAND': 'NZ',
        'NEW ZEALANDER': 'NZ',
        'FIJIAN': 'FJ',
        'FIJI': 'FJ',
        'PAPUA NEW GUINEA': 'PG',

        # Additional common variations
        'STATELESS': 'XX',
        'UNKNOWN': 'XX',
        'NOT SPECIFIED': 'XX',
    }
    
    def __init__(self, file_path: str, validate_only: bool = False):
        """Initialize with Excel file path
        
        Args:
            file_path: Path to the Excel file
            validate_only: If True, only validate without importing
        """
        self.file_path = file_path
        self.image_mapping = {}
        self.extracted_images = {}
        self.workers_data = []
        self.validate_only = validate_only
        
    def get_exact_image_mapping(self) -> Dict[str, Dict]:
        """
        Get the exact cell-to-image mapping from Excel metadata
        Direct implementation from extract_xlsx_final.py
        """
        mapping = {}
        
        with zipfile.ZipFile(self.file_path, 'r') as zip_ref:
            # Get image relationships (rId -> image filename)
            image_rel_map = {}
            
            if 'xl/richData/_rels/richValueRel.xml.rels' in zip_ref.namelist():
                content = zip_ref.read('xl/richData/_rels/richValueRel.xml.rels')
                root = ET.fromstring(content)
                
                for elem in root:
                    if 'Relationship' in elem.tag:
                        rel_id = elem.get('Id')
                        target = elem.get('Target')
                        if '../media/' in target:
                            image_name = target.replace('../media/', '')
                            image_rel_map[rel_id] = image_name
            
            # Get cell-to-vm mapping from worksheet
            cell_vm_map = {}  # cell -> vm_id
            vm_cell_map = {}  # vm_id -> cell
            
            if 'xl/worksheets/sheet1.xml' in zip_ref.namelist():
                content = zip_ref.read('xl/worksheets/sheet1.xml')
                content_str = content.decode('utf-8')
                
                # Find all cells with vm attribute
                cell_vm_pattern = r'<c r="([^"]+)"[^>]*vm="(\d+)"'
                matches = re.findall(cell_vm_pattern, content_str)
                
                for cell_ref, vm_id in matches:
                    vm_id = int(vm_id)
                    cell_vm_map[cell_ref] = vm_id
                    vm_cell_map[vm_id] = cell_ref
            
            # Get the ordering of rIds from richValueRel
            rel_order = []
            if 'xl/richData/richValueRel.xml' in zip_ref.namelist():
                content = zip_ref.read('xl/richData/richValueRel.xml')
                root = ET.fromstring(content)
                
                # Find all rel elements in order
                for elem in root.iter():
                    if 'rel' in elem.tag.lower() and elem.tag != root.tag:
                        for attr_name, attr_value in elem.attrib.items():
                            if 'id' in attr_name.lower() and attr_value.startswith('rId'):
                                rel_order.append(attr_value)
                                break
            
            # Map vm_id to rId (assuming they're in order)
            for vm_id in sorted(vm_cell_map.keys()):
                idx = vm_id - 1  # vm_id is 1-based
                if idx < len(rel_order):
                    rel_id = rel_order[idx]
                    if rel_id in image_rel_map:
                        cell_ref = vm_cell_map[vm_id]
                        image_name = image_rel_map[rel_id]
                        
                        # Parse cell reference (e.g., B2 -> row=2, col=B)
                        col_letter = ''.join(c for c in cell_ref if c.isalpha())
                        row_num = int(''.join(c for c in cell_ref if c.isdigit()))
                        col_num = openpyxl.utils.column_index_from_string(col_letter)
                        
                        matrix_pos = f"{row_num}x{col_num}"
                        mapping[matrix_pos] = {
                            'image': image_name,
                            'cell': cell_ref,
                            'row': row_num,
                            'col': col_num,
                            'col_letter': col_letter
                        }
        
        self.image_mapping = mapping
        return mapping
    
    def extract_images(self) -> Dict[str, bytes]:
        """
        Extract images from Excel file
        Returns dict of matrix_position -> image_data
        """
        images = {}
        
        with zipfile.ZipFile(self.file_path, 'r') as zip_ref:
            for matrix_pos, info in self.image_mapping.items():
                source_path = f"xl/media/{info['image']}"
                
                if source_path in zip_ref.namelist():
                    image_data = zip_ref.read(source_path)
                    images[matrix_pos] = {
                        'data': image_data,
                        'original_name': info['image'],
                        'cell': info['cell'],
                        'row': info['row'],
                        'col': info['col']
                    }
        
        self.extracted_images = images
        return images
    
    def extract_workers_data(self) -> List[Dict]:
        """
        Extract all worker data from Excel
        """
        # Get image mapping first
        self.get_exact_image_mapping()
        self.extract_images()
        
        # Load workbook with proper context management
        wb = None
        workers = []
        
        try:
            wb = openpyxl.load_workbook(self.file_path, data_only=True)
            sheet = wb.active
            
            # Skip header row, process data rows
            for row_idx in range(2, sheet.max_row + 1):
                worker_data = {
                    'row': row_idx,
                    'data': {},
                    'images': {}
                }
                
                # Process each column
                for col_idx in range(1, min(20, sheet.max_column + 1)):  # Process up to column S
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    
                    # Handle different value types and errors
                    if value is None:
                        value = ''
                    elif isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d')
                    elif str(value).strip() in ['#VALUE!', '#ERROR!', '#REF!', '#N/A']:
                        value = ''  # Convert Excel errors to empty string
                    else:
                        value = str(value).strip() if value is not None else ''
                    
                    # Store in data dict using column mapping
                    if col_idx in self.COLUMN_MAPPING:
                        field_name = self.COLUMN_MAPPING[col_idx]
                        worker_data['data'][field_name] = value
                    
                    # Check for image at this position
                    matrix_pos = f"{row_idx}x{col_idx}"
                    if matrix_pos in self.extracted_images:
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        worker_data['images'][col_letter] = self.extracted_images[matrix_pos]
                
                # Only add if worker has a number (not empty row)
                if worker_data['data'].get('no'):
                    workers.append(worker_data)
        
        finally:
            # Always close the workbook to release file handle
            if wb:
                wb.close()
        
        self.workers_data = workers
        return workers
    
    def parse_date(self, date_value: Any) -> Optional[datetime.date]:
        """Parse date from various formats"""
        if not date_value:
            return None
        
        if isinstance(date_value, datetime):
            return date_value.date()
        
        if isinstance(date_value, str):
            # Try common date formats
            for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                try:
                    return datetime.strptime(date_value.strip(), fmt).date()
                except:
                    continue
        
        return None
    
    def parse_date_range(self, date_range: str) -> Tuple[Optional[datetime.date], Optional[datetime.date]]:
        """Parse date range like '14/11/2018-14/11/2028' or various other formats"""
        if not date_range or not isinstance(date_range, str):
            return None, None
        
        date_range = str(date_range).strip()
        if '-' not in date_range:
            # Single date - treat as issue date only
            single_date = self.parse_date(date_range)
            return single_date, None
        
        try:
            parts = date_range.split('-')
            if len(parts) != 2:
                return None, None
                
            start_str, end_str = parts[0].strip(), parts[1].strip()
            
            # Try multiple date formats
            date_formats = ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y']
            
            start_date = None
            end_date = None
            
            for fmt in date_formats:
                if not start_date:
                    try:
                        start_date = datetime.strptime(start_str, fmt).date()
                    except ValueError:
                        continue
                        
                if not end_date:
                    try:
                        end_date = datetime.strptime(end_str, fmt).date()
                    except ValueError:
                        continue
                        
                if start_date and end_date:
                    break
            
            return start_date, end_date

        except Exception as e:
            return None, None
    
    def parse_name(self, full_name: str) -> Tuple[str, str]:
        """Parse full name into first and last name"""
        if not full_name:
            return '', ''
        
        name_parts = str(full_name).strip().split(' ', 1)
        first_name = name_parts[0]
        last_name = name_parts[1] if len(name_parts) > 1 else ''
        
        return first_name, last_name
    
    def format_document_number(self, doc_number: Any) -> str:
        """Format document number, handling scientific notation"""
        if not doc_number:
            return ''
        
        doc_str = str(doc_number).strip()
        
        # Handle scientific notation
        if 'e+' in doc_str.lower():
            try:
                doc_str = str(int(float(doc_str)))
            except:
                pass
        
        return doc_str
    
    def map_nationality(self, nationality: str) -> str:
        """Map nationality to country code"""
        if not nationality:
            return ''
        
        nationality_upper = str(nationality).strip().upper()
        return self.NATIONALITY_MAPPING.get(nationality_upper, '')
    
    def validate_data(self) -> Dict[str, Any]:
        """
        Validate extracted data without importing to database
        Returns validation results with detailed error messages
        """
        from zone.models import Document, Building
        
        results = {
            'total_rows': len(self.workers_data),
            'valid_rows': 0,
            'error_rows': 0,
            'warning_rows': 0,
            'validation_errors': [],
            'warnings': [],
            'summary': {
                'can_import': True,
                'critical_issues': [],
                'recommendations': []
            }
        }
        
        # Extract data if not already done
        if not self.workers_data:
            self.extract_workers_data()
        
        for worker_info in self.workers_data:
            row_num = worker_info['row']
            data = worker_info['data']
            row_errors = []
            row_warnings = []
            
            # Validate name (required)
            name = str(data.get('name', '')).strip()
            if not name:
                row_errors.append({
                    'field': 'Name',
                    'value': '',
                    'error': 'Please provide worker name',
                    'column': 'C',
                    'suggestion': 'Enter full name in column C'
                })
            elif len(name.split()) < 2:
                row_warnings.append({
                    'field': 'Name',
                    'value': name,
                    'warning': 'Consider adding last name',
                    'column': 'C',
                    'suggestion': f'Current: "{name}" - Try: "{name} [LastName]"'
                })
            
            # Validate sex
            if data.get('sex'):
                sex_value = str(data['sex']).strip()
                sex = sex_value.upper()[:1] if sex_value else ''
                if sex and sex not in ['M', 'F']:
                    row_warnings.append({
                        'field': 'Gender',
                        'value': sex_value,
                        'warning': 'Use M for Male or F for Female',
                        'column': 'D',
                        'suggestion': 'Accepted values: M, F, Male, Female'
                    })
            
            # Validate passport duplication
            passport_no = data.get('passport_no')
            if passport_no:
                formatted_passport = self.format_document_number(passport_no)
                if formatted_passport:
                    # Check for duplicates
                    existing_docs = Document.objects.filter(
                        document_type='passport',
                        document_number=formatted_passport
                    ).select_related('worker')
                    
                    if existing_docs.exists():
                        existing_worker = existing_docs.first().worker
                        row_errors.append({
                            'field': 'Passport Number',
                            'value': passport_no,
                            'error': f'Duplicate - Already assigned to {existing_worker.get_full_name()}',
                            'column': 'L',
                            'duplicate_worker_id': existing_worker.id,
                            'suggestion': f'Check row {row_num} or update existing worker'
                        })
                else:
                    row_warnings.append({
                        'field': 'Passport Number',
                        'value': passport_no,
                        'warning': 'Format may be incorrect',
                        'column': 'L',
                        'suggestion': 'Remove special characters except hyphens'
                    })
            
            # Validate building
            building_name = data.get('building', '').strip()
            if not building_name:
                row_warnings.append({
                    'field': 'Building',
                    'value': '',
                    'warning': 'No building specified',
                    'column': 'G',
                    'suggestion': 'Worker will be assigned to default location'
                })
            
            # Validate dates
            if data.get('dob'):
                try:
                    parsed_date = self.parse_date(data['dob'])
                    # Check if date is reasonable (not in future, not too old)
                    from datetime import datetime, timedelta
                    today = datetime.now().date()
                    if parsed_date and parsed_date > today:
                        row_errors.append({
                            'field': 'Date of Birth',
                            'value': data.get('dob'),
                            'error': 'Date cannot be in the future',
                            'column': 'E',
                            'suggestion': 'Check year - should be in the past'
                        })
                    elif parsed_date and parsed_date < (today - timedelta(days=365*100)):
                        row_warnings.append({
                            'field': 'Date of Birth',
                            'value': data.get('dob'),
                            'warning': 'Date seems unusually old',
                            'column': 'E',
                            'suggestion': 'Verify the year is correct'
                        })
                except:
                    row_warnings.append({
                        'field': 'Date of Birth',
                        'value': data.get('dob'),
                        'warning': 'Date format not recognized',
                        'column': 'E',
                        'suggestion': 'Use format: DD/MM/YYYY or YYYY-MM-DD'
                    })
            
            # Record results
            if row_errors:
                results['error_rows'] += 1
                results['validation_errors'].append({
                    'row': row_num,
                    'worker_name': name or 'Unknown',
                    'errors': row_errors,
                    'warnings': row_warnings
                })
            elif row_warnings:
                results['warning_rows'] += 1
                results['warnings'].append({
                    'row': row_num,
                    'worker_name': name or 'Unknown',
                    'warnings': row_warnings
                })
                results['valid_rows'] += 1  # Still valid, just with warnings
            else:
                results['valid_rows'] += 1
        
        # Generate summary
        if results['error_rows'] > 0:
            results['summary']['can_import'] = False
            results['summary']['critical_issues'].append(
                f"{results['error_rows']} row{'s' if results['error_rows'] > 1 else ''} with critical errors must be fixed"
            )
        
        if results['warning_rows'] > 0:
            results['summary']['recommendations'].append(
                f"{results['warning_rows']} row{'s' if results['warning_rows'] > 1 else ''} have warnings - review recommended but not required"
            )
        
        # Add specific recommendations based on common issues
        passport_errors = sum(1 for err in results['validation_errors'] 
                            for e in err['errors'] if e['field'] == 'Passport Number')
        if passport_errors > 0:
            results['summary']['recommendations'].append(
                f"Found {passport_errors} duplicate passport{'s' if passport_errors > 1 else ''} - verify these are different workers"
            )
        
        name_warnings = sum(1 for warn in results['warnings'] 
                          for w in warn['warnings'] if w['field'] == 'Name')
        if name_warnings > 0:
            results['summary']['recommendations'].append(
                f"{name_warnings} worker{'s' if name_warnings > 1 else ''} may have incomplete names - consider adding last names"
            )
        
        return results
    
    @transaction.atomic
    def import_to_database(self, created_by) -> Dict[str, Any]:
        """
        Import workers to database
        Returns results dictionary with counts and messages
        """
        from zone.models import Worker, Document, Building, Position, Zone, Department, Floor
        
        results = {
            'success_count': 0,
            'error_count': 0,
            'workers_created': [],
            'errors': [],
            'warnings': [],
            'validation_errors': []  # Detailed validation errors per row
        }
        
        # Extract data if not already done
        if not self.workers_data:
            self.extract_workers_data()
        
        for worker_info in self.workers_data:
            row_num = worker_info['row']
            data = worker_info['data']
            images = worker_info['images']
            row_errors = []  # Collect validation errors for this row
            row_warnings = []  # Collect warnings for this row
            
            try:
                # Validate name (required field)
                name = str(data.get('name', '')).strip()
                if not name:
                    row_errors.append({
                        'field': 'Name',
                        'value': '',
                        'error': 'Please provide worker name',
                        'column': 'C',
                        'suggestion': 'Enter full name in column C'
                    })
                
                # Parse basic info
                first_name, last_name = self.parse_name(name)
                
                # Warn if name seems incomplete
                if name and not last_name:
                    row_warnings.append({
                        'field': 'Name',
                        'value': name,
                        'warning': 'Consider adding last name',
                        'column': 'C',
                        'suggestion': f'Current: "{name}" - Try: "{name} [LastName]"'
                    })
                
                # Get or create building
                building = None
                zone = None
                floor_number = None
                if data.get('building'):
                    building_raw = str(data['building']).strip()
                    
                    # Parse zone-building-floor format (e.g., "3H-B14-F1")
                    zone_name = None
                    building_name = building_raw
                    floors_count = 1
                    
                    # Split building name by '-' to extract zone, building, and floor info
                    if '-' in building_raw:
                        parts = building_raw.split('-')
                        if len(parts) >= 3:
                            zone_name = parts[0].strip()  # First part is zone (e.g., "3H")
                            building_name = parts[1].strip()  # Middle part is building (e.g., "B14")
                            # Extract floor number from last part (e.g., 'F1' -> 1)
                            floor_part = parts[-1].strip().upper()
                            if floor_part.startswith('F') and floor_part[1:].isdigit():
                                floors_count = int(floor_part[1:])
                                floor_number = floors_count  # Store specific floor for this worker
                        elif len(parts) == 2:
                            # Handle case like "B14-F1" (no zone)
                            building_name = parts[0].strip()
                            floor_part = parts[1].strip().upper()
                            if floor_part.startswith('F') and floor_part[1:].isdigit():
                                floors_count = int(floor_part[1:])
                                floor_number = floors_count
                    
                    # Try to get existing building by the parsed building name (not the full raw name)
                    try:
                        building = Building.objects.get(name=building_name)
                    except Building.DoesNotExist:
                        
                        # Get or create zone if specified
                        if zone_name:
                            try:
                                zone = Zone.objects.get(name=zone_name)
                            except Zone.DoesNotExist:
                                # Create new zone
                                zone = Zone.objects.create(
                                    name=zone_name,
                                    created_by=created_by
                                )
                        
                        # Create new building with unique code
                        import re
                        # Generate a clean code from parsed building name (not the full raw name)
                        building_code = re.sub(r'[^a-zA-Z0-9\-]', '', building_name)[:20]
                        if not building_code:
                            building_code = f"BLD_{Building.objects.count() + 1}"
                        
                        # Ensure code is unique
                        original_code = building_code
                        counter = 1
                        while Building.objects.filter(code=building_code).exists():
                            building_code = f"{original_code[:17]}_{counter}"
                            counter += 1
                        
                        building = Building.objects.create(
                            name=building_name,  # This is now the parsed building name (e.g., "B14")
                            code=building_code,
                            address=f"Building {building_name}",
                            total_floors=floors_count,  # Extract from building name
                            zone=zone,  # Associate with zone if found
                            created_by=created_by
                        )
                
                # Get or create position
                position = None
                if data.get('position'):
                    position_name = str(data['position']).strip()
                    
                    # Try to get existing position first
                    try:
                        position = Position.objects.get(name=position_name)
                    except Position.DoesNotExist:
                        # Create default department if needed
                        department, created = Department.objects.get_or_create(
                            name='General',
                            defaults={
                                'code': 'GEN',
                                'description': 'General Department',
                                'created_by': created_by
                            }
                        )
                        
                        # Generate position code and level
                        import re
                        position_code = re.sub(r'[^a-zA-Z0-9]', '', position_name)[:20].upper()
                        if not position_code:
                            position_code = f"POS_{Position.objects.count() + 1}"
                        
                        # Ensure code is unique
                        original_code = position_code
                        counter = 1
                        while Position.objects.filter(code=position_code).exists():
                            position_code = f"{original_code[:17]}_{counter}"
                            counter += 1
                        
                        # Assign level based on position name (basic mapping)
                        level = 3  # Default level
                        position_upper = position_name.upper()
                        if any(word in position_upper for word in ['MANAGER', 'HEAD', 'LEAD', 'SUPERVISOR']):
                            level = 2
                        elif any(word in position_upper for word in ['DIRECTOR', 'CEO', 'VP', 'PRESIDENT']):
                            level = 1
                        elif any(word in position_upper for word in ['INTERN', 'TRAINEE', 'JUNIOR']):
                            level = 4
                        
                        position = Position.objects.create(
                            name=position_name,
                            code=position_code,
                            department=department,
                            level=level,
                            description=f"Position: {position_name}",
                            created_by=created_by
                        )
                
                # Determine worker status based on resigned_date
                worker_status = 'active'
                if data.get('resigned_date'):
                    # If there's a resigned date, mark as inactive
                    worker_status = 'inactive'
                
                # Validate passport number duplication
                passport_no = data.get('passport_no')
                if passport_no:
                    formatted_passport = self.format_document_number(passport_no)
                    if formatted_passport:
                        # Check if passport number already exists
                        existing_documents = Document.objects.filter(
                            document_type='passport',
                            document_number=formatted_passport
                        ).select_related('worker')
                        
                        if existing_documents.exists():
                            existing_worker = existing_documents.first().worker
                            row_errors.append({
                                'field': 'Passport Number',
                                'value': passport_no,
                                'error': f'Duplicate - Already assigned to {existing_worker.get_full_name()}',
                                'column': 'L',
                                'duplicate_worker_id': existing_worker.id,
                                'suggestion': 'Check if this is the same worker or verify passport number'
                            })
                        passport_no = formatted_passport  # Use formatted version
                    else:
                        row_warnings.append({
                            'field': 'Passport Number',
                            'value': passport_no,
                            'warning': 'Format may be incorrect',
                            'column': 'L',
                            'suggestion': 'Remove special characters except hyphens'
                        })
                        passport_no = None  # Don't use invalid passport
                
                # Check if we have critical errors before proceeding
                if row_errors:
                    results['error_count'] += 1
                    results['validation_errors'].append({
                        'row': row_num,
                        'worker_name': name or 'Unknown',
                        'errors': row_errors,
                        'warnings': row_warnings
                    })
                    continue  # Skip this worker
                
                # Get or create floor if specified
                floor = None
                if building and floor_number:
                    try:
                        floor = Floor.objects.get(building=building, floor_number=floor_number)
                    except Floor.DoesNotExist:
                        # Create new floor
                        floor = Floor.objects.create(
                            building=building,
                            floor_number=floor_number,
                            name=f"Floor {floor_number}",
                            description=f"Floor {floor_number} in building {building.name}",
                            created_by=created_by
                        )

                # Get zone for worker (required field)
                # If building has a zone, use it; otherwise get first zone or create default
                if zone is None:  # Only set zone if not already set from building creation
                    if building and building.zone:
                        zone = building.zone
                    else:
                        # Get or create a default zone
                        zone, created = Zone.objects.get_or_create(
                            name='Default Zone',
                            defaults={
                                'created_by': created_by
                            }
                        )
                
                # Create worker with correct field names
                worker = Worker.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    sex=str(data.get('sex', '')).upper() if data.get('sex') else '',
                    dob=self.parse_date(data.get('dob')),
                    nationality=self.map_nationality(data.get('nationality')),
                    zone=zone,  # Required field
                    building=building,
                    floor=floor,  # Set the specific floor
                    position=position,
                    # Note: date_joined is auto_now_add, so it's set automatically
                    status=worker_status,
                    created_by=created_by
                )
                
                pass
                
                # Save photo if exists
                if 'B' in images:  # Column B is photo
                    photo_data = images['B']['data']
                    photo_name = f"worker_{worker.id}_photo.jpg"
                    worker.photo.save(photo_name, ContentFile(photo_data), save=True)
                
                # Create documents
                self._create_worker_documents(worker, data, images, created_by)
                
                # Record warnings if any
                if row_warnings:
                    results['warnings'].append({
                        'row': row_num,
                        'worker_name': worker.get_full_name(),
                        'warnings': row_warnings
                    })
                
                results['success_count'] += 1
                results['workers_created'].append({
                    'id': worker.id,
                    'name': worker.get_full_name(),
                    'row': row_num,
                    'warnings_count': len(row_warnings)
                })
                
            except Exception as e:
                results['error_count'] += 1
                # Add system error to row errors
                row_errors.append({
                    'field': 'System',
                    'value': '',
                    'error': f'Unexpected error: {str(e)}',
                    'column': 'N/A'
                })
                results['validation_errors'].append({
                    'row': row_num,
                    'worker_name': name if 'name' in locals() else 'Unknown',
                    'errors': row_errors,
                    'warnings': row_warnings
                })
        
        return results
    
    def _create_worker_documents(self, worker, data: Dict, images: Dict, created_by):
        """Create document records for worker"""
        from zone.models import Document
        
        # Document configurations based on correct column mapping
        # B=photo, K=passport, N=work permit, P=visa
        doc_configs = [
            {
                'type': 'passport',
                'number_field': 'passport_no',
                'dates_field': 'passport_dates',
                'file_column': 'K',  # Column K
                'authority': 'Immigration'
            },
            {
                'type': 'work_permit',
                'number_field': None,  # We don't have workpermit_no field
                'dates_field': 'workpermit_dates',
                'file_column': 'N',  # Column N
                'authority': 'Labor Department'
            },
            {
                'type': 'visa',
                'number_field': None,  # We don't have visa_no field
                'dates_field': 'visa_dates',
                'file_column': 'P',  # Column P (corrected from Q)
                'authority': 'Immigration'
            }
        ]

        for config in doc_configs:
            
            # Get document number if field exists
            doc_number = None
            if config['number_field'] and config['number_field'] in data:
                doc_number = self.format_document_number(data.get(config['number_field']))
            
            # Check if we have dates or images for this document type
            dates_data = data.get(config['dates_field'])
            has_image = config['file_column'] in images
            
            # Create document if we have either dates or image (or both)
            if dates_data or has_image or doc_number:
                # Parse dates if available
                issue_date, expiry_date = None, None
                if dates_data:
                    issue_date, expiry_date = self.parse_date_range(dates_data)
                
                try:
                    # Generate document number if not provided
                    if not doc_number:
                        doc_number = f"{config['type'].upper()}_{worker.id}_{int(datetime.now().timestamp())}"
                    
                    document = Document.objects.create(
                        worker=worker,
                        document_type=config['type'],
                        document_number=doc_number,
                        issue_date=issue_date,
                        expiry_date=expiry_date,
                        issuing_authority=config['authority'],
                        created_by=created_by
                    )
                    
                    # Save document image if exists
                    if has_image:
                        try:
                            image_info = images[config['file_column']]
                            file_data = image_info['data']

                            if file_data and len(file_data) > 0:
                                file_name = f"worker_{worker.id}_{config['type']}.jpg"

                                # Save the file to the correct field
                                document.document_file.save(file_name, ContentFile(file_data), save=True)

                        except Exception as file_error:
                            pass

                except Exception as e:
                    pass


def import_workers_from_excel(file_path: str, created_by) -> Dict[str, Any]:
    """
    Main import function - clean interface for views
    """
    importer = ExcelWorkerImporter(file_path)
    return importer.import_to_database(created_by)