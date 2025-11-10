# import cv2  # Temporarily commented out for Docker startup
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
import numpy as np
from core.decorators import manager_or_admin_required
from django.contrib import messages
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.db.models import Q, Count, Avg, Case, When, F
from django.http import JsonResponse, HttpResponse, Http404
from django.views.decorators.http import require_http_methods
from django.urls import reverse


from .models import Zone, Worker, Building, Floor, Document, WorkerProbationPeriod, WorkerProbationExtension, ProbationExtensionRequest, Position, Department, WorkerAssignment
from .utils.progress_tracker import ImportProgressTracker, ProgressStages, STAGE_MESSAGES

from user_management.models import UserRoleAssignment
from .models import Zone, Worker, Building, Floor, Document, WorkerProbationPeriod, WorkerProbationExtension, Position, Department, WorkerAssignment

from .forms import (UploadExcelForm, WorkerForm, WorkerWithDocumentsForm, WorkerDocumentFormSet, WorkerSearchForm, 
                     BuildingForm, FloorForm, ZoneForm, DocumentForm, PositionForm, DepartmentForm,
                     WorkerProbationPeriodForm, WorkerProbationExtensionForm, ProbationSearchForm, WorkerProbationTerminationForm,
                     ProbationExtensionRequestForm, ProbationExtensionRequestReviewForm, WorkerExcelImportForm, BatchProbationForm)

# Import the new views for web-based import
from .views_import import (
    import_workers_view,
    import_workers_preview,
    import_workers_results,
    import_workers_ajax,
    download_template
)
from django.contrib.auth.models import User
from django.db import transaction
from django.utils import timezone
from datetime import timedelta, datetime
from django import forms
import os
from django.conf import settings
import pandas as pd
import tempfile
import os
from django.views.decorators.csrf import csrf_exempt
import logging
import json
import openpyxl
import zipfile
from datetime import datetime
from django.core.files.base import ContentFile
from openpyxl.styles import Font, PatternFill, Alignment
from django.template.loader import render_to_string
from django.db import IntegrityError
from django.core.exceptions import ValidationError
from core.models import get_user_type
from core.file_encryption import FileEncryptionHandler
from django.urls import reverse
from auditlog.models import LogEntry
from django.contrib.contenttypes.models import ContentType

from PIL import Image
import io

logger = logging.getLogger(__name__)

@login_required
def zone_dashboard(request):
    """Dashboard view for zone management."""
    # Show zones created by the current user
    zones = Zone.objects.filter(created_by=request.user)
    total_workers = Worker.objects.filter(zone__created_by=request.user).count()
    
    context = {
        'zones': zones,
        'total_workers': total_workers,
        'page_title': 'Zone Dashboard',
    }
    return render(request, 'zone/dashboard.html', context)

@login_required
def zone_profile(request):
    """View for zone profile management."""
    # Since zones are no longer tied to users, show all zones the user created
    zones = Zone.objects.filter(created_by=request.user)
    
    context = {
        'zones': zones,
        'page_title': 'Zone Profile',
    }
    return render(request, 'zone/profile.html', context)


@login_required
def zone_list(request):
    """List all zones with search and filter capabilities."""
    search_query = request.GET.get('search', '')
    building_filter = request.GET.get('building', '')
    floor_filter = request.GET.get('floor', '')
    status_filter = request.GET.get('status', '')
    
    zones = Zone.objects.select_related('created_by').annotate(
        worker_count=Count('workers', distinct=True),
        building_count=Count('buildings', distinct=True)
    )
    
    # Apply search
    if search_query:
        zones = zones.filter(
            Q(name__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(created_by__first_name__icontains=search_query) |
            Q(created_by__last_name__icontains=search_query)
        )
    
    # Apply filters - now filter by buildings that belong to the zone
    if building_filter:
        zones = zones.filter(buildings__id=building_filter)
    
    if floor_filter:
        zones = zones.filter(buildings__floors__id=floor_filter)
    
    if status_filter:
        if status_filter == 'active':
            zones = zones.filter(is_active=True)
        elif status_filter == 'inactive':
            zones = zones.filter(is_active=False)
    
    # Ensure consistent ordering for pagination
    zones = zones.order_by('name')
    
    # Pagination
    paginator = Paginator(zones, 20)  # Show 20 zones per page
    page_number = request.GET.get('page')
    zones = paginator.get_page(page_number)
    
    # Get data for filters
    buildings = Building.objects.filter(is_active=True).values('id', 'name')
    floors = Floor.objects.filter(is_active=True).select_related('building').values(
        'id', 'name', 'building__name'
    )
    
    # Get statistics
    total_zones = Zone.objects.count()
    active_zones = Zone.objects.filter(is_active=True).count()
    total_workers = Worker.objects.count()
    total_buildings = Building.objects.count()
    
    context = {
        'zones': zones,
        'page_obj': zones,  # Template expects page_obj for pagination
        'search_query': search_query,
        'building_filter': building_filter,
        'floor_filter': floor_filter,
        'status_filter': status_filter,
        'buildings': buildings,
        'floors': floors,
        'total_zones': total_zones,
        'active_zones': active_zones,
        'total_workers': total_workers,
        'total_buildings': total_buildings,
        'page_title': 'Zone Management'
    }
    return render(request, 'zone/zone_list.html', context)

@login_required
def zone_detail(request, zone_id):
    """View zone details with their buildings."""
    zone = get_object_or_404(Zone, id=zone_id)
    buildings = Building.objects.filter(zone=zone).annotate(
        worker_count=Count('workers', distinct=True)
    )
    workers = Worker.objects.filter(zone=zone)
    
    # Calculate building statistics
    building_stats = {
        'total': buildings.count(),
        'active': buildings.filter(is_active=True).count(),
    }
    
    # Calculate worker statistics
    worker_stats = {
        'total': workers.count(),
        'active': workers.filter(status='active').count(),
    }
    
    context = {
        'zone': zone,
        'buildings': buildings,
        'building_stats': building_stats,
        'worker_stats': worker_stats,
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
        'page_title': f'Zone Details - {zone.name}'
    }
    return render(request, 'zone/zone_detail.html', context)

@login_required
@permission_required('zone.add_zone', raise_exception=True)
def zone_create(request):
    """Create a new zone."""
    if request.method == 'POST':
        form = ZoneForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                zone = form.save(commit=False)
                zone.created_by = request.user
                zone.save()
                messages.success(request, f'Zone "{zone.name}" created successfully.')
                return redirect('zone:zone_detail', zone_id=zone.id)
    else:
        form = ZoneForm()
    
    context = {
        'form': form,
        'title': 'Create Zone',
        'submit_text': 'Create Zone',
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
        'page_title': 'Create New Zone'
    }
    return render(request, 'zone/zone_form.html', context)

@login_required
def zone_edit(request, zone_id):
    """Edit an existing zone."""
    zone = get_object_or_404(Zone, id=zone_id)
    
    if request.method == 'POST':
        form = ZoneForm(request.POST, instance=zone)
        if form.is_valid():
            zone = form.save()
            messages.success(request, f'Zone "{zone.name}" updated successfully.')
            return redirect('zone:zone_detail', zone_id=zone.id)
    else:
        form = ZoneForm(instance=zone)
    
    context = {
        'form': form,
        'zone': zone,
        'title': f'Edit Zone - {zone.name}',
        'submit_text': 'Update Zone',
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
        'page_title': f'Edit Zone - {zone.name}'
    }
    return render(request, 'zone/zone_form.html', context)

@login_required
def zone_delete(request, zone_id):
    """Delete a zone and all related data."""
    zone = get_object_or_404(Zone, id=zone_id)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                zone_name = zone.name
                
                # Get counts for reporting
                worker_count = Worker.objects.filter(zone=zone).count()
                building_count = Building.objects.filter(zone=zone).count()
                assignment_count = WorkerAssignment.objects.filter(zone=zone).count()
                
                # Additional related data counts for reporting
                document_count = 0
                probation_count = 0
                extension_count = 0
                certificate_count = 0
                invoice_count = 0
                visa_service_count = 0
                
                # Get related data counts before deletion
                workers = Worker.objects.filter(zone=zone)
                for worker in workers:
                    document_count += worker.documents.count()
                    probation_count += worker.probation_periods.count()
                    extension_count += worker.extension_requests.count()
                    certificate_count += worker.certificate_requests.count()
                    invoice_count += worker.invoices.count()
                    visa_service_count += worker.visa_services.count()
                
                # Delete the zone - Django will handle CASCADE relationships automatically
                # This will delete:
                # - All Workers in this zone (CASCADE)
                #   - Which cascades to delete all Documents for those workers
                #   - Which cascades to delete all WorkerProbationPeriods for those workers
                #     - Which cascades to delete all WorkerProbationExtensions
                #   - Which cascades to delete all ExtensionRequests for those workers
                #   - Which cascades to delete all Invoices for those workers (optional FK)
                #   - Which cascades to delete all VisaServiceRecords for those workers (optional FK)
                # - All WorkerAssignments in this zone (CASCADE)
                # - Buildings in this zone will have their zone set to NULL (SET_NULL)
                zone.delete()
                
                # Create detailed success message
                delete_summary = []
                if worker_count > 0:
                    delete_summary.append(f"{worker_count} workers")
                if document_count > 0:
                    delete_summary.append(f"{document_count} documents")
                if probation_count > 0:
                    delete_summary.append(f"{probation_count} probation periods")
                if extension_count > 0:
                    delete_summary.append(f"{extension_count} extension requests")
                if certificate_count > 0:
                    delete_summary.append(f"{certificate_count} certificate requests")
                if invoice_count > 0:
                    delete_summary.append(f"{invoice_count} invoices")
                if visa_service_count > 0:
                    delete_summary.append(f"{visa_service_count} visa service records")
                if assignment_count > 0:
                    delete_summary.append(f"{assignment_count} worker assignments")
                if building_count > 0:
                    delete_summary.append(f"{building_count} buildings (unlinked from zone)")
                
                success_message = f'Zone "{zone_name}" deleted successfully.'
                if delete_summary:
                    success_message += f' Also deleted: {", ".join(delete_summary)}.'
                
                messages.success(request, success_message)
                
            return redirect('zone:zone_list')
                
        except Exception as e:
            messages.error(request, f'Error deleting zone: {str(e)}')
            return redirect('zone:zone_detail', zone_id=zone.id)
    
    # For GET requests, redirect to zone list (shouldn't happen with modal)
    return redirect('zone:zone_list')

# ============================================================================
# WORKERS MANAGEMENT VIEWS
# ============================================================================

@login_required
@permission_required('zone.view_worker', raise_exception=True)
def worker_list(request):
    """List all workers with search, filter, and pagination."""
    
    # Check for import results in session
    import_results = None
    if 'import_results' in request.session:
        import_results = request.session.pop('import_results')  # Get and remove from session
    
    # Prepare form data with defaults if not provided
    from datetime import datetime, timedelta
    form_data = request.GET.copy()
    
    # Set default date range if not provided in GET parameters
    if 'date_joined_from' not in request.GET and 'date_joined_to' not in request.GET:
        today = datetime.now().date()
        one_month_ago = today - timedelta(days=30)
        form_data['date_joined_from'] = one_month_ago.strftime('%Y-%m-%d')
        form_data['date_joined_to'] = today.strftime('%Y-%m-%d')
    
    # Get search form with potentially modified data
    search_form = WorkerSearchForm(form_data, request=request)
    
    # Start with all workers - including ID card data for the new column
    workers = Worker.objects.select_related('zone', 'building', 'floor').prefetch_related('id_cards').all()
    
    # Apply search filters
    try:
        if search_form.is_valid():
            search_query = search_form.cleaned_data.get('search')
            position = search_form.cleaned_data.get('position')
            zone = search_form.cleaned_data.get('zone')
            building = search_form.cleaned_data.get('building')
            status = search_form.cleaned_data.get('status')
            worker_type = search_form.cleaned_data.get('worker_type')
            sex = search_form.cleaned_data.get('sex')
            nationality = search_form.cleaned_data.get('nationality')
            date_joined_from = search_form.cleaned_data.get('date_joined_from')
            date_joined_to = search_form.cleaned_data.get('date_joined_to')
            
            if search_query:
                # Search in non-encrypted fields only
                workers = workers.filter(
                    Q(first_name__icontains=search_query) |
                    Q(last_name__icontains=search_query) |
                    Q(worker_id__icontains=search_query) |
                    Q(nickname__icontains=search_query)
                )
                
            if position:
                workers = workers.filter(position=position)
                
            if zone:
                workers = workers.filter(zone=zone)
                
            if building:
                workers = workers.filter(building=building)
                
            if status:
                workers = workers.filter(status=status)
                
            if worker_type:
                if worker_type == 'vip':
                    workers = workers.filter(is_vip=True)
                elif worker_type == 'worker':
                    workers = workers.filter(is_vip=False)
                
            if sex:
                workers = workers.filter(sex=sex)
                
            if nationality:
                workers = workers.filter(nationality__icontains=nationality)
            
            # Apply date range filter
            if date_joined_from:
                workers = workers.filter(date_joined__gte=date_joined_from)
            
            if date_joined_to:
                workers = workers.filter(date_joined__lte=date_joined_to)
    except Exception as e:
        # If there's any error with filtering, log it and continue with unfiltered results
        import logging
        logger = logging.getLogger(__name__)
        # search_form will show validation errors if any
    
    # Sorting
    sort_by = request.GET.get('sort', 'worker_id')
    order = request.GET.get('order', 'asc')
    
    valid_sort_fields = ['worker_id', 'first_name', 'last_name', 'zone__name', 
                        'position__name', 'status', 'date_joined', 'created_at']
    
    if sort_by in valid_sort_fields:
        if order == 'desc':
            sort_by = f'-{sort_by}'
        workers = workers.order_by(sort_by)
    else:
        workers = workers.order_by('worker_id')
    
    # Pagination
    per_page = request.GET.get('per_page', '20')
    try:
        per_page = int(per_page)
        if per_page not in [10, 20, 25, 50, 100, 200]:
            per_page = 20
    except (ValueError, TypeError):
        per_page = 20
    
    paginator = Paginator(workers, per_page)
    page = request.GET.get('page')
    
    try:
        workers_page = paginator.page(page)
    except PageNotAnInteger:
        workers_page = paginator.page(1)
    except EmptyPage:
        workers_page = paginator.page(paginator.num_pages)
    
    # Calculate additional stats for template
    active_count = Worker.objects.filter(status='active').count()
    vip_count = Worker.objects.filter(is_vip=True).count()
    regular_worker_count = Worker.objects.filter(is_vip=False).count()
    zones_count = Zone.objects.count()
    buildings_count = Building.objects.count()

    # Check if we're viewing VIP workers only
    is_vip_view = (search_form.is_valid() and 
                   search_form.cleaned_data.get('worker_type') == 'vip')
    
    context = {
        'workers': workers_page,
        'search_form': search_form,
        'current_sort': request.GET.get('sort', 'worker_id'),
        'current_order': request.GET.get('order', 'asc'),
        'total_count': paginator.count,
        'per_page': per_page,
        'active_count': active_count,
        'vip_count': vip_count,
        'regular_worker_count': regular_worker_count,
        'zones_count': zones_count,
        'buildings_count': buildings_count,
        'is_vip_view': is_vip_view,
        'import_results': import_results,  # Pass import results to template
    }
    
    return render(request, 'zone/worker_list.html', context)

@login_required
@permission_required('zone.view_worker', raise_exception=True)
def worker_detail(request, worker_id):
    """View worker details with documents and audit log."""
    worker = get_object_or_404(Worker, id=worker_id)
    # Use consistent document filtering with the edit form
    from .forms import WorkerDocumentFormSet
    formset = WorkerDocumentFormSet(instance=worker)
    # Only show documents that exist (have files) to match edit form template logic
    documents = worker.documents.filter(document_file__isnull=False).exclude(document_file='').order_by('created_at')
    
    # Get audit log data for the worker
    from auditlog.models import LogEntry
    from django.contrib.contenttypes.models import ContentType
    
    content_type = ContentType.objects.get_for_model(Worker)
    audit_logs = LogEntry.objects.filter(
        content_type=content_type,
        object_id=worker.id
    ).select_related('actor').order_by('-timestamp')[:50]  # Limit to 50 recent entries
    
    # Process audit logs to add display information
    all_logs = []
    has_audit_logs = False
    
    for log in audit_logs:
        has_audit_logs = True
        log_entry = {
            'log_type': 'change_log',
            'action': log.action,
            'timestamp': log.timestamp,
            'actor': log.actor,
            'changes': log.changes or {},
            'filtered_changes': {}
        }
        
        # Filter out excluded encrypted fields from display
        excluded_fields = [
            'phone_number', 'email', 'address', 'emergency_phone',
            'passport_number', 'id_card_number', 'notes'
        ]
        
        if log.changes:
            for field, change in log.changes.items():
                if field not in excluded_fields:
                    log_entry['filtered_changes'][field] = change
        
        # Add method to get action display
        def get_action_display():
            action_map = {
                0: 'Created',
                1: 'Updated', 
                2: 'Deleted'
            }
            return action_map.get(log.action, 'Modified')
        
        log_entry['get_action_display'] = get_action_display()
        all_logs.append(log_entry)
    
    context = {
        'worker': worker,
        'documents': documents,
        'formset': formset,  # Include formset for consistency with edit template structure
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
        'all_logs': all_logs,
        'has_audit_logs': has_audit_logs,
    }
    return render(request, 'zone/worker_detail.html', context)

@login_required
@permission_required('zone.add_worker', raise_exception=True)
def worker_create(request):
    """Create a new worker with documents."""
    
    form_action = "not_submit"
    if request.method == 'POST':
        form_action = "submited"
        
        # Log received data (excluding sensitive info)
        form = WorkerForm(request.POST, request.FILES)
        formset = WorkerDocumentFormSet(request.POST, request.FILES)
        # Pass request to formset for user feedback about document replacement
        formset._request = request
        
        form_valid = form.is_valid()
        formset_valid = formset.is_valid()
        if form_valid and formset_valid:
            try:
                with transaction.atomic():
                    # Save the worker
                    worker = form.save(commit=False)
                    if request.user.is_authenticated:
                        worker.created_by = request.user
                        worker._created_by = request.user  # For automatic probation creation
                    worker.save()
                    
                    # Save other documents
                    formset.instance = worker
                    formset.save()
                    
                    messages.success(request, f'Worker {worker.get_full_name()} created successfully with 15-day probation period!')
                    return redirect('zone:worker_detail', worker_id=worker.id)
            except Exception as e:
                messages.error(request, f'Error creating worker: {str(e)}')
    else:
        form_action = "not_submit"
        form = WorkerForm()
        formset = WorkerDocumentFormSet()
    
    # Get data for chain selects
    zones = Zone.objects.filter(is_active=True).select_related('created_by')
    buildings = Building.objects.filter(is_active=True)

    
    context = {
        'form': form,
        'formset': formset,
        'title': 'Create New Entry',
        'submit_text': 'Save',
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
        'zones_data': [{'id': zone.id, 'name': zone.name} for zone in zones],
        'buildings_data': [{'id': building.id, 'name': building.name, 'zone_id': building.zone_id} for building in buildings],
        'form_action': form_action,
        'is_edit': 'create'  # Explicitly set to 'create' to ensure template logic works correctly
    }
    return render(request, 'zone/worker_form_create.html', context)

@login_required
@permission_required('zone.change_worker', raise_exception=True)
def worker_edit(request, worker_id):
    """Edit an existing worker with documents."""
    worker = get_object_or_404(Worker, id=worker_id)

    if request.method == 'POST':
        form = WorkerForm(request.POST, request.FILES, instance=worker)
        # Disable status field in edit mode - status should be changed through workflows
        if 'status' in form.fields:
            form.fields['status'].disabled = True
        
        formset = WorkerDocumentFormSet(request.POST, request.FILES, instance=worker)
        # Pass request to formset for user feedback about document replacement
        formset._request = request
        
        # Validate form and formset
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    # Track what changed for better feedback
                    changed_fields = []
                    if form.changed_data:
                        for field_name in form.changed_data:
                            # Exclude status from change tracking since it's disabled
                            if field_name in ['zone', 'building', 'floor', 'position']:
                                old_value = getattr(worker, field_name)
                                new_value = form.cleaned_data.get(field_name)
                                if old_value != new_value:
                                    changed_fields.append(f"{field_name.title()}: {old_value} → {new_value}")
                    
                    # Save the worker (status won't be changed since field is disabled)
                    saved_worker = form.save()
                    
                    # Save the documents
                    formset.instance = saved_worker
                    formset.save()
                    
                    # Provide detailed success feedback
                    if changed_fields:
                        changes_text = ", ".join(changed_fields)
                        success_message = f'Worker {saved_worker.get_full_name()} updated successfully! Changes: {changes_text}'
                    else:
                        success_message = f'Worker {saved_worker.get_full_name()} updated successfully!'
                    
                    messages.success(request, success_message)
                    return redirect('zone:worker_detail', worker_id=saved_worker.id)
                    
            except Exception as e:
                messages.error(request, f'Failed to save worker: {str(e)}')
        else:
            # Handle validation errors
            if not form.is_valid():
                error_messages = []
                for field, errors in form.errors.items():
                    for error in errors:
                        if field == '__all__':
                            error_messages.append(f'Form error: {error}')
                        else:
                            error_messages.append(f'{field.title()}: {error}')
                
                if error_messages:
                    messages.error(request, 'Please correct the following errors: ' + '; '.join(error_messages))
            
            if not formset.is_valid():
                # Create detailed error message for user
                error_details = []
                for i, form_errors in enumerate(formset.errors):
                    if form_errors:
                        for field, errors in form_errors.items():
                            for error in errors:
                                error_details.append(f"Document {i+1} - {field}: {error}")
                
                if formset.non_form_errors():
                    for error in formset.non_form_errors():
                        error_details.append(f"Document formset: {error}")
                
                if error_details:
                    detailed_message = 'Document errors: ' + '; '.join(error_details)
                    messages.error(request, detailed_message)
                else:
                    messages.error(request, 'Please check the document information for errors.')
            
            # Don't recreate forms - keep the existing ones with POST data and errors
    else:
        # GET request - clear any preserved files from previous attempts
        clear_preserved_files_from_session(request)
        
        form = WorkerForm(instance=worker)
        # Disable status field in edit mode - status should be changed through workflows
        if 'status' in form.fields:
            form.fields['status'].disabled = True
            form.fields['status'].help_text = 'Status cannot be changed directly. Use probation workflows or other appropriate actions to change worker status.'
        
        formset = WorkerDocumentFormSet(instance=worker)
    
    # Get data for chain selects
    zones = Zone.objects.filter(is_active=True).select_related('created_by')
    buildings = Building.objects.filter(is_active=True)
    
    # Ensure worker is never None (safety check)
    if not worker:
        worker = get_object_or_404(Worker, id=worker_id)
    
    context = {
        'form': form,
        'formset': formset,
        'worker': worker,
        'title': f'Edit Worker: {worker.get_full_name()}',
        'submit_text': 'Update Worker',
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
        'zones_data': [{'id': zone.id, 'name': zone.name} for zone in zones],
        'buildings_data': [{'id': building.id, 'name': building.name, 'zone_id': building.zone_id} for building in buildings],
        'is_edit': 'edit'
    }
    return render(request, 'zone/worker_form_create.html', context)

@login_required
@permission_required('zone.delete_worker', raise_exception=True)
def worker_delete(request, worker_id):
    """Delete a worker."""
    worker = get_object_or_404(Worker, id=worker_id)
    
    if request.method == 'POST':
        worker_name = worker.get_full_name()
        
        try:
            # Log deletion before deleting the worker
            
            # Check related records before deletion
            change_history_count = worker.change_history.count()
            documents_count = worker.documents.count()
            assignments_count = worker.assignments.count()
            probation_count = worker.probation_periods.count() if hasattr(worker, 'probation_periods') else 0
            
            # Collect files to delete before deletion
            files_to_delete = []
            
            # Collect worker photo file
            if worker.photo and worker.photo.name:
                try:
                    files_to_delete.append(worker.photo.path)
                except Exception:
                    pass  # File might not exist
            
            # Collect document files
            for document in worker.documents.all():
                if document.document_file and document.document_file.name:
                    try:
                        files_to_delete.append(document.document_file.path)
                    except Exception:
                        pass  # File might not exist
            
            # Set a flag to disable signals during worker deletion
            import threading
            if not hasattr(threading.current_thread(), 'deleting_worker'):
                threading.current_thread().deleting_worker = True
                
            # Log the worker deletion BEFORE deleting (since post_delete signal was removed)
            from .models import WorkerChangeHistory
            try:
                WorkerChangeHistory.objects.create(
                    worker=worker,
                    action='deleted',
                    description=f'Worker {worker_name} ({worker.worker_id}) deleted by {request.user.get_full_name() or request.user.username}',
                    changed_by=request.user,
                    ip_address=request.META.get('REMOTE_ADDR')
                )
            except Exception as log_error:
                pass
                
            try:
                # Use Django's built-in cascade delete - should handle all related records automatically
                deleted_info = worker.delete()
                
                # Clean up files from filesystem
                import os
                import logging
                logger = logging.getLogger(__name__)
                
                files_deleted = 0
                for file_path in files_to_delete:
                    try:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                            files_deleted += 1
                    except Exception as e:
                        pass
                
                
            finally:
                # Clear the flag after deletion
                if hasattr(threading.current_thread(), 'deleting_worker'):
                    delattr(threading.current_thread(), 'deleting_worker')
            
            
            messages.success(request, f'Worker "{worker_name}" and all related records deleted successfully.')
            return redirect('zone:worker_list')
            
        except Exception as e:
            import traceback
            
            # More user-friendly error messages
            if 'foreign key constraint' in str(e).lower():
                messages.error(request, f'Database constraint error: Cannot delete worker "{worker_name}". This may be due to database inconsistency. Please contact system administrator.')
            elif 'violates foreign key constraint' in str(e).lower():
                messages.error(request, f'Foreign key constraint violation: Worker "{worker_name}" has related records that prevent deletion. Please contact system administrator.')
            else:
                messages.error(request, f'Error deleting worker "{worker_name}": {str(e)}')
            
            # Stay on the confirmation page to show the error
            context = {
                'worker': worker,
                'total_zones': Zone.objects.count(),
                'total_workers': Worker.objects.count(),
                'change_history_count': worker.change_history.count(),
                'documents_count': worker.documents.count(),
                'assignments_count': worker.assignments.count(),
                'error_message': str(e),
            }
            return render(request, 'zone/worker_confirm_delete.html', context)
    
    # Get related records count for confirmation page
    change_history_count = worker.change_history.count()
    documents_count = worker.documents.count()
    assignments_count = worker.assignments.count()
    
    context = {
        'worker': worker,
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
        'change_history_count': change_history_count,
        'documents_count': documents_count,
        'assignments_count': assignments_count,
    }
    return render(request, 'zone/worker_confirm_delete.html', context)


@login_required
@manager_or_admin_required("Worker Management")
def bulk_delete_workers(request):
    """Bulk delete multiple workers."""
    if request.method == 'POST':
        worker_ids = request.POST.getlist('worker_ids[]')
        
        if not worker_ids:
            return JsonResponse({'success': False, 'message': 'No workers selected for deletion.'})
        
        try:
            # Get workers to delete with related data
            workers = Worker.objects.filter(id__in=worker_ids).prefetch_related('documents')
            worker_count = workers.count()
            
            if worker_count == 0:
                return JsonResponse({'success': False, 'message': 'No valid workers found to delete.'})
            
            # Set a flag to disable signals during bulk deletion
            import threading
            if not hasattr(threading.current_thread(), 'deleting_worker'):
                threading.current_thread().deleting_worker = True
            
            try:
                # Collect all files to delete and log deletion for each worker
                files_to_delete = []
                from .models import WorkerChangeHistory
                
                for worker in workers:
                    # Collect worker photo file
                    if worker.photo and worker.photo.name:
                        files_to_delete.append(worker.photo.path)
                    
                    # Collect document files
                    for document in worker.documents.all():
                        if document.document_file and document.document_file.name:
                            files_to_delete.append(document.document_file.path)
                    
                    # Log deletion
                    try:
                        WorkerChangeHistory.objects.create(
                            worker=worker,
                            action='deleted',
                            description=f'Worker {worker.get_full_name()} ({worker.worker_id}) bulk deleted by {request.user.get_full_name() or request.user.username}',
                            changed_by=request.user,
                            ip_address=request.META.get('REMOTE_ADDR')
                        )
                    except Exception:
                        pass  # Continue even if logging fails
                
                # Delete all workers (this will cascade delete documents)
                workers.delete()
                
                # Clean up files from filesystem
                import os
                import logging
                logger = logging.getLogger(__name__)
                
                files_deleted = 0
                for file_path in files_to_delete:
                    try:
                        if os.path.exists(file_path):
                            os.remove(file_path)
                            files_deleted += 1
                    except Exception as e:
                        pass
                
                
            finally:
                # Clear the flag after deletion
                if hasattr(threading.current_thread(), 'deleting_worker'):
                    delattr(threading.current_thread(), 'deleting_worker')
            
            return JsonResponse({
                'success': True, 
                'message': f'Successfully deleted {worker_count} worker(s).'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'Error during bulk deletion: {str(e)}'
            })
    
    return JsonResponse({'success': False, 'message': 'Invalid request method.'})


@login_required
def worker_probation_status(request, worker_id):
    """View all probation periods for a specific worker."""
    worker = get_object_or_404(
        Worker.objects.select_related(
            'zone', 'zone__created_by', 'building', 'position'
        ),
        id=worker_id
    )
    
    # Get all probation periods for this worker
    probations = WorkerProbationPeriod.objects.filter(
        worker=worker
    ).select_related('created_by').prefetch_related('extensions').order_by('-start_date')
    
    # Calculate statistics
    total_probations = probations.count()
    active_probations = probations.filter(worker__status__in=['probation', 'extended']).count()
    completed_probations = probations.filter(worker__status='passed').count()
    failed_probations = probations.filter(worker__status='failed').count()
    
    # Total extension days
    total_extension_days = 0
    for probation in probations:
        total_extension_days += sum(ext.extension_duration_days for ext in probation.extensions.all())
    
    context = {
        'worker': worker,
        'probations': probations,
        'probation_periods': probations,  # Add correct template variable
        'total_probations': total_probations,
        'active_probations': active_probations,
        'completed_probations': completed_probations,
        'failed_probations': failed_probations,
        'total_extension_days': total_extension_days,
    }
    return render(request, 'zone/worker_probation_status.html', context)

@login_required
def worker_probation_extensions(request, worker_id):
    """View all probation extensions for a specific worker."""
    worker = get_object_or_404(
        Worker.objects.select_related(
            'zone', 'zone__created_by', 'building', 'position'
        ),
        id=worker_id
    )
    
    # Get all extensions for this worker's probations
    extensions = WorkerProbationExtension.objects.filter(
        probation_period__worker=worker
    ).select_related(
        'probation_period', 'approved_by', 'created_by'
    ).order_by('-created_at')
    
    # Apply filters
    status_filter = request.GET.get('status')
    date_filter = request.GET.get('date_range')
    
    if status_filter:
        extensions = extensions.filter(probation_period__worker__status=status_filter)
    
    if date_filter:
        from datetime import timedelta
        today = timezone.now().date()
        
        if date_filter == 'last_30':
            extensions = extensions.filter(created_at__gte=today - timedelta(days=30))
        elif date_filter == 'last_90':
            extensions = extensions.filter(created_at__gte=today - timedelta(days=90))
        elif date_filter == 'this_year':
            extensions = extensions.filter(created_at__year=today.year)
    
    # Calculate statistics
    total_extensions = extensions.count()
    total_extension_days = sum(ext.extension_duration_days for ext in extensions)
    # Count active probations for this worker (check worker status directly)
    active_probations = WorkerProbationPeriod.objects.filter(
        worker=worker
    ).count() if worker.status in ['probation', 'extended'] else 0
    
    # Pagination
    paginator = Paginator(extensions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'worker': worker,
        'extensions': page_obj,
        'total_extensions': total_extensions,
        'total_extension_days': total_extension_days,
        'active_probations': active_probations,
        'status_filter': status_filter,
        'date_filter': date_filter,
    }
    return render(request, 'zone/worker_probation_extensions.html', context)

# ============================================================================
# BUILDING MANAGEMENT VIEWS
# ============================================================================

@login_required
def building_list(request):
    """List all buildings with search, sorting, and filtering."""
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    sort_by = request.GET.get('sort', 'name')
    
    buildings = Building.objects.annotate(
        floor_count=Count('floors'),
        worker_count=Count('workers')
    )
    
    # Apply search filter
    if search_query:
        buildings = buildings.filter(
            Q(name__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Apply status filter
    if status_filter:
        buildings = buildings.filter(is_active=(status_filter == 'active'))
    
    # Apply sorting
    valid_sorts = ['name', '-name', 'created_at', '-created_at', 'total_floors', '-total_floors']
    if sort_by in valid_sorts:
        buildings = buildings.order_by(sort_by)
    else:
        buildings = buildings.order_by('name')
    
    # Pagination
    paginator = Paginator(buildings, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'total_buildings': Building.objects.count(),
        'active_buildings': Building.objects.filter(is_active=True).count(),
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
    }
    return render(request, 'zone/building_list.html', context)

@login_required
def building_detail(request, building_id):
    """View building details with floors, zones, and workers."""
    building = get_object_or_404(Building, id=building_id)
    floors = Floor.objects.filter(building=building).annotate(
        worker_count=Count('workers')
    ).order_by('floor_number')
    zones = []  # Building no longer has zones - zones have buildings
    if building.zone:
        zones = [building.zone]  # Show the zone this building belongs to
    workers = Worker.objects.filter(building=building).select_related('zone', 'floor')
    
    # Get statistics
    stats = {
        'total_floors': floors.count(),
        'active_floors': floors.filter(is_active=True).count(),
        'total_zones': len(zones),
        'active_zones': len([z for z in zones if z.is_active]) if zones else 0,
        'total_workers': workers.count(),
        'active_workers': workers.filter(status='active').count(),
    }
    
    context = {
        'building': building,
        'floors': floors,
        'zones': zones,  # Show the zone this building belongs to
        'workers': workers[:10],  # Show first 10
        'stats': stats,
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
    }
    return render(request, 'zone/building_detail.html', context)

@login_required
def building_create(request):
    """Create a new building."""
    if request.method == 'POST':
        form = BuildingForm(request.POST)
        if form.is_valid():
            building = form.save(commit=False)
            building.created_by = request.user
            building.save()
            
            # If total_floors is 0, create a default floor
            if building.total_floors == 0:
                Floor.objects.create(
                    building=building,
                    floor_number=0,
                    name='F0',
                    description='no floor',
                    created_by=request.user
                )
            
            messages.success(request, f'Building "{building.name}" created successfully.')
            return redirect('zone:building_detail', building_id=building.id)
    else:
        form = BuildingForm()
    
    context = {
        'form': form,
        'title': 'Create Building',
        'submit_text': 'Create Building',
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
    }
    return render(request, 'zone/building_form.html', context)

@login_required
def building_edit(request, building_id):
    """Edit an existing building."""
    building = get_object_or_404(Building, id=building_id)
    
    if request.method == 'POST':
        form = BuildingForm(request.POST, instance=building)
        if form.is_valid():
            building = form.save()
            messages.success(request, f'Building "{building.name}" updated successfully.')
            return redirect('zone:building_detail', building_id=building.id)
    else:
        form = BuildingForm(instance=building)
    
    context = {
        'form': form,
        'building': building,
        'title': f'Edit Building - {building.name}',
        'submit_text': 'Update Building',
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
    }
    return render(request, 'zone/building_form.html', context)

@login_required
def building_delete(request, building_id):
    """Delete a building."""
    building = get_object_or_404(Building, id=building_id)
    
    if request.method == 'POST':
        building_name = building.name
        building.delete()
        messages.success(request, f'Building "{building_name}" deleted successfully.')
        return redirect('zone:building_list')
    
    # Check for dependencies
    floor_count = building.floors.count()
    zone_count = 1 if building.zone else 0  # Building belongs to one zone or none
    worker_count = building.workers.count()
    
    context = {
        'building': building,
        'floor_count': floor_count,
        'zone_count': zone_count,
        'worker_count': worker_count,
        'has_dependencies': floor_count > 0 or worker_count > 0,  # zone dependency removed since building->zone
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
    }
    return render(request, 'zone/building_confirm_delete.html', context)

# ============================================================================
# FLOOR MANAGEMENT VIEWS
# ============================================================================

@login_required
def floor_list(request):
    """List all floors with search, sorting, and filtering."""
    search_query = request.GET.get('search', '')
    building_filter = request.GET.get('building', '')
    status_filter = request.GET.get('status', '')
    sort_by = request.GET.get('sort', 'building__name')
    
    floors = Floor.objects.select_related('building').annotate(
        worker_count=Count('workers')
    )
    
    # Apply search filter
    if search_query:
        floors = floors.filter(
            Q(name__icontains=search_query) |
            Q(building__name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(floor_number__icontains=search_query)
        )
    
    # Apply building filter
    if building_filter:
        floors = floors.filter(building_id=building_filter)
    
    # Apply status filter
    if status_filter:
        floors = floors.filter(is_active=(status_filter == 'active'))
    
    # Apply sorting
    valid_sorts = ['building__name', '-building__name', 'floor_number', '-floor_number', 
                   'name', '-name', 'created_at', '-created_at']
    if sort_by in valid_sorts:
        floors = floors.order_by(sort_by)
    else:
        floors = floors.order_by('building__name', 'floor_number')
    
    # Pagination
    paginator = Paginator(floors, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    buildings = Building.objects.filter(is_active=True).order_by('name')
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'building_filter': building_filter,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'buildings': buildings,
        'total_floors': Floor.objects.count(),
        'active_floors': Floor.objects.filter(is_active=True).count(),
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
    }
    return render(request, 'zone/floor_list.html', context)

@login_required
def floor_detail(request, floor_id):
    """View floor details with zones and workers."""
    floor = get_object_or_404(Floor.objects.select_related('building'), id=floor_id)
    zones = []  # Floors no longer have direct zone relationship
    if floor.building and floor.building.zone:
        zones = [floor.building.zone]  # Show the zone this floor's building belongs to
    workers = Worker.objects.filter(floor=floor).select_related('zone')
    
    # Get statistics
    stats = {
        'total_zones': len(zones),
        'active_zones': len([z for z in zones if z.is_active]) if zones else 0,
        'total_workers': workers.count(),
        'active_workers': workers.filter(status='active').count(),
    }
    
    context = {
        'floor': floor,
        'zones': zones,
        'workers': workers,
        'stats': stats,
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
    }
    return render(request, 'zone/floor_detail.html', context)

@login_required
def floor_create(request):
    """Create a new floor."""
    if request.method == 'POST':
        form = FloorForm(request.POST)
        if form.is_valid():
            floor = form.save(commit=False)
            floor.created_by = request.user
            floor.save()
            messages.success(request, f'Floor "{floor.name}" created successfully.')
            return redirect('zone:floor_detail', floor_id=floor.id)
    else:
        # Pre-select building if provided in URL parameter
        initial_data = {}
        building_id = request.GET.get('building')
        if building_id:
            try:
                building = Building.objects.get(id=building_id, is_active=True)
                initial_data['building'] = building
            except Building.DoesNotExist:
                pass
        
        form = FloorForm(initial=initial_data)
    
    # Get building data for JavaScript
    buildings = Building.objects.filter(is_active=True).values('id', 'name', 'address', 'total_floors')
    
    context = {
        'form': form,
        'title': 'Create Floor',
        'submit_text': 'Create Floor',
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
        'buildings_data': list(buildings),
    }
    return render(request, 'zone/floor_form.html', context)

@login_required
def floor_edit(request, floor_id):
    """Edit an existing floor."""
    floor = get_object_or_404(Floor, id=floor_id)
    
    if request.method == 'POST':
        form = FloorForm(request.POST, instance=floor)
        if form.is_valid():
            floor = form.save()
            messages.success(request, f'Floor "{floor.name}" updated successfully.')
            return redirect('zone:floor_detail', floor_id=floor.id)
    else:
        form = FloorForm(instance=floor)
    
    # Get building data for JavaScript
    buildings = Building.objects.filter(is_active=True).values('id', 'name', 'address', 'total_floors')
    
    context = {
        'form': form,
        'floor': floor,
        'title': f'Edit Floor - {floor.name}',
        'submit_text': 'Update Floor',
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
        'buildings_data': list(buildings),
    }
    return render(request, 'zone/floor_form.html', context)

@login_required
def floor_delete(request, floor_id):
    """Delete a floor."""
    floor = get_object_or_404(Floor, id=floor_id)
    
    if request.method == 'POST':
        floor_name = floor.name
        building = floor.building
        floor.delete()
        messages.success(request, f'Floor "{floor_name}" deleted successfully.')
        return redirect('zone:building_detail', building_id=building.id)
    
    # Check for dependencies
    zone_count = 0  # Floors no longer have direct zone relationships
    worker_count = floor.workers.count()
    
    context = {
        'floor': floor,
        'zone_count': zone_count,
        'worker_count': worker_count,
        'has_dependencies': worker_count > 0,  # Only workers are a dependency now
        'total_zones': Zone.objects.count(),
        'total_workers': Worker.objects.count(),
    }
    return render(request, 'zone/floor_confirm_delete.html', context)

# ============================================================================
# PROBATION VIEWS
# ============================================================================

# Original probation_list function removed - replaced by probation_list_improved from views_probation_simple.py

@login_required
@permission_required('hr.view_probationperiod', raise_exception=True)
def probation_detail(request, probation_id):
    """Redirect to probation list - detail page disabled."""
    messages.info(request, 'Probation details page has been disabled. Redirected to probation list.')
    return redirect('zone:probation_list')

def is_manager(user):
    """Check if user has manager privileges for probation approvals."""
    return user.is_staff or user.is_superuser or user.groups.filter(name='Managers').exists()

@login_required
# @permission_required('hr.change_probationperiod', raise_exception=True)  # Temporarily disabled
def probation_extend(request, probation_id):
    """Extend a probation period with maker-checker workflow."""
    
    probation = get_object_or_404(WorkerProbationPeriod, id=probation_id)
    user_is_manager = is_manager(request.user)
    
    if request.method == 'POST':
        print(f"DEBUG: POST data received: {dict(request.POST)}")
        if user_is_manager:
            print(f"DEBUG: User is manager, using WorkerProbationExtensionForm")
            # Managers can directly extend (legacy flow)
            form = WorkerProbationExtensionForm(request.POST, exclude_probation_field=True)
            print(f"DEBUG: Form data: {form.data}")
            print(f"DEBUG: Form is valid: {form.is_valid()}")
            if not form.is_valid():
                print(f"DEBUG: Form errors: {form.errors}")
            if form.is_valid():
                extension = form.save(commit=False)
                extension.probation_period = probation
                extension.created_by = request.user
                extension.save()
                
                # Update worker status to 'extended' if it was in an active probation state
                worker = probation.worker
                if worker.status in ['probation', 'extended']:
                    worker.status = 'extended'
                    worker.save()
                
                messages.success(request, f'Probation period extended by {extension.extension_duration_days} days.')
                return redirect('zone:probation_list')
        else:
            print(f"DEBUG: User is not manager, using ProbationExtensionRequestForm")
            # Regular users create extension requests
            form = ProbationExtensionRequestForm(request.POST, probation_period=probation)
            print(f"DEBUG: Form data: {form.data}")
            print(f"DEBUG: Form is valid: {form.is_valid()}")
            if not form.is_valid():
                print(f"DEBUG: Form errors: {form.errors}")
            if form.is_valid():
                extension_request = form.save(commit=False)
                extension_request.probation_period = probation
                extension_request.requested_by = request.user
                extension_request.save()
                
                messages.success(request, f'Extension request submitted for approval. Requested duration: {extension_request.extension_duration_days} days.')
                return redirect('zone:probation_list')
    else:
        if user_is_manager:
            # Show direct extension form to managers
            form = WorkerProbationExtensionForm(
                exclude_probation_field=True,
                initial={'approved_by': request.user}
            )
        else:
            # Show extension request form to regular users
            form = ProbationExtensionRequestForm(probation_period=probation)
    
    context = {
        'probation': probation,
        'worker': probation.worker,
        'form': form,
        'user_is_manager': user_is_manager,
        'title': f'{"Extend" if user_is_manager else "Request Extension for"} Probation - {probation.worker.get_full_name()}',
    }
    return render(request, 'zone/probation_extend.html', context)

@login_required
@permission_required('hr.change_probationperiod', raise_exception=True)
def probation_complete(request, probation_id):
    """Complete a probation period."""
    probation = get_object_or_404(WorkerProbationPeriod, id=probation_id)
    
    if request.method == 'POST':
        # Update worker status to 'passed' to indicate successful completion of probation
        probation.worker.status = 'passed'
        probation.worker.save()
        
        probation.actual_end_date = timezone.now().date()
        probation.save()
        
        messages.success(request, f'Probation period completed for {probation.worker.get_full_name()}.')
        return redirect('zone:probation_list')
    
    context = {
        'probation': probation,
        'worker': probation.worker,
        'title': f'Complete Probation - {probation.worker.get_full_name()}',
    }
    return render(request, 'zone/probation_complete.html', context)

@login_required
@permission_required('hr.change_probationperiod', raise_exception=True)
def probation_pass(request, probation_id):
    """Pass a worker after successful probation completion - no confirmation needed."""
    probation = get_object_or_404(WorkerProbationPeriod, id=probation_id)
    
    # Check if probation can be passed
    if probation.worker.status not in ['probation', 'extended']:
        messages.error(request, 'Only workers with probation or extended status can have their probation passed.')
        return redirect('zone:probation_list')
    
    # Directly pass the probation without confirmation
    worker = probation.worker
    worker.status = 'passed'
    worker.save()
    
    # Update probation period details
    probation.actual_end_date = timezone.now().date()
    probation.save()
    
    messages.success(
        request, 
        f'Worker {probation.worker.get_full_name()} has successfully passed probation! '
        f'Worker status has been updated to "Passed".'
    )
    return redirect('zone:probation_list')

@login_required
@permission_required('hr.change_probationperiod', raise_exception=True)
def probation_terminate(request, probation_id):
    """Terminate a worker during probation period."""
    probation = get_object_or_404(WorkerProbationPeriod, id=probation_id)
    
    # Check if probation can be terminated
    if probation.worker.status not in ['probation', 'extended']:
        messages.error(request, 'Only workers with probation or extended status can have their probation terminated.')
        return redirect('zone:probation_list')
    
    if request.method == 'POST':
        form = WorkerProbationTerminationForm(request.POST, instance=probation)
        if form.is_valid():
            # Update worker status to terminated
            probation.worker.status = 'terminated'
            probation.worker.save()
            
            # Update probation period details
            probation.termination_reason = form.cleaned_data['termination_reason']
            probation.terminated_date = form.cleaned_data['terminated_date']
            probation.terminated_by = request.user
            probation.actual_end_date = form.cleaned_data['terminated_date']
            probation.save()
            
            messages.success(
                request, 
                f'Worker {probation.worker.get_full_name()} has been terminated during probation. '
                f'Worker status has been updated to "Terminated".'
            )
            return redirect('zone:probation_list')
    else:
        form = WorkerProbationTerminationForm(instance=probation)
    
    context = {
        'probation': probation,
        'worker': probation.worker,
        'form': form,
        'title': f'Terminate Probation - {probation.worker.get_full_name()}',
    }
    return render(request, 'zone/probation_terminate.html', context)

@login_required
@permission_required('zone.view_workerprobationperiod', raise_exception=True)
def probation_edit(request, probation_id):
    """Edit a probation period."""
    probation = get_object_or_404(WorkerProbationPeriod, id=probation_id)
    
    # Check if user is manager or admin
    user_is_manager = False
    if request.user.is_superuser:
        user_is_manager = True
    else:
        try:
            from user_management.models import UserRoleAssignment
            role_assignment = UserRoleAssignment.objects.get(
                user=request.user,
                is_active=True
            )
            user_is_manager = role_assignment.role.name in ["Manager", "Admin"]
        except UserRoleAssignment.DoesNotExist:
            user_is_manager = False
    
    if request.method == 'POST':
        form = WorkerProbationPeriodForm(request.POST, instance=probation, exclude_worker_field=True)
        
        # If user is not a manager, preserve the original worker_status
        if not user_is_manager and 'worker_status' in form.data:
            # Create a mutable copy of POST data and restore original worker_status
            post_data = request.POST.copy()
            post_data['worker_status'] = probation.worker.status
            form = WorkerProbationPeriodForm(post_data, instance=probation, exclude_worker_field=True)
        
        if form.is_valid():
            updated_probation = form.save()
            
            # Status is now managed only at the Worker level - no syncing needed
            worker = updated_probation.worker
            messages.success(request, f'Probation period for {worker.get_full_name()} has been updated successfully.')
            
            return redirect('zone:probation_list')
    else:
        form = WorkerProbationPeriodForm(instance=probation, exclude_worker_field=True)
    
    context = {
        'form': form,
        'probation': probation,
        'worker': probation.worker,
        'title': f'Edit Probation - {probation.worker.get_full_name()}',
        'user_is_manager': user_is_manager,
    }
    return render(request, 'zone/probation_edit.html', context)

@login_required
@permission_required('hr.delete_probationperiod', raise_exception=True)
def probation_cancel(request, probation_id):
    """Cancel a probation period and reset worker status (keeps record for audit)."""
    from django.utils import timezone
    
    probation = get_object_or_404(WorkerProbationPeriod, id=probation_id)
    
    if request.method == 'POST':
        worker_name = probation.worker.get_full_name()
        worker = probation.worker
        
        # Reset worker status if they were on probation
        if worker.status in ['probation', 'extended']:
            worker.status = 'active'  # Reset to active status
            worker.save()
            
        # Mark probation as cancelled (don't delete the record)
        probation.terminated_date = timezone.now().date()
        probation.terminated_by = request.user
        probation.termination_reason = 'Probation cancelled by administrator'
        probation.save()
        
        messages.success(request, f'Probation period for {worker_name} has been cancelled and worker status reset to active.')
        return redirect('zone:probation_list')
    
    return redirect('zone:probation_list')

@login_required  
@permission_required('hr.delete_probationperiod', raise_exception=True)
def probation_delete(request, probation_id):
    """Delete a probation period completely (use with caution)."""
    probation = get_object_or_404(WorkerProbationPeriod, id=probation_id)
    
    if request.method == 'POST':
        worker_name = probation.worker.get_full_name()
        worker = probation.worker
        
        # Reset worker status if they were on probation
        if worker.status in ['probation', 'extended']:
            worker.status = 'active'  # Reset to active status
            worker.save()
            
        # Delete the probation record completely
        probation.delete()
        
        messages.success(request, f'Probation period for {worker_name} has been permanently deleted and worker status reset to active.')
        return redirect('zone:probation_list')
    
    return redirect('zone:probation_list')

@login_required
def probation_extensions_list(request):
    """Global list of all probation extensions across all workers."""
    # Basic implementation
    extensions = WorkerProbationExtension.objects.select_related(
        'probation_period__worker', 
        'probation_period__worker__zone__created_by',
        'approved_by', 
        'created_by'
    ).order_by('-created_at')
    
    # Pagination
    paginator = Paginator(extensions, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'total_extensions': extensions.count(),
    }
    return render(request, 'zone/probation_extensions_list.html', context)

# ============================================================================
# HELPER VIEWS
# ============================================================================

@login_required
@require_http_methods(["GET"])
def get_floors_by_building(request):
    """Get floors for a specific building (AJAX)."""
    building_id = request.GET.get('building_id')
    if building_id:
        try:
            building = Building.objects.get(id=building_id)
            floors = Floor.objects.filter(building_id=building_id, is_active=True)
            floor_data = [{'id': floor.id, 'name': floor.name} for floor in floors]
            return JsonResponse({
                'floors': floor_data,
                'building_no_floor': building.no_floor
            })
        except Building.DoesNotExist:
            return JsonResponse({'floors': [], 'building_no_floor': False})
    return JsonResponse({'floors': [], 'building_no_floor': False})

@login_required
@require_http_methods(["GET"])
def get_floors_by_building_id(request, building_id):
    """Get floors for a specific building via URL parameter (AJAX)."""
    try:
        building = Building.objects.get(id=building_id)
        floors = Floor.objects.filter(building_id=building_id, is_active=True)
        floor_data = [{'id': floor.id, 'name': floor.name} for floor in floors]
        return JsonResponse({
            'floors': floor_data,
            'building_no_floor': building.no_floor
        })
    except Building.DoesNotExist:
        return JsonResponse({'floors': [], 'building_no_floor': False})

@login_required
@require_http_methods(["GET"])
def get_buildings_by_zone(request):
    """Get buildings for a specific zone (AJAX)."""
    zone_id = request.GET.get('zone_id')
    if zone_id:
        buildings = Building.objects.filter(zone_id=zone_id, is_active=True)
        building_data = [{'id': building.id, 'name': building.name} for building in buildings]
        return JsonResponse({'buildings': building_data})
    return JsonResponse({'buildings': []})

# OCR functionality removed

@login_required
def worker_reports(request):
    """Worker Reports dashboard with compact UI and ultra minimal design."""
    
    # Get filter parameters
    building_filter = request.GET.get('building', '')
    floor_filter = request.GET.get('floor', '')
    status_filter = request.GET.get('status', '')
    nationality_filter = request.GET.get('nationality', '')
    vip_filter = request.GET.get('vip_filter', '')  # Add VIP filter
    date_range = request.GET.get('date_range', '30')  # Default to 30 days
    
    # Base queryset
    workers = Worker.objects.select_related('building', 'floor', 'zone', 'position')
    
    # Apply filters
    if building_filter:
        workers = workers.filter(building_id=building_filter)
    if floor_filter:
        workers = workers.filter(floor_id=floor_filter)
    if status_filter:
        workers = workers.filter(status=status_filter)
    if nationality_filter:
        workers = workers.filter(nationality=nationality_filter)
    if vip_filter:
        if vip_filter == 'vip':
            workers = workers.filter(is_vip=True)
        elif vip_filter == 'regular':
            workers = workers.filter(is_vip=False)
    
    # Date range filter for new joiners
    if date_range:
        try:
            days = int(date_range)
            cutoff_date = timezone.now().date() - timedelta(days=days)
            new_joiners = workers.filter(date_joined__gte=cutoff_date)
        except (ValueError, TypeError):
            new_joiners = workers.filter(date_joined__gte=timezone.now().date() - timedelta(days=30))
    else:
        new_joiners = workers.none()
    
    # Statistics
    total_workers = workers.count()
    active_workers = workers.filter(status='active').count()
    inactive_workers = workers.filter(status='inactive').count()
    on_leave_workers = workers.filter(status='on_leave').count()
    terminated_workers = workers.filter(status='terminated').count()
    new_joiners_count = new_joiners.count()
    
    # VIP Statistics
    total_vips = workers.filter(is_vip=True).count()
    vip_active = workers.filter(is_vip=True, status='active').count()
    regular_workers = workers.filter(is_vip=False).count()
    vip_new_joiners = new_joiners.filter(is_vip=True).count()
    
    # Workers on probation
    on_probation = workers.filter(
        status__in=['probation', 'extended'],
        probation_periods__actual_end_date__isnull=True
    ).distinct().count()
    
    # Document expiry warnings
    expiring_docs = workers.filter(
        documents__expiry_date__lte=timezone.now().date() + timedelta(days=30),
        documents__expiry_date__gte=timezone.now().date()
    ).distinct().count()
    
    expired_docs = workers.filter(
        documents__expiry_date__lt=timezone.now().date()
    ).distinct().count()
    
    # Performance distribution
    performance_stats = workers.filter(performance_rating__isnull=False).aggregate(
        avg_rating=Avg('performance_rating'),
        excellent=Count(Case(When(performance_rating=5, then=1))),
        good=Count(Case(When(performance_rating=4, then=1))),
        average=Count(Case(When(performance_rating=3, then=1))),
        below_avg=Count(Case(When(performance_rating=2, then=1))),
        poor=Count(Case(When(performance_rating=1, then=1)))
    )
    
    # Top buildings and floors by worker count
    building_stats = Building.objects.annotate(
        worker_count=Count('workers')
    ).filter(worker_count__gt=0).order_by('-worker_count')[:5]
    
    floor_stats = Floor.objects.annotate(
        worker_count=Count('workers')
    ).filter(worker_count__gt=0).order_by('-worker_count')[:5]
    
    # Nationality distribution
    nationality_stats = workers.values('nationality').annotate(
        count=Count('id')
    ).filter(count__gt=0).order_by('-count')[:10]
    
    # Recent activity (last 7 days)
    recent_joiners = workers.filter(
        date_joined__gte=timezone.now().date() - timedelta(days=7)
    ).order_by('-date_joined')[:10]
    
    # VIP reports data
    recent_vip_joiners = workers.filter(
        is_vip=True,
        date_joined__gte=timezone.now().date() - timedelta(days=30)
    ).order_by('-date_joined')[:10]
    
    # VIP by building distribution
    vip_building_stats = Building.objects.annotate(
        vip_count=Count('workers', filter=Q(workers__is_vip=True))
    ).filter(vip_count__gt=0).order_by('-vip_count')[:5]
    
    # VIP by nationality
    vip_nationality_stats = workers.filter(is_vip=True).values('nationality').annotate(
        count=Count('id')
    ).filter(count__gt=0).order_by('-count')[:10]
    
    # Filter options
    buildings = Building.objects.filter(is_active=True).order_by('name')
    floors = Floor.objects.filter(is_active=True).select_related('building').order_by('building__name', 'floor_number')
    nationalities = Worker.objects.values_list('nationality', flat=True).distinct()
    
    context = {
        'total_workers': total_workers,
        'active_workers': active_workers,
        'inactive_workers': inactive_workers,
        'on_leave_workers': on_leave_workers,
        'terminated_workers': terminated_workers,
        'new_joiners_count': new_joiners_count,
        'on_probation': on_probation,
        'expiring_docs': expiring_docs,
        'expired_docs': expired_docs,
        'performance_stats': performance_stats,
        'building_stats': building_stats,
        'floor_stats': floor_stats,
        'nationality_stats': nationality_stats,
        'recent_joiners': recent_joiners,
        # VIP data
        'total_vips': total_vips,
        'vip_active': vip_active,
        'regular_workers': regular_workers,
        'vip_new_joiners': vip_new_joiners,
        'recent_vip_joiners': recent_vip_joiners,
        'vip_building_stats': vip_building_stats,
        'vip_nationality_stats': vip_nationality_stats,
        # Filter options
        'buildings': buildings,
        'floors': floors,
        'nationalities': sorted([n for n in nationalities if n]),
        'building_filter': building_filter,
        'floor_filter': floor_filter,
        'status_filter': status_filter,
        'nationality_filter': nationality_filter,
        'vip_filter': vip_filter,
        'date_range': date_range,
    }
    
    return render(request, 'zone/worker_reports.html', context)

@login_required
def worker_document_list(request, worker_id):
    """List all documents for a specific worker."""
    worker = get_object_or_404(Worker, id=worker_id)
    documents = worker.documents.all().order_by('-created_at')
    
    context = {
        'worker': worker,
        'documents': documents,
        'page_title': f'Documents - {worker.get_full_name()}',
    }
    return render(request, 'zone/worker_document_list.html', context)

@login_required
def worker_document_add(request, worker_id):
    """Add a new document for a worker."""
    worker = get_object_or_404(Worker, id=worker_id)
    
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.worker = worker
            document.created_by = request.user
            document.save()
            messages.success(request, 'Document added successfully.')
            return redirect('zone:worker_detail', worker_id=worker.id)
    else:
        form = DocumentForm()
    
    context = {
        'worker': worker,
        'form': form,
        'page_title': f'Add Document - {worker.get_full_name()}',
    }
    return render(request, 'zone/worker_document_form.html', context)

@login_required
def document_edit(request, document_id):
    """Edit a document."""
    document = get_object_or_404(Document, id=document_id)
    
    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES, instance=document)
        if form.is_valid():
            form.save()
            messages.success(request, 'Document updated successfully.')
            return redirect('zone:worker_detail', worker_id=document.worker.id)
    else:
        form = DocumentForm(instance=document)
    
    context = {
        'document': document,
        'form': form,
        'page_title': f'Edit Document - {document.get_document_type_display()}',
    }
    return render(request, 'zone/document_form.html', context)

@login_required
def document_delete(request, document_id):
    """Delete a document."""
    document = get_object_or_404(Document, id=document_id)
    worker_id = document.worker.id
    
    if request.method == 'POST':
        document.delete()
        messages.success(request, 'Document deleted successfully.')
        return redirect('zone:worker_detail', worker_id=worker_id)
    
    context = {
        'document': document,
        'page_title': f'Delete Document - {document.get_document_type_display()}',
    }
    return render(request, 'zone/document_confirm_delete.html', context)

@login_required
def document_download(request, document_id):
    """Download a document file."""
    document = get_object_or_404(Document, id=document_id)
    
    if not document.document_file:
        messages.error(request, 'No file attached to this document.')
        return redirect('zone:worker_detail', worker_id=document.worker.id)
    
    # Use the secure media view for downloading
    return redirect('core:serve_worker_document', document_id=document.id)

@login_required
@permission_required('zone.add_workerprobationperiod', raise_exception=True)
def probation_create(request, worker_id=None):
    """Create a new probation period for a worker (VIPs and employees are not eligible)."""
    
    worker = None
    if worker_id:
        worker = get_object_or_404(Worker, id=worker_id)

    if request.method == 'POST':
        # If worker_id is provided, ensure it's included in POST data
        post_data = request.POST.copy()
        if worker_id and not post_data.get('worker'):
            post_data['worker'] = str(worker_id)
            
        form = WorkerProbationPeriodForm(post_data, worker_id=worker_id)
        if form.is_valid():
            probation = form.save(commit=False)
            probation.created_by = request.user
            
            # If worker_id is provided but no worker is selected in form, use the provided worker
            if worker_id and not probation.worker:
                probation.worker = worker
                
            probation.save()
            
            # Automatically set worker status to "probation" when creating a new probation period
            selected_worker = probation.worker
            selected_worker.status = 'probation'
            selected_worker.save(update_fields=['status'])
            
            messages.success(request, f'Probation period created successfully for Worker {selected_worker.get_full_name()}. Status updated to "Probation".')
            
            # Redirect to worker's probation status page if we came from a specific worker
            if worker_id:
                return redirect('zone:worker_probation_status', worker_id=selected_worker.id)
            else:
                return redirect('zone:probation_list')
        else:
            # Add error messages to help debug form issues
            if form.errors:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"Form error in {field}: {error}")
            if form.non_field_errors():
                for error in form.non_field_errors():
                    messages.error(request, f"Form validation error: {error}")
    else:
        form = WorkerProbationPeriodForm(worker_id=worker_id)
    
    context = {
        'form': form,
        'worker': worker,
        'page_title': f'Create Probation Period{" for " + worker.get_full_name() if worker else ""}',
        'title': f'Create Probation Period{" for " + worker.get_full_name() if worker else ""}',
    }
    return render(request, 'zone/probation_form.html', context)


@login_required
@permission_required('zone.add_workerprobationperiod', raise_exception=True)
def probation_batch_create(request):
    """Create probation periods for multiple workers at once."""
    
    if request.method == 'POST':
        form = BatchProbationForm(request.POST)
        if form.is_valid():
            workers = form.cleaned_data['workers']
            start_date = form.cleaned_data['start_date']
            original_end_date = form.cleaned_data['original_end_date']
            evaluation_notes = form.cleaned_data['evaluation_notes']
            batch_name = form.cleaned_data['batch_name']
            
            # Prepare batch tracking information
            batch_info = f"Batch: {batch_name}"
            if evaluation_notes:
                full_evaluation_notes = f"{batch_info}\n\n{evaluation_notes}"
            else:
                full_evaluation_notes = batch_info
            
            created_probations = []
            failed_workers = []
            
            # Create probation periods for each selected worker
            for worker in workers:
                try:
                    # Check if worker already has an active probation period
                    existing_probation = worker.probation_periods.filter(
                        actual_end_date__isnull=True
                    ).first()
                    
                    if existing_probation:
                        failed_workers.append({
                            'worker': worker,
                            'reason': 'Already has an active probation period'
                        })
                        continue
                    
                    # Create the probation period
                    probation = WorkerProbationPeriod.objects.create(
                        worker=worker,
                        start_date=start_date,
                        original_end_date=original_end_date,
                        evaluation_notes=full_evaluation_notes,
                        created_by=request.user
                    )
                    
                    # Update worker status to "probation"
                    worker.status = 'probation'
                    worker.save(update_fields=['status'])
                    
                    created_probations.append(probation)
                    
                except Exception as e:
                    failed_workers.append({
                        'worker': worker,
                        'reason': f'Error: {str(e)}'
                    })
            
            # Provide feedback to user
            if created_probations:
                messages.success(
                    request, 
                    f'Successfully created {len(created_probations)} probation period(s) for batch "{batch_name}". '
                    f'Worker statuses updated to "Probation".'
                )
            
            if failed_workers:
                failed_names = []
                for item in failed_workers:
                    failed_names.append(f"{item['worker'].get_full_name()} ({item['reason']})")
                messages.warning(
                    request,
                    f'Failed to create probation periods for {len(failed_workers)} worker(s): '
                    f'{", ".join(failed_names)}'
                )
            
            return redirect('zone:probation_list')
        else:
            # Add error messages for form validation
            if form.errors:
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f"Form error in {field}: {error}")
            if form.non_field_errors():
                for error in form.non_field_errors():
                    messages.error(request, f"Form validation error: {error}")
    else:
        form = BatchProbationForm(request.GET or None)

    context = {
        'form': form,
        'page_title': 'Batch Probation Creation',
        'title': 'Batch Probation Creation',
    }
    return render(request, 'zone/probation_batch_create.html', context)


@login_required
@permission_required('hr.change_probationperiod', raise_exception=True)
def extension_request_review(request, request_id):
    """Review (approve/reject) a probation extension request - Manager only."""
    extension_request = get_object_or_404(ProbationExtensionRequest, id=request_id)
    
    # Check if user has manager privileges
    if not is_manager(request.user):
        messages.error(request, 'You do not have permission to review extension requests.')
        return redirect('zone:probation_list')
    
    # Check if request is still pending
    if extension_request.status != 'pending':
        messages.error(request, f'This request has already been {extension_request.get_status_display().lower()}.')
        return redirect('zone:probation_list')
    
    if request.method == 'POST':
        form = ProbationExtensionRequestReviewForm(request.POST, instance=extension_request)
        if form.is_valid():
            review = form.save(commit=False)
            review.reviewed_by = request.user
            review.save()
            
            if review.status == 'approved':
                messages.success(request, f'Extension request approved. Probation period extended by {extension_request.extension_duration_days} days.')
            else:
                messages.info(request, 'Extension request rejected.')
                
            return redirect('zone:probation_list')
    else:
        form = ProbationExtensionRequestReviewForm(instance=extension_request)
    
    context = {
        'extension_request': extension_request,
        'probation': extension_request.probation_period,
        'worker': extension_request.probation_period.worker,
        'form': form,
        'title': f'Review Extension Request - {extension_request.probation_period.worker.get_full_name()}',
    }
    return render(request, 'zone/extension_request_review.html', context)

@login_required
@permission_required('hr.change_probationperiod', raise_exception=True)
def extension_request_cancel(request, request_id):
    """Cancel a pending extension request - Requester only."""
    extension_request = get_object_or_404(ProbationExtensionRequest, id=request_id)
    
    # Check if user is the requester
    if extension_request.requested_by != request.user:
        messages.error(request, 'You can only cancel your own extension requests.')
        return redirect('zone:probation_list')
    
    # Check if request can be cancelled
    if not extension_request.can_be_cancelled():
        messages.error(request, 'This request cannot be cancelled as it has already been reviewed.')
        return redirect('zone:probation_list')
    
    if request.method == 'POST':
        extension_request.status = 'cancelled'
        extension_request.save()
        
        messages.info(request, 'Extension request cancelled.')
        return redirect('zone:probation_list')
    
    context = {
        'extension_request': extension_request,
        'probation': extension_request.probation_period,
        'worker': extension_request.probation_period.worker,
        'title': f'Cancel Extension Request - {extension_request.probation_period.worker.get_full_name()}',
    }
    return render(request, 'zone/extension_request_cancel.html', context)

@login_required
@permission_required('hr.view_probationperiod', raise_exception=True)
def extension_requests_list(request):
    """List all probation extension requests - Manager view."""
    if not is_manager(request.user):
        messages.error(request, 'You do not have permission to view all extension requests.')
        return redirect('zone:probation_list')
    
    # Get filter parameters
    status_filter = request.GET.get('status', 'pending')
    search_query = request.GET.get('q', '')
    
    # Base queryset
    requests = ProbationExtensionRequest.objects.select_related(
        'probation_period__worker', 'requested_by', 'reviewed_by'
    ).order_by('-requested_at')
    
    # Apply status filter
    if status_filter and status_filter != 'all':
        requests = requests.filter(status=status_filter)
    
    # Apply search filter
    if search_query:
        requests = requests.filter(
            Q(probation_period__worker__first_name__icontains=search_query) |
            Q(probation_period__worker__last_name__icontains=search_query) |
            Q(probation_period__worker__worker_id__icontains=search_query) |
            Q(reason__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(requests, 20)
    page = request.GET.get('page')
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Calculate counts for dashboard
    total_requests = ProbationExtensionRequest.objects.count()
    pending_count = ProbationExtensionRequest.objects.filter(status='pending').count()
    approved_count = ProbationExtensionRequest.objects.filter(status='approved').count()
    rejected_count = ProbationExtensionRequest.objects.filter(status='rejected').count()
    
    context = {
        'page_obj': page_obj,
        'total_requests': total_requests,
        'pending_count': pending_count,
        'approved_count': approved_count,
        'rejected_count': rejected_count,
        'status_filter': status_filter,
        'search_query': search_query,
        'title': 'Probation Extension Requests',
    }
    return render(request, 'zone/extension_requests_list.html', context)

@login_required
def get_floors_ajax(request):
    """AJAX endpoint to get floors for a building."""
    building_id = request.GET.get('building_id')
    floors = Floor.objects.filter(building_id=building_id, is_active=True).values('id', 'name')
    return JsonResponse({'floors': list(floors)})

@login_required
def worker_photo_ajax(request, worker_id):
    """AJAX endpoint to get worker photo."""
    worker = get_object_or_404(Worker, id=worker_id)
    if worker.photo:
        # Return the secure photo URL instead of direct photo.url
        photo_url = reverse('zone:serve_worker_photo', kwargs={'worker_id': worker.id})
        return JsonResponse({'photo_url': photo_url})
    return JsonResponse({'photo_url': None})

@login_required
def serve_worker_photo(request, worker_id):
    """Serve encrypted worker photo."""
    try:
        worker = get_object_or_404(Worker, id=worker_id)
        
        # Check if worker has a photo
        if not worker.photo or not worker.photo.name:
            # Return default avatar instead of 404
            return redirect('/static/images/default-worker-avatar.png')
        
        # Get the file path
        file_path = worker.photo.path
        
        if not os.path.exists(file_path):
            # Return default avatar instead of 404 when file is missing
            return redirect('/static/images/default-worker-avatar.png')
        
        # Read and decrypt if needed
        try:
            if worker.photo.name.endswith('.enc'):
                # Encrypted file - decrypt it
                with open(file_path, 'rb') as f:
                    encrypted_content = f.read().decode('utf-8')
                
                decrypted_content = FileEncryptionHandler.decrypt_file_content(encrypted_content)
                file_content = decrypted_content
                original_filename = worker.photo.name.replace('.enc', '')
            else:
                # Non-encrypted file - serve directly
                with open(file_path, 'rb') as f:
                    file_content = f.read()
                original_filename = worker.photo.name
            
            # Determine content type
            import mimetypes
            content_type, _ = mimetypes.guess_type(original_filename)
            if not content_type or not content_type.startswith('image/'):
                content_type = 'image/jpeg'  # Default to JPEG for images
            
            # Create response
            response = HttpResponse(file_content, content_type=content_type)
            response['Content-Disposition'] = f'inline; filename="worker_{worker.worker_id}_photo.{original_filename.split(".")[-1]}"'
            response['Content-Length'] = len(file_content)
            
            return response
            
        except Exception as e:
            raise Http404("Photo could not be processed")
            
    except Worker.DoesNotExist:
        raise Http404("Worker not found")

@login_required
def serve_worker_professional_photo(request, worker_id):
    """Serve encrypted worker professional photo."""
    try:
        worker = get_object_or_404(Worker, id=worker_id)
        
        # Check if worker has a professional photo
        if not worker.professional_photo or not worker.professional_photo.name:
            raise Http404("Professional photo not found")
        
        # Get the file path
        file_path = worker.professional_photo.path
        
        if not os.path.exists(file_path):
            raise Http404("Professional photo file not found on disk")
        
        # Read and decrypt if needed
        try:
            if worker.professional_photo.name.endswith('.enc'):
                # Encrypted file - decrypt it
                with open(file_path, 'rb') as f:
                    encrypted_content = f.read().decode('utf-8')
                
                decrypted_content = FileEncryptionHandler.decrypt_file_content(encrypted_content)
                file_content = decrypted_content
                original_filename = worker.professional_photo.name.replace('.enc', '')
            else:
                # Non-encrypted file - serve directly
                with open(file_path, 'rb') as f:
                    file_content = f.read()
                original_filename = worker.professional_photo.name
            
            # Determine content type
            import mimetypes
            content_type, _ = mimetypes.guess_type(original_filename)
            if not content_type or not content_type.startswith('image/'):
                content_type = 'image/jpeg'  # Default to JPEG for images
            
            # Create response
            response = HttpResponse(file_content, content_type=content_type)
            response['Content-Disposition'] = f'inline; filename="worker_{worker.worker_id}_professional_photo.{original_filename.split(".")[-1]}"'
            response['Content-Length'] = len(file_content)
            
            return response
            
        except Exception as e:
            raise Http404("Professional photo could not be processed")
            
    except Worker.DoesNotExist:
        raise Http404("Worker not found") 

#check passport duplicate
@login_required
def check_passport_duplicate(request):
    """Check if a passport number already exists in the database."""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            import json
            data = json.loads(request.body)
            document_number = data.get('passport_number', '').strip()
            worker_id = data.get('worker_id','').strip()
            
            if not document_number:
                return JsonResponse({'exists': False})
            
            # Check in Worker records
            worker_qs = Worker.objects.filter(worker_id=worker_id)
            if worker_id:
                worker_qs = worker_qs.exclude(pk=worker_id)
            
            if worker_qs.exists():
                existing_worker = worker_qs.first()
                return JsonResponse({
                    'exists': True,
                    'worker_name': "",
                    'worker_id': existing_worker.pk
                })
            
            # Check in VIP documents
            doc_qs = Document.objects.filter(
                document_number=document_number,
                document_type='passport'
            )
            
            if doc_qs.exists():
                existing_doc = doc_qs.first()
                return JsonResponse({
                    'exists': True,
                    'worker_name': "",
                    'worker_id': existing_doc.worker.pk
                })
            
            return JsonResponse({'exists': False})
            
        except Exception as e:
            return JsonResponse({'exists': False, 'error': str(e)})
    
    return JsonResponse({'exists': False, 'error': 'Invalid request'})

@login_required
def serve_encrypted_image(request, pk):
    try:
        # Get the document
        obj = Document.objects.get(pk=pk)
        
        # Check if document has a file
        if not obj.document_file or not obj.document_file.name:
            # Return placeholder instead of 404
            placeholder_svg = f'''<?xml version="1.0" encoding="UTF-8"?>
            <svg width="200" height="150" xmlns="http://www.w3.org/2000/svg">
                <rect width="100%" height="100%" fill="#f8f9fa" stroke="#dee2e6" stroke-width="2"/>
                <text x="50%" y="40%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="14" fill="#6c757d">
                    No File Attached
                </text>
                <text x="50%" y="60%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="10" fill="#6c757d">
                    Document ID: {pk}
                </text>
            </svg>'''
            return HttpResponse(placeholder_svg, content_type="image/svg+xml")
            
        enc_file = obj.document_file  # EncryptedImageField
        
        # Check if file exists
        if not enc_file.file:
            # Return placeholder instead of 404
            placeholder_svg = f'''<?xml version="1.0" encoding="UTF-8"?>
            <svg width="200" height="150" xmlns="http://www.w3.org/2000/svg">
                <rect width="100%" height="100%" fill="#f8f9fa" stroke="#dee2e6" stroke-width="2"/>
                <text x="50%" y="40%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="14" fill="#6c757d">
                    File Corrupted
                </text>
                <text x="50%" y="60%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="10" fill="#6c757d">
                    Document ID: {pk}
                </text>
            </svg>'''
            return HttpResponse(placeholder_svg, content_type="image/svg+xml")
            
    except Document.DoesNotExist:
        # Return placeholder instead of 404 to prevent template breaking
        placeholder_svg = f'''<?xml version="1.0" encoding="UTF-8"?>
        <svg width="200" height="150" xmlns="http://www.w3.org/2000/svg">
            <rect width="100%" height="100%" fill="#f8f9fa" stroke="#dee2e6" stroke-width="2"/>
            <text x="50%" y="40%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="14" fill="#6c757d">
                Document Not Found
            </text>
            <text x="50%" y="60%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="10" fill="#6c757d">
                ID: {pk}
            </text>
        </svg>'''
        return HttpResponse(placeholder_svg, content_type="image/svg+xml")
    except Exception as e:
        # Return placeholder instead of 404 to prevent template breaking  
        placeholder_svg = f'''<?xml version="1.0" encoding="UTF-8"?>
        <svg width="200" height="150" xmlns="http://www.w3.org/2000/svg">
            <rect width="100%" height="100%" fill="#f8f9fa" stroke="#dee2e6" stroke-width="2"/>
            <text x="50%" y="40%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="14" fill="#6c757d">
                Access Error
            </text>
            <text x="50%" y="60%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="10" fill="#6c757d">
                ID: {pk}
            </text>
        </svg>'''
        return HttpResponse(placeholder_svg, content_type="image/svg+xml")

    try:
        # Decrypt file with proper file handling
        with enc_file.file as decrypted_file:
            # Reset file pointer to beginning
            decrypted_file.seek(0)
            image_data = decrypted_file.read()
            
        # Check if we got data
        if not image_data:
            # Return placeholder instead of 404
            placeholder_svg = f'''<?xml version="1.0" encoding="UTF-8"?>
            <svg width="200" height="150" xmlns="http://www.w3.org/2000/svg">
                <rect width="100%" height="100%" fill="#f8f9fa" stroke="#dee2e6" stroke-width="2"/>
                <text x="50%" y="40%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="14" fill="#6c757d">
                    Empty File
                </text>
                <text x="50%" y="60%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="10" fill="#6c757d">
                    Document ID: {pk}
                </text>
            </svg>'''
            return HttpResponse(placeholder_svg, content_type="image/svg+xml")

        # Determine content type from file extension or actual file content
        file_name = enc_file.name.lower()
        if file_name.endswith(('.jpg', '.jpeg')):
            content_type = 'image/jpeg'
        elif file_name.endswith('.png'):
            content_type = 'image/png'
        elif file_name.endswith('.gif'):
            content_type = 'image/gif'
        elif file_name.endswith('.pdf'):
            content_type = 'application/pdf'
        elif file_name.endswith(('.doc', '.docx')):
            content_type = 'application/msword'
        else:
            content_type = 'application/octet-stream'  # fallback

        return HttpResponse(image_data, content_type=content_type)
        
    except Exception as e:
        import traceback
        
        # Instead of raising Http404, return a placeholder SVG image
        # This prevents template rendering from breaking when one document has issues
        placeholder_svg = '''<?xml version="1.0" encoding="UTF-8"?>
        <svg width="200" height="150" xmlns="http://www.w3.org/2000/svg">
            <rect width="100%" height="100%" fill="#f8f9fa" stroke="#dee2e6" stroke-width="2"/>
            <text x="50%" y="40%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="14" fill="#6c757d">
                Document Error
            </text>
            <text x="50%" y="60%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="10" fill="#6c757d">
                ID: ''' + str(pk) + '''
            </text>
            <text x="50%" y="75%" dominant-baseline="middle" text-anchor="middle" font-family="Arial" font-size="8" fill="#dc3545">
                File Access Failed
            </text>
        </svg>'''
        
        response = HttpResponse(placeholder_svg, content_type="image/svg+xml")
        response['X-Document-Error'] = f'File Access Failed: {str(e)}'
        return response

# ocr image
@login_required
def worker_ocr_image(request):
    """OCR image."""
    if request.method == 'POST' and request.FILES.get('image') and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        
        # reader = easyocr.Reader(['en'])
        
        try:
            image_file  = request.FILES['image']
            
            image_bytes = image_file.read()

            # Convert to PIL image
            image = Image.open(io.BytesIO(image_bytes))

            # result = reader.readtext(np.array(image))
            res = []
            # for (bbox, text, prob) in result:
            #     res.append(text)

            return JsonResponse({'message': 'Image uploaded', 'data': res})
            
        except Exception as e:
            return JsonResponse({'data': None, 'error': str(e)})
    
    return JsonResponse({'data': None, 'error': 'Invalid request'})


# ===== FILE PRESERVATION SYSTEM =====
# Prevents file loss during form validation failures

def preserve_uploaded_files_in_session(request):
    """Store uploaded files in session to prevent loss during validation failures."""
    if 'preserved_files' not in request.session:
        request.session['preserved_files'] = {}
    
    # Store document files with their field names
    for key, file in request.FILES.items():
        if 'document_file' in key and hasattr(file, 'read'):
            try:
                # Store file data in session (for files up to 5MB)
                if file.size <= 5 * 1024 * 1024:  # 5MB limit
                    # Read file content
                    file_content = file.read()
                    file.seek(0)  # Reset file pointer for Django processing
                    
                    file_data = {
                        'name': file.name,
                        'content_type': file.content_type,
                        'size': file.size,
                        'data': file_content,
                        'field_name': key  # Store the exact field name
                    }
                    request.session['preserved_files'][key] = file_data
                    
                    # Also store base64 for JavaScript restoration
                    import base64
                    base64_data = base64.b64encode(file_content).decode('utf-8')
                    data_url = f"data:{file.content_type};base64,{base64_data}"
                    request.session['preserved_files'][f"{key}_data_url"] = data_url
                    
            except Exception as e:
                pass


def restore_preserved_files_to_forms(request, formset):
    """Restore preserved files to formset after validation failure AND prepare for JavaScript restoration."""
    preserved_files = request.session.get('preserved_files', {})
    
    if not preserved_files:
        return
    
    # Prepare JavaScript restoration data
    js_restoration_data = {}
    
    try:
        for form in formset:
            if hasattr(form, 'prefix') and form.prefix:
                file_key = f"{form.prefix}-document_file"
                
                if file_key in preserved_files:
                    file_data = preserved_files[file_key]
                    
                    if isinstance(file_data, dict) and 'data' in file_data:
                        # 1. Restore to Django form
                        from django.core.files.uploadedfile import SimpleUploadedFile
                        temp_file = SimpleUploadedFile(
                            file_data['name'],
                            file_data['data'],
                            content_type=file_data['content_type']
                        )
                        
                        # Inject back into form files
                        if hasattr(form, 'files'):
                            form.files['document_file'] = temp_file
                        if hasattr(form, 'cleaned_data'):
                            form.cleaned_data['document_file'] = temp_file
                        
                        # 2. Prepare JavaScript restoration data
                        data_url_key = f"{file_key}_data_url"
                        if data_url_key in preserved_files:
                            # Extract form index from prefix (e.g., "documents-0" -> 0)
                            try:
                                form_index = int(form.prefix.split('-')[1])
                                js_restoration_data[form_index] = {
                                    'data_url': preserved_files[data_url_key],
                                    'filename': file_data['name'],
                                    'size': file_data['size']
                                }
                            except (IndexError, ValueError):
                                pass
                            
    except Exception as e:
        pass
        
    # Store restoration data for JavaScript access
    request.js_file_restoration = js_restoration_data


def clear_preserved_files_from_session(request):
    """Clear preserved files from session."""
    if 'preserved_files' in request.session:
        del request.session['preserved_files']


@require_http_methods(["GET"])
@login_required
def smart_search(request):
    """Smart search endpoint for auto-suggestions across workers, VIPs, employees, and menu items."""
    try:
        query = request.GET.get('q', '').strip()

        if len(query) < 2:  # Minimum 2 characters to search
            return JsonResponse({'results': []})

        results = []

        # Search menu items from sidebar
        try:
            from core.context_processors import sidebar_menu as get_sidebar_menu
            sidebar_data = get_sidebar_menu(request)
            menu_items = sidebar_data.get('sidebar_menu', [])

            # Extract all menu items and search them
            for section in menu_items:
                section_name = section.get('name', '')

                # Search through main items in the section
                for item in section.get('items', []):
                    item_name = item.get('name', '')
                    item_url = item.get('url', '#')
                    item_icon = item.get('icon', 'bi-circle')

                    # Check if query matches item name or section name
                    if (query.lower() in item_name.lower() or
                        query.lower() in section_name.lower()):

                        results.append({
                            'id': f"menu_{section_name}_{item_name}".replace(' ', '_'),
                            'title': item_name,
                            'subtitle': f"Navigate to {section_name}",
                            'type': 'Menu Item',
                            'type_class': 'menu',
                            'url': item_url,
                            'avatar': None,
                            'icon': item_icon,
                            'status': None,
                            'zone': None,
                        })

                # Search through subsection items
                for subsection in section.get('subsections', []):
                    subsection_name = subsection.get('name', '')
                    for item in subsection.get('items', []):
                        item_name = item.get('name', '')
                        item_url = item.get('url', '#')
                        item_icon = item.get('icon', 'bi-circle')

                        # Check if query matches item name, subsection name, or section name
                        if (query.lower() in item_name.lower() or
                            query.lower() in subsection_name.lower() or
                            query.lower() in section_name.lower()):

                            results.append({
                                'id': f"menu_{section_name}_{subsection_name}_{item_name}".replace(' ', '_'),
                                'title': item_name,
                                'subtitle': f"Navigate to {section_name} → {subsection_name}",
                                'type': 'Menu Item',
                                'type_class': 'menu',
                                'url': item_url,
                                'avatar': None,
                                'icon': item_icon,
                                'status': None,
                                'zone': None,
                            })
        except Exception as menu_error:
            pass
            # If menu search fails, log but continue with other searches

        # Search workers (including VIPs)
        workers = Worker.objects.filter(
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(worker_id__icontains=query) |
            Q(phone_number__icontains=query)
        ).select_related('zone', 'building').order_by('first_name', 'last_name')[:10]

        for worker in workers:
            # Determine worker type
            worker_type = "VIP Worker" if worker.is_vip else "Worker"

            # Build subtitle with relevant info
            subtitle_parts = []
            if worker.worker_id:
                subtitle_parts.append(f"ID: {worker.worker_id}")
            if worker.zone:
                subtitle_parts.append(f"Zone: {worker.zone.name}")
            if worker.building:
                subtitle_parts.append(f"Building: {worker.building.name}")
            if worker.phone_number:
                subtitle_parts.append(f"Phone: {worker.phone_number}")

            subtitle = " • ".join(subtitle_parts) if subtitle_parts else "No additional info"

            results.append({
                'id': worker.id,
                'title': worker.get_full_name(),
                'subtitle': subtitle,
                'type': worker_type,
                'type_class': 'vip' if worker.is_vip else 'worker',
                'url': reverse('zone:worker_detail', kwargs={'worker_id': worker.id}),
                'avatar': worker.photo.url if worker.photo else None,
                'icon': None,
                'status': worker.get_status_display(),
                'zone': worker.zone.name if worker.zone else None,
            })

        # Search employees if there's an Employee model (check if it exists)
        try:
            from hr.models import Employee
            employees = Employee.objects.filter(
                Q(first_name__icontains=query) |
                Q(last_name__icontains=query) |
                Q(employee_id__icontains=query) |
                Q(phone_number__icontains=query) |
                Q(email__icontains=query)
            ).order_by('first_name', 'last_name')[:5]

            for employee in employees:
                subtitle_parts = []
                if employee.employee_id:
                    subtitle_parts.append(f"ID: {employee.employee_id}")
                if hasattr(employee, 'department') and employee.department:
                    subtitle_parts.append(f"Dept: {employee.department}")
                if employee.phone_number:
                    subtitle_parts.append(f"Phone: {employee.phone_number}")

                subtitle = " • ".join(subtitle_parts) if subtitle_parts else "No additional info"

                results.append({
                    'id': employee.id,
                    'title': employee.full_name,
                    'subtitle': subtitle,
                    'type': 'Employee',
                    'type_class': 'employee',
                    'url': reverse('hr:employee_detail', kwargs={'pk': employee.id}),
                    'avatar': employee.photo.url if hasattr(employee, 'photo') and employee.photo else None,
                    'icon': None,
                    'status': employee.get_status_display() if hasattr(employee, 'get_status_display') else 'Active',
                    'zone': None,
                })
        except ImportError:
            # HR module not available, skip employee search
            pass

        # Sort results by relevance (exact matches first, then partial matches)
        def sort_key(item):
            title_lower = item['title'].lower()
            query_lower = query.lower()

            # Menu items get lower priority than people
            if item['type_class'] == 'menu':
                base_priority = 10
            else:
                base_priority = 0

            # Exact match gets highest priority
            if title_lower == query_lower:
                return base_priority + 0
            # Starts with query gets second priority
            elif title_lower.startswith(query_lower):
                return base_priority + 1
            # Contains query gets third priority
            elif query_lower in title_lower:
                return base_priority + 2
            else:
                return base_priority + 3

        results.sort(key=sort_key)

        return JsonResponse({
            'results': results[:15],  # Limit to 15 results total
            'query': query,
            'total_found': len(results)
        })

    except Exception as e:
        # Return error response for debugging
        return JsonResponse({
            'error': True,
            'message': str(e),
            'results': []
        }, status=500)

# =============================================================================
# WORKER POSITION MANAGEMENT VIEWS
# =============================================================================

@login_required
@manager_or_admin_required("Worker Management Settings")
def worker_position_list(request):
    """List all worker positions with search and pagination."""
    positions = Position.objects.select_related('department').prefetch_related('workers')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    department_filter = request.GET.get('department', '')
    
    if search_query:
        positions = positions.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if department_filter:
        positions = positions.filter(department_id=department_filter)
    
    # Only show active positions by default
    show_inactive = request.GET.get('show_inactive', '')
    if not show_inactive:
        positions = positions.filter(is_active=True)
    
    # Pagination
    per_page = request.GET.get('per_page', '20')
    try:
        per_page = int(per_page)
        if per_page not in [10, 20, 50, 100]:
            per_page = 20
    except ValueError:
        per_page = 20
    
    paginator = Paginator(positions, per_page)
    page = request.GET.get('page')
    
    try:
        positions = paginator.page(page)
    except PageNotAnInteger:
        positions = paginator.page(1)
    except EmptyPage:
        positions = paginator.page(paginator.num_pages)
    
    # Get departments for filter dropdown
    from zone.models import Department
    departments = Department.objects.filter(is_active=True).order_by('name')
    
    context = {
        'positions': positions,
        'departments': departments,
        'search_query': search_query,
        'department_filter': department_filter,
        'show_inactive': show_inactive,
        'per_page': per_page,
    }
    
    return render(request, 'zone/worker_position_list.html', context)


@login_required
@manager_or_admin_required("Worker Management Settings")
def worker_position_detail(request, pk):
    """Display details of a specific worker position."""
    position = get_object_or_404(Position, pk=pk)
    
    # Get workers assigned to this position
    workers = position.workers.all()
    
    context = {
        'position': position,
        'workers': workers,
    }
    
    return render(request, 'zone/worker_position_detail.html', context)


@login_required
@manager_or_admin_required("Worker Management Settings")
def worker_position_create(request):
    """Create a new worker position."""
    if request.method == 'POST':
        form = PositionForm(request.POST)
        if form.is_valid():
            position = form.save(commit=False)
            position.created_by = request.user
            position.save()
            messages.success(request, f'Worker position "{position.name}" created successfully.')
            return redirect('zone:worker_position_detail', pk=position.pk)
    else:
        form = PositionForm()
    
    context = {
        'form': form,
        'title': 'Create Worker Position'
    }
    return render(request, 'zone/worker_position_form.html', context)


@login_required
@manager_or_admin_required("Worker Management Settings")
def worker_position_edit(request, pk):
    """Edit an existing worker position."""
    position = get_object_or_404(Position, pk=pk)
    
    if request.method == 'POST':
        form = PositionForm(request.POST, instance=position)
        if form.is_valid():
            form.save()
            messages.success(request, f'Worker position "{position.name}" updated successfully.')
            return redirect('zone:worker_position_detail', pk=position.pk)
    else:
        form = PositionForm(instance=position)
    
    context = {
        'form': form,
        'position': position,
        'title': f'Edit Worker Position: {position.name}'
    }
    return render(request, 'zone/worker_position_form.html', context)


@login_required
@manager_or_admin_required("Worker Management Settings")
def worker_position_delete(request, pk):
    """Delete a worker position."""
    position = get_object_or_404(Position, pk=pk)
    
    if request.method == 'POST':
        position_name = position.name
        
        # Check if position has workers assigned
        worker_count = position.workers.count()
        if worker_count > 0:
            messages.error(request, 
                f'Cannot delete position "{position_name}" because it has {worker_count} workers assigned to it.')
            return redirect('zone:worker_position_detail', pk=position.pk)
        
        position.delete()
        messages.success(request, f'Worker position "{position_name}" deleted successfully.')
        return redirect('zone:worker_position_list')
    
    # Get workers assigned to this position for display
    workers = position.workers.all()
    
    context = {
        'position': position,
        'workers': workers,
    }
    
    return render(request, 'zone/worker_position_delete.html', context)


# =============================================================================
# WORKER DEPARTMENT MANAGEMENT VIEWS
# =============================================================================

@login_required
@manager_or_admin_required("Worker Management Settings")
def worker_department_list(request):
    """List all worker departments with search and pagination."""
    departments = Department.objects.all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    parent_filter = request.GET.get('parent', '')
    
    if search_query:
        departments = departments.filter(
            Q(name__icontains=search_query) |
            Q(code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if parent_filter:
        departments = departments.filter(parent_id=parent_filter)
    
    # Only show active departments by default
    show_inactive = request.GET.get('show_inactive', '')
    if not show_inactive:
        departments = departments.filter(is_active=True)
    
    # Pagination
    per_page = request.GET.get('per_page', '20')
    try:
        per_page = int(per_page)
        if per_page not in [10, 20, 50, 100]:
            per_page = 20
    except ValueError:
        per_page = 20
    
    paginator = Paginator(departments, per_page)
    page = request.GET.get('page')
    
    try:
        departments = paginator.page(page)
    except PageNotAnInteger:
        departments = paginator.page(1)
    except EmptyPage:
        departments = paginator.page(paginator.num_pages)
    
    # Get parent departments for filter dropdown
    parent_departments = Department.objects.filter(is_active=True, parent__isnull=True).order_by('name')
    
    context = {
        'departments': departments,
        'parent_departments': parent_departments,
        'search_query': search_query,
        'parent_filter': parent_filter,
        'show_inactive': show_inactive,
        'per_page': per_page,
    }
    
    return render(request, 'zone/worker_department_list.html', context)


@login_required
@manager_or_admin_required("Worker Management Settings")
def worker_department_detail(request, pk):
    """Display details of a specific worker department."""
    department = get_object_or_404(Department, pk=pk)
    
    # Get child departments
    child_departments = department.children.all()
    
    # Get positions in this department
    positions = department.positions.all()
    
    context = {
        'department': department,
        'child_departments': child_departments,
        'positions': positions,
    }
    
    return render(request, 'zone/worker_department_detail.html', context)


@login_required
@manager_or_admin_required("Worker Management Settings")
def worker_department_create(request):
    """Create a new worker department."""
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            department = form.save(commit=False)
            department.created_by = request.user
            department.save()
            messages.success(request, f'Worker department "{department.name}" created successfully.')
            return redirect('zone:worker_department_detail', pk=department.pk)
    else:
        form = DepartmentForm()
    
    context = {
        'form': form,
        'title': 'Create Worker Department'
    }
    return render(request, 'zone/worker_department_form.html', context)


@login_required
@manager_or_admin_required("Worker Management Settings")
def worker_department_edit(request, pk):
    """Edit an existing worker department."""
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
            messages.success(request, f'Worker department "{department.name}" updated successfully.')
            return redirect('zone:worker_department_detail', pk=department.pk)
    else:
        form = DepartmentForm(instance=department)
    
    context = {
        'form': form,
        'department': department,
        'title': f'Edit Worker Department: {department.name}'
    }
    return render(request, 'zone/worker_department_form.html', context)


@login_required
@manager_or_admin_required("Worker Management Settings")
def worker_department_delete(request, pk):
    """Delete a worker department."""
    department = get_object_or_404(Department, pk=pk)
    
    if request.method == 'POST':
        department_name = department.name
        
        # Check if department has child departments
        child_count = department.children.count()
        if child_count > 0:
            messages.error(request, 
                f'Cannot delete department "{department_name}" because it has {child_count} child department(s).')
            return redirect('zone:worker_department_detail', pk=department.pk)
        
        # Check if department has positions
        position_count = department.positions.count()
        if position_count > 0:
            messages.error(request, 
                f'Cannot delete department "{department_name}" because it has {position_count} position(s) assigned to it.')
            return redirect('zone:worker_department_detail', pk=department.pk)
        
        department.delete()
        messages.success(request, f'Worker department "{department_name}" deleted successfully.')
        return redirect('zone:worker_department_list')
    
    # Get child departments and positions for display
    child_departments = department.children.all()
    positions = department.positions.all()
    
    context = {
        'department': department,
        'child_departments': child_departments,
        'positions': positions,
    }
    
    return render(request, 'zone/worker_department_delete.html', context)


@login_required
@permission_required('zone.view_workerprobationperiod', raise_exception=True)
def worker_search_for_probation_api(request):
    """API endpoint for worker auto-suggestions when creating probation periods"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:  # Require at least 2 characters
        return JsonResponse({'workers': []})
    
    # Filter workers eligible for probation periods - only active, non-VIP workers
    eligible_workers = Worker.objects.filter(
        status='active',  # Only active workers can be put on probation
        is_vip=False  # Exclude VIPs - probation only for regular workers
    ).select_related('zone', 'position', 'building').filter(
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(worker_id__icontains=query) |
        Q(nickname__icontains=query)
    )
    
    # Limit results and format for auto-suggestion
    suggestions = []
    for worker in eligible_workers[:10]:  # Limit to 10 results
        # Check if worker already has active probation
        # Since WorkerProbationPeriod no longer has status field, 
        # check if worker has probation status and active probation period
        has_active_probation = (
            worker.status in ['probation', 'extended'] and
            WorkerProbationPeriod.objects.filter(worker=worker).exists()
        )
        
        suggestions.append({
            'worker_id': worker.id,
            'worker_name': worker.get_full_name(),
            'worker_code': worker.worker_id,
            'position': worker.position.name if worker.position else 'N/A',
            'zone': worker.zone.name if worker.zone else 'N/A',
            'building': worker.building.name if worker.building else 'N/A',
            'status': worker.get_status_display(),
            'has_active_probation': has_active_probation,
            'display_text': f"{worker.get_full_name()} ({worker.worker_id}) - {worker.position.name if worker.position else 'No Position'}"
        })
    
    return JsonResponse({
        'workers': suggestions,
        'message': f'Found {len(suggestions)} eligible workers for probation periods'
    })


@login_required
def worker_import(request):
    """Import workers from Excel file with SinoSecu API integration."""
    
    if request.method == 'POST':
        form = WorkerExcelImportForm(request.POST, request.FILES)
        if not form.is_valid():
            pass
        
        if form.is_valid():
            # ALWAYS validate first, even if not a preview request
            validation_results = _validate_excel_file(form)
            
            # Check if this is a preview request OR if there are validation issues
            if 'preview' in request.POST or validation_results['invalid_rows'] > 0 or validation_results['duplicates']:
                # Get template download URL
                template_url = request.build_absolute_uri('/') + 'worker_eform.xlsx'
                
                context = {
                    'form': form,
                    'template_url': template_url,
                    'validation_results': validation_results,
                    'is_preview': True,
                    'page_title': 'Import Workers - Validation Results',
                }
                return render(request, 'zone/worker_import_compact.html', context)
            else:
                # Only process if validation passed and not a preview
                return _process_excel_import(request, form)
    else:
        form = WorkerExcelImportForm()
    
    # Get template download URL
    template_url = request.build_absolute_uri('/') + 'worker_eform.xlsx'
    
    context = {
        'form': form,
        'template_url': template_url,
        'page_title': 'Import Workers from Excel',
    }
    # Use compact UI template
    return render(request, 'zone/worker_import_compact.html', context)


def _validate_excel_file(form):
    """Preview and validate Excel file before import."""
    import tempfile
    from openpyxl import load_workbook
    
    excel_file = form.cleaned_data['excel_file']
    
    validation_results = {
        'total_rows': 0,
        'valid_rows': 0,
        'invalid_rows': 0,
        'rows_data': [],
        'errors': [],
        'warnings': [],
        'duplicates': [],
        'missing_fields': [],
    }
    
    tmp_file_name = None
    try:
        # Save Excel file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file_name = tmp_file.name
            for chunk in excel_file.chunks():
                tmp_file.write(chunk)
            tmp_file.flush()
        
        # Read Excel file
        df = pd.read_excel(tmp_file_name, sheet_name='Workers')
        
        # Filter out completely empty rows
        df = df.dropna(how='all')
        validation_results['total_rows'] = len(df)
        
        # Validate each row
        for index, row in df.iterrows():
            row_validation = _validate_worker_row(row, index + 2)  # +2 for Excel row numbering
            validation_results['rows_data'].append(row_validation)
            
            if row_validation['is_valid']:
                validation_results['valid_rows'] += 1
            else:
                validation_results['invalid_rows'] += 1
                
            # Collect errors and warnings
            validation_results['errors'].extend(row_validation['errors'])
            validation_results['warnings'].extend(row_validation['warnings'])
            
            # Check for duplicates
            if row_validation['is_duplicate']:
                validation_results['duplicates'].append(row_validation)
    
    except Exception as e:
        validation_results['errors'].append(f"File processing error: {str(e)}")
    
    finally:
        # Clean up temp file
        if tmp_file_name and os.path.exists(tmp_file_name):
            try:
                os.unlink(tmp_file_name)
            except OSError as e:
                pass
    
    # Calculate how many workers will actually be imported (valid - duplicates)
    validation_results['will_import'] = validation_results['valid_rows'] - len(validation_results['duplicates'])
    
    return validation_results


def _preview_excel_import(request, form):
    """Preview Excel file with validation results."""
    validation_results = _validate_excel_file(form)
    
    # Get template download URL
    template_url = request.build_absolute_uri('/') + 'worker_eform.xlsx'
    
    context = {
        'form': form,
        'template_url': template_url,
        'validation_results': validation_results,
        'is_preview': True,
        'page_title': 'Import Workers - Validation Preview',
    }
    
    return render(request, 'zone/worker_import_compact.html', context)


def _validate_worker_row(row, row_number):
    """Validate a single worker row and return validation results."""
    
    validation = {
        'row_number': row_number,
        'is_valid': True,
        'is_duplicate': False,
        'errors': [],
        'warnings': [],
        'data': {},
    }
    
    # Required fields validation
    required_fields = ['NAME', 'SEX', 'NATIONALITY']
    
    for field in required_fields:
        value = row.get(field)
        if pd.isna(value) or str(value).strip() == '':
            validation['errors'].append(f"Row {row_number}: Missing required field '{field}'")
            validation['is_valid'] = False
    
    # Extract and validate worker data
    try:
        worker_data = _extract_worker_data(row)
        validation['data'] = worker_data
        
        # Validate specific fields
        
        # Name validation
        if worker_data.get('first_name'):
            if len(worker_data['first_name']) < 2:
                validation['warnings'].append(f"Row {row_number}: First name seems too short")
        
        # Sex validation
        sex = str(row.get('SEX', '')).strip().upper()
        if sex and sex not in ['M', 'MALE', 'F', 'FEMALE']:
            validation['errors'].append(f"Row {row_number}: Invalid sex value '{sex}'. Use M/Male or F/Female")
            validation['is_valid'] = False
            
        # Phone number validation
        phone = str(row.get('PHONE_NUMBER', '')).strip()
        if phone and phone != 'nan':
            # Basic phone validation - should be digits with possible + or spaces
            import re
            # Clean phone for validation
            clean_phone = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('+', '')
            
            if not clean_phone.isdigit():
                validation['errors'].append(f"Row {row_number}: Phone number '{phone}' contains invalid characters")
                validation['is_valid'] = False
            elif len(clean_phone) < 6:
                validation['errors'].append(f"Row {row_number}: Phone number '{phone}' is too short (minimum 6 digits)")
                validation['is_valid'] = False
            elif len(clean_phone) > 15:
                validation['warnings'].append(f"Row {row_number}: Phone number '{phone}' seems too long (more than 15 digits)")
        
        # Nationality validation
        nationality = str(row.get('NATIONALITY', '')).strip()
        if nationality and len(nationality) < 2:
            validation['warnings'].append(f"Row {row_number}: Nationality seems too short")
            
        # Date validation
        dob = row.get('DOB')
        if not pd.isna(dob) and isinstance(dob, str):
            try:
                pd.to_datetime(dob)
            except:
                validation['warnings'].append(f"Row {row_number}: Date of birth format may be invalid")
        
        joined_date = row.get('JOINEDDATE')
        if not pd.isna(joined_date) and isinstance(joined_date, str):
            try:
                pd.to_datetime(joined_date)
            except:
                validation['warnings'].append(f"Row {row_number}: Joined date format may be invalid")
        
        # Check for duplicate workers
        duplicate_info = _is_duplicate_worker(worker_data)
        if duplicate_info['is_duplicate']:
            validation['is_duplicate'] = True
            validation['duplicate_info'] = duplicate_info
            
            # Determine severity based on match type
            match_type = duplicate_info.get('match_type', 'possible')
            
            if match_type == 'exact':
                # Phone number match - definite duplicate
                for match_reason in duplicate_info['match_reasons']:
                    validation['errors'].append(f"Row {row_number}: ❌ DUPLICATE - {match_reason}")
                    validation['is_valid'] = False
            elif match_type == 'strong':
                # Name + DOB match - very likely duplicate
                for match_reason in duplicate_info['match_reasons']:
                    validation['errors'].append(f"Row {row_number}: ⚠️ LIKELY DUPLICATE - {match_reason}")
                    validation['is_valid'] = False
            else:
                # Possible match - warning only
                for match_reason in duplicate_info['match_reasons']:
                    validation['warnings'].append(f"Row {row_number}: ⚠️ POSSIBLE DUPLICATE - {match_reason}")
            
            # Add detailed info about existing workers
            for existing_worker in duplicate_info['matching_workers']:
                details = [
                    f"Name: {existing_worker['name']}",
                    f"ID: {existing_worker['worker_id']}",
                    f"Zone: {existing_worker['zone']}",
                    f"Status: {existing_worker['status']}"
                ]
                if existing_worker.get('phone') and existing_worker['phone'] != 'N/A':
                    details.append(f"Phone: {existing_worker['phone']}")
                if existing_worker.get('dob') and existing_worker['dob'] != 'N/A':
                    details.append(f"DOB: {existing_worker['dob']}")
                
                validation['warnings'].append(
                    f"Row {row_number}: Existing worker - {' | '.join(details)}"
                )
        
        # Position and Department validation
        position = str(row.get('POSITION', '')).strip()
        if position and position != 'nan':
            # Check if position exists
            if not Position.objects.filter(name__iexact=position).exists():
                validation['warnings'].append(f"Row {row_number}: Position '{position}' doesn't exist and will be created")
        
        department = str(row.get('DEPARTMENT', '')).strip()
        if department and department != 'nan':
            # Check if department exists
            if not Department.objects.filter(name__iexact=department).exists():
                validation['warnings'].append(f"Row {row_number}: Department '{department}' doesn't exist and will be created")
        
        # Zone validation
        zone_name = str(row.get('ZONE_NAME', '')).strip()
        if zone_name and zone_name != 'nan':
            if not Zone.objects.filter(name__iexact=zone_name).exists():
                validation['errors'].append(f"Row {row_number}: Zone '{zone_name}' doesn't exist")
                validation['is_valid'] = False
        
        # Building and Floor validation (handles format like "B14-F1")
        building_floor = str(row.get('BUILDINGNAME', '')).strip()
        if building_floor and building_floor != 'nan':
            # Parse building and floor from format like "B14-F1"
            if '-' in building_floor:
                parts = building_floor.split('-', 1)
                building_name = parts[0].strip()
                floor_name = parts[1].strip() if len(parts) > 1 else ''
            else:
                building_name = building_floor
                floor_name = ''
            
            # Validate building exists
            if building_name:
                if not Building.objects.filter(name__iexact=building_name).exists():
                    validation['errors'].append(f"Row {row_number}: Building '{building_name}' doesn't exist")
                    validation['is_valid'] = False
                elif floor_name:
                    # If building exists, validate floor exists for that building
                    building = Building.objects.filter(name__iexact=building_name).first()
                    if building and not Floor.objects.filter(building=building, name__iexact=floor_name).exists():
                        validation['warnings'].append(f"Row {row_number}: Floor '{floor_name}' doesn't exist in building '{building_name}'")
                        # Not setting is_valid to False for floor as it might be optional
        
    except Exception as e:
        validation['errors'].append(f"Row {row_number}: Data extraction error - {str(e)}")
        validation['is_valid'] = False
    
    return validation


def _process_excel_import(request, form):
    """Process the Excel file import."""
    import tempfile
    import json
    from django.core.files.base import ContentFile
    from sinosecu.models import PassportScan, ScanImage, ScanResult
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from PIL import Image as PILImage
    import io
    
    excel_file = form.cleaned_data['excel_file']
    process_photos = form.cleaned_data.get('process_photos', True)
    skip_duplicates = form.cleaned_data.get('skip_duplicates', True)
    
    results = {
        'total_rows': 0,
        'processed': 0,
        'skipped': 0,
        'duplicates': [],  # Track duplicate details separately
        'errors': [],
        'created_workers': [],
    }
    
    tmp_file_name = None
    try:
        # Save Excel file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file_name = tmp_file.name
            for chunk in excel_file.chunks():
                tmp_file.write(chunk)
            tmp_file.flush()
        
        # Extract embedded images from Excel
        embedded_images = _extract_embedded_images_from_excel(tmp_file_name)
        
        # Read Excel file
        df = pd.read_excel(tmp_file_name, sheet_name='Workers')
        
        # Filter out empty rows
        df = df.dropna(subset=['NAME'])
        results['total_rows'] = len(df)
        
        # Process each row
        for index, row in df.iterrows():
            try:
                worker_data = _extract_worker_data(row)
                
                # Check for duplicates
                duplicate_info = _is_duplicate_worker(worker_data)
                if skip_duplicates and duplicate_info['is_duplicate']:
                    results['skipped'] += 1
                    # Store duplicate details separately for clearer messaging
                    duplicate_detail = {
                        'row': index + 2,
                        'worker_name': f"{worker_data['first_name']} {worker_data['last_name']}",
                        'matching_workers': duplicate_info['matching_workers'],
                        'reasons': duplicate_info['match_reasons']
                    }
                    results['duplicates'].append(duplicate_detail)
                    continue
                
                # Create worker with transaction
                with transaction.atomic():
                    worker = _create_worker_from_data(worker_data, request.user)
                    
                    # Process photos if enabled
                    if process_photos:
                        # Pass embedded images along with row data
                        _process_worker_photos(worker, row, request, embedded_images, index)
                    else:
                        pass
                    
                    results['processed'] += 1
                    results['created_workers'].append({
                        'name': worker.get_full_name(),
                        'worker_id': worker.worker_id,
                    })
                    
            except Exception as e:
                error_msg = f"Row {index + 2}: {str(e)}"
                results['errors'].append(error_msg)
    
    except Exception as e:
        results['errors'].append(f"File processing error: {str(e)}")
    finally:
        # Clean up temp file
        if tmp_file_name and os.path.exists(tmp_file_name):
            try:
                os.unlink(tmp_file_name)
            except OSError as e:
                pass
    
    # Show results with detailed messages
    if results['processed'] > 0:
        messages.success(request, f"✓ Successfully imported {results['processed']} new worker(s).")
    
    # Show detailed duplicate information
    if results['duplicates']:
        duplicate_msg = f"⚠ Skipped {results['skipped']} duplicate worker(s):"
        messages.warning(request, duplicate_msg)
        
        # Add details for each duplicate (limit to first 5 for readability)
        for dup in results['duplicates'][:5]:
            existing_info = []
            for worker in dup['matching_workers']:
                existing_info.append(f"{worker['name']} (ID: {worker['worker_id']}, Zone: {worker['zone']})")
            
            detail_msg = f"• Row {dup['row']}: {dup['worker_name']} matches existing: {', '.join(existing_info)}"
            messages.info(request, detail_msg)
        
        if len(results['duplicates']) > 5:
            messages.info(request, f"... and {len(results['duplicates']) - 5} more duplicate(s)")
    
    # Show errors if any
    if results['errors']:
        error_summary = f"✗ Encountered {len(results['errors'])} error(s):"
        messages.error(request, error_summary)
        
        # Show first 3 errors
        for error in results['errors'][:3]:
            messages.error(request, f"• {error}")
        
        if len(results['errors']) > 3:
            messages.error(request, f"... and {len(results['errors']) - 3} more error(s)")
    
    # Summary message
    total_attempted = results['total_rows']
    if total_attempted > 0:
        success_rate = (results['processed'] / total_attempted) * 100
        summary = f"Import Summary: {results['processed']}/{total_attempted} workers imported ({success_rate:.0f}% success rate)"
        if results['skipped'] > 0:
            summary += f", {results['skipped']} duplicates skipped"
        messages.info(request, summary)
    
    # Store import results in session for display on worker list page
    request.session['import_results'] = {
        'total_rows': results['total_rows'],
        'processed': results['processed'],
        'skipped': results['skipped'],
        'duplicates': [
            {
                'row': dup['row'],
                'worker_name': dup['worker_name'],
                'matches': ', '.join([f"{w['name']} (ID: {w['worker_id']})" for w in dup['matching_workers'][:2]])
            }
            for dup in results['duplicates'][:10]  # Store first 10 for display
        ],
        'errors': results['errors'][:5],  # Store first 5 errors
        'timestamp': timezone.now().isoformat()
    }
    
    return redirect('zone:worker_list')


def _extract_worker_data(row):
    """Extract worker data from Excel row."""
    
    # Split full name into first and last name
    full_name = str(row.get('NAME', '')).strip()
    name_parts = full_name.split(' ', 1)
    first_name = name_parts[0] if name_parts else ''
    last_name = name_parts[1] if len(name_parts) > 1 else ''
    
    # Convert DOB to proper date format
    dob = row.get('DOB')
    if pd.isna(dob):
        dob = None
    elif isinstance(dob, str):
        try:
            dob = pd.to_datetime(dob).date()
        except:
            dob = None
    
    # Convert joined date
    joined_date = row.get('JOINEDDATE')
    if pd.isna(joined_date):
        joined_date = timezone.now().date()
    elif isinstance(joined_date, str):
        try:
            joined_date = pd.to_datetime(joined_date).date()
        except:
            joined_date = timezone.now().date()
    
    # Map status
    status_value = str(row.get('ACTIVE/NOACTIVE', 'ACTIVE')).upper()
    status = 'active' if status_value == 'ACTIVE' else 'inactive'
    
    # Parse building and floor from format like "B14-F1"
    building_floor = str(row.get('BUILDINGNAME', '')).strip()
    building_name = ''
    floor_name = ''
    
    if building_floor and building_floor != 'nan':
        if '-' in building_floor:
            parts = building_floor.split('-', 1)
            building_name = parts[0].strip()
            floor_name = parts[1].strip() if len(parts) > 1 else ''
        else:
            building_name = building_floor
    
    return {
        'first_name': first_name,
        'last_name': last_name,
        'sex': str(row.get('SEX', '')).upper(),
        'dob': dob,
        'nationality': str(row.get('NATIONALITY', '')),
        'phone_number': str(row.get('PHONE_NUMBER', '')),
        'zone_name': str(row.get('ZONE_NAME', '')),
        'building_name': building_name,
        'floor_name': floor_name,
        'position_name': str(row.get('POSITION', '')),
        'joined_date': joined_date,
        'status': status,
        'photo_filename': str(row.get('PHOTO', '')),
        'id_document_filename': str(row.get('IDKH/PASSPORT', '')),
        'work_permit_filename': str(row.get('WORKINGPERMIT', '')),
        'visa_filename': str(row.get('VISA', '')),
    }


def _is_duplicate_worker(worker_data):
    """Check if worker already exists and return detailed information."""
    duplicate_info = {
        'is_duplicate': False,
        'matching_workers': [],
        'match_reasons': [],
        'match_type': None  # 'exact', 'strong', 'possible'
    }
    
    matched_worker_ids = set()  # Track already matched workers to avoid duplicates
    
    # 1. Check by phone number (high confidence match)
    if worker_data['phone_number'] and worker_data['phone_number'] != 'nan' and worker_data['phone_number'].strip():
        # Clean phone number for comparison
        clean_phone = worker_data['phone_number'].strip().replace(' ', '').replace('-', '').replace('+', '')
        if clean_phone:
            phone_matches = Worker.objects.filter(phone_number__icontains=clean_phone)
            if phone_matches.exists():
                duplicate_info['is_duplicate'] = True
                duplicate_info['match_type'] = 'exact'
                for worker in phone_matches:
                    if worker.id not in matched_worker_ids:
                        matched_worker_ids.add(worker.id)
                        duplicate_info['matching_workers'].append({
                            'id': worker.id,
                            'name': worker.get_full_name(),
                            'worker_id': worker.worker_id,
                            'zone': worker.zone.name if worker.zone else 'N/A',
                            'status': worker.status,
                            'phone': worker.phone_number,
                            'dob': worker.dob.strftime('%Y-%m-%d') if worker.dob else 'N/A'
                        })
                duplicate_info['match_reasons'].append(f"Phone number '{worker_data['phone_number']}' already exists")
    
    # 2. Check by DOB + Name combination (high confidence match)
    if worker_data['dob'] and worker_data['first_name'] and worker_data['last_name']:
        dob_name_matches = Worker.objects.filter(
            dob=worker_data['dob'],
            first_name__iexact=worker_data['first_name'],
            last_name__iexact=worker_data['last_name']
        )
        if dob_name_matches.exists():
            duplicate_info['is_duplicate'] = True
            if not duplicate_info['match_type']:
                duplicate_info['match_type'] = 'strong'
            for worker in dob_name_matches:
                if worker.id not in matched_worker_ids:
                    matched_worker_ids.add(worker.id)
                    duplicate_info['matching_workers'].append({
                        'id': worker.id,
                        'name': worker.get_full_name(),
                        'worker_id': worker.worker_id,
                        'zone': worker.zone.name if worker.zone else 'N/A',
                        'status': worker.status,
                        'phone': worker.phone_number if worker.phone_number else 'N/A',
                        'dob': worker.dob.strftime('%Y-%m-%d') if worker.dob else 'N/A'
                    })
            dob_str = worker_data['dob'].strftime('%Y-%m-%d') if hasattr(worker_data['dob'], 'strftime') else str(worker_data['dob'])
            duplicate_info['match_reasons'].append(
                f"Same name '{worker_data['first_name']} {worker_data['last_name']}' with DOB {dob_str} already exists"
            )
    
    # 3. Check by exact name match only (medium confidence)
    elif worker_data['first_name'] and worker_data['last_name']:
        name_matches = Worker.objects.filter(
            first_name__iexact=worker_data['first_name'],
            last_name__iexact=worker_data['last_name']
        )
        if name_matches.exists():
            # Check if any have matching DOB (if DOB provided)
            strong_match = False
            if worker_data['dob']:
                for worker in name_matches:
                    if worker.dob == worker_data['dob']:
                        strong_match = True
                        break
            
            if strong_match or len(name_matches) == 1:  # Single name match or DOB match
                duplicate_info['is_duplicate'] = True
                if not duplicate_info['match_type']:
                    duplicate_info['match_type'] = 'strong' if strong_match else 'possible'
                for worker in name_matches:
                    if worker.id not in matched_worker_ids:
                        matched_worker_ids.add(worker.id)
                        duplicate_info['matching_workers'].append({
                            'id': worker.id,
                            'name': worker.get_full_name(),
                            'worker_id': worker.worker_id,
                            'zone': worker.zone.name if worker.zone else 'N/A',
                            'status': worker.status,
                            'phone': worker.phone_number if worker.phone_number else 'N/A',
                            'dob': worker.dob.strftime('%Y-%m-%d') if worker.dob else 'N/A'
                        })
                duplicate_info['match_reasons'].append(
                    f"Name '{worker_data['first_name']} {worker_data['last_name']}' already exists"
                )
    
    # 4. Check by DOB + partial name (possible match)
    if worker_data['dob'] and (worker_data['first_name'] or worker_data['last_name']):
        dob_matches = Worker.objects.filter(dob=worker_data['dob'])
        
        for worker in dob_matches:
            # Check for partial name match
            name_match = False
            if worker_data['first_name'] and worker.first_name:
                if worker_data['first_name'].lower() in worker.first_name.lower() or \
                   worker.first_name.lower() in worker_data['first_name'].lower():
                    name_match = True
            if worker_data['last_name'] and worker.last_name:
                if worker_data['last_name'].lower() in worker.last_name.lower() or \
                   worker.last_name.lower() in worker_data['last_name'].lower():
                    name_match = True
            
            if name_match and worker.id not in matched_worker_ids:
                duplicate_info['is_duplicate'] = True
                if not duplicate_info['match_type']:
                    duplicate_info['match_type'] = 'possible'
                matched_worker_ids.add(worker.id)
                duplicate_info['matching_workers'].append({
                    'id': worker.id,
                    'name': worker.get_full_name(),
                    'worker_id': worker.worker_id,
                    'zone': worker.zone.name if worker.zone else 'N/A',
                    'status': worker.status,
                    'phone': worker.phone_number if worker.phone_number else 'N/A',
                    'dob': worker.dob.strftime('%Y-%m-%d') if worker.dob else 'N/A'
                })
                if f"Similar worker with same DOB" not in str(duplicate_info['match_reasons']):
                    dob_str = worker_data['dob'].strftime('%Y-%m-%d') if hasattr(worker_data['dob'], 'strftime') else str(worker_data['dob'])
                    duplicate_info['match_reasons'].append(f"Similar worker with same DOB {dob_str} found")
    
    return duplicate_info


def _create_worker_from_data(worker_data, created_by):
    """Create worker from extracted data."""
    
    # Get or create zone
    zone, created = Zone.objects.get_or_create(
        name=worker_data['zone_name'],
        defaults={
            'created_by': created_by,
            'is_active': True
        }
    )
    if created:
        pass
    
    # Get or create building
    building, created = Building.objects.get_or_create(
        name=worker_data['building_name'],
        zone=zone,  # Also match by zone to avoid conflicts
        defaults={
            'created_by': created_by,
            'is_active': True,
            'address': f"Building {worker_data['building_name']} in {zone.name}",
            'total_floors': 10  # Default value
        }
    )
    if created:
        pass
    
    # Get floor (optional)
    floor = None
    if worker_data.get('floor_name') and building:
        floor_name = worker_data['floor_name']
        try:
            # First try matching by name (exact match)
            floor = Floor.objects.get(building=building, name__iexact=floor_name)
        except Floor.DoesNotExist:
            try:
                # Try extracting floor number from various formats
                floor_number = None
                
                if floor_name.upper().startswith('F') and floor_name[1:].isdigit():
                    # Format like "F1", "F11"
                    floor_number = int(floor_name[1:])
                elif floor_name.isdigit():
                    # Format like "1", "11"
                    floor_number = int(floor_name)
                elif floor_name.lower().startswith('floor') and floor_name[5:].strip().isdigit():
                    # Format like "Floor 1", "Floor 11"
                    floor_number = int(floor_name[5:].strip())
                
                if floor_number:
                    # Try to find floor by number first
                    try:
                        floor = Floor.objects.get(building=building, floor_number=floor_number)
                    except Floor.DoesNotExist:
                        # If no exact floor number match, create one with the expected name format
                        floor, created = Floor.objects.get_or_create(
                            building=building,
                            floor_number=floor_number,
                            defaults={
                                'name': f'F{floor_number}',
                                'description': f'Floor {floor_number}',
                                'created_by': None  # Set to None as we don't have user context here
                            }
                        )
                        if created:
                            pass
                            
            except (ValueError, TypeError):
                # Floor is optional, so we don't raise an error
                pass
    
    # Get position (optional)
    position = None
    if worker_data['position_name']:
        try:
            position = Position.objects.get(name__iexact=worker_data['position_name'])
        except Position.DoesNotExist:
            pass  # Position is optional
    
    # Validate required fields
    if not worker_data['first_name']:
        raise ValueError("First name is required")
    if not worker_data['last_name']:
        raise ValueError("Last name is required")
    if not worker_data['sex'] or worker_data['sex'] not in ['M', 'F', 'O']:
        raise ValueError("Valid sex (M/F/O) is required")
    if not worker_data['nationality']:
        raise ValueError("Nationality is required")
    # Phone number is optional since not all Excel files have this field
    
    # Create worker
    worker = Worker.objects.create(
        first_name=worker_data['first_name'],
        last_name=worker_data['last_name'],
        sex=worker_data['sex'],
        dob=worker_data['dob'],
        nationality=worker_data['nationality'],
        phone_number=worker_data['phone_number'],
        zone=zone,
        building=building,
        floor=floor,
        position=position,
        date_joined=worker_data['joined_date'],
        status=worker_data['status'],
        created_by=created_by,
    )
    
    return worker


def _process_worker_photos(worker, row, request, embedded_images=None, row_index=0):
    """Process worker photos from embedded images in Excel cells."""
    
    # Map column names to field types and column letters
    photo_columns = {
        'photo': ('PHOTO', 'Q'),
        'id_document': ('IDKH/PASSPORT', 'N'),
        'work_permit': ('WORKINGPERMIT', 'O'),
        'visa': ('VISA', 'P'),
    }
    
    for field_type, (column_name, column_letter) in photo_columns.items():
        # Check for embedded image in this cell
        if embedded_images:
            # IMPORTANT: row_index is the pandas DataFrame index after dropna()
            # pandas.read_excel() creates DataFrame where index 0 = Excel row 2 (first data row)
            # After dropna(), the original index is preserved
            # So pandas index N corresponds to Excel row N+2 (N+1 for 0->1 based, +1 for header)
            excel_row = row_index + 2  # Convert pandas index to Excel row number
            cell_ref = f"{column_letter}{excel_row}"
            
            
            if cell_ref in embedded_images:
                # Process embedded image
                try:
                    image_data = embedded_images[cell_ref]
                    
                    if field_type == 'photo':
                        _process_worker_portrait_from_bytes(worker, image_data, f"worker_{worker.id}_{field_type}.jpg")
                    else:
                        _process_document_from_bytes_with_sinosecu(worker, image_data, field_type, request.user,'', [])
                except Exception as e:
                    pass
            else:
                pass
                # Log if no image found in expected cell


def _extract_embedded_images_from_excel(excel_path):
    """
    Extract embedded images from Excel file.
    Returns a dictionary mapping cell references to image data.
    Example: {'N2': image_bytes, 'O2': image_bytes, ...}
    """
    from openpyxl import load_workbook
    import zipfile
    import io
    
    embedded_images = {}
    
    try:
        # Primary method: Use openpyxl-image-loader if available
        try:
            from openpyxl_image_loader import SheetImageLoader
            
            wb = load_workbook(excel_path)
            ws = wb['Workers'] if 'Workers' in wb.sheetnames else wb.active
            
            # Use SheetImageLoader to extract images
            image_loader = SheetImageLoader(ws)
            
            # Check cells where we expect images (N, O, P, Q columns)
            for row in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                for col_letter, col_name in [('N', 'IDKH/PASSPORT'), ('O', 'WORKINGPERMIT'), 
                                             ('P', 'VISA'), ('Q', 'PHOTO')]:
                    cell_ref = f"{col_letter}{row}"
                    if image_loader.image_in(cell_ref):
                        try:
                            # Get PIL Image
                            pil_image = image_loader.get(cell_ref)
                            # Convert to bytes
                            img_byte_arr = io.BytesIO()
                            pil_image.save(img_byte_arr, format="JPEG")
                            embedded_images[cell_ref] = img_byte_arr.getvalue()
                        except Exception as e:
                            pass
                    # else:
                    #     logger.debug(f"No image in cell {cell_ref}")
            
            wb.close()
            
        except ImportError:
            
            # Alternative method: Extract from Excel ZIP structure
            with zipfile.ZipFile(excel_path, 'r') as zip_file:
                # Get list of image files in the Excel package
                image_files = [f for f in zip_file.namelist() if f.startswith('xl/media/')]
                
                # Extract drawing relationships to map images to cells
                if 'xl/drawings/_rels/drawing1.xml.rels' in zip_file.namelist():
                    rels_content = zip_file.read('xl/drawings/_rels/drawing1.xml.rels')
                    # Parse relationships to map images
                    # This is simplified - full implementation would parse XML properly
                    
                # Extract all images (simplified approach)
                for idx, image_file in enumerate(image_files):
                    image_data = zip_file.read(image_file)
                    # Map to approximate cell (this is a simplified mapping)
                    # In production, parse drawing1.xml to get exact cell mappings
                    row = 2 + (idx // 4)  # Distribute across rows
                    col_idx = idx % 4
                    col_letter = ['N', 'O', 'P', 'Q'][col_idx]
                    cell_ref = f"{col_letter}{row}"
                    embedded_images[cell_ref] = image_data
        
    except Exception as e:
        pass
    
    return embedded_images


def _find_photo_file(filename_or_path, photos_folder):
    """
    Find photo file from Excel cell value (filename or path).
    Excel cells can contain:
    - Just filename: 'john_passport.jpg'
    - Relative path: 'documents/john_passport.jpg'  
    - Full path: 'C:/Photos/john_passport.jpg'
    """
    
    if not filename_or_path or pd.isna(filename_or_path):
        return None
        
    filename_or_path = str(filename_or_path).strip()
    search_paths = []
    
    # 1. If it's already a full absolute path, check it directly
    if os.path.isabs(filename_or_path):
        if os.path.exists(filename_or_path):
            return filename_or_path
        search_paths.append(filename_or_path)
    
    # 2. If it contains path separators, treat as relative path
    elif '/' in filename_or_path or '\\' in filename_or_path:
        # Relative path from photos_folder
        if photos_folder:
            search_paths.append(os.path.join(photos_folder, filename_or_path))
        
        # Relative path from media root
        media_root = getattr(settings, 'MEDIA_ROOT', '')
        if media_root:
            search_paths.extend([
                os.path.join(media_root, filename_or_path),
                os.path.join(media_root, 'worker_photos', filename_or_path),
                os.path.join(media_root, 'worker_documents', filename_or_path),
            ])
    
    # 3. Just a filename - search in various folders
    else:
        # Custom photos folder
        if photos_folder:
            search_paths.extend([
                os.path.join(photos_folder, filename_or_path),
                os.path.join(photos_folder, 'passport', filename_or_path),
                os.path.join(photos_folder, 'visa', filename_or_path),
                os.path.join(photos_folder, 'nid', filename_or_path),
                os.path.join(photos_folder, 'work_permit', filename_or_path),
            ])
        
        # Default media folders
        media_root = getattr(settings, 'MEDIA_ROOT', '')
        if media_root:
            search_paths.extend([
                os.path.join(media_root, 'worker_photos', filename_or_path),
                os.path.join(media_root, 'worker_documents', filename_or_path),
                os.path.join(media_root, 'temp_uploads', filename_or_path),
                # Document type specific folders
                os.path.join(media_root, 'documents', 'passport', filename_or_path),
                os.path.join(media_root, 'documents', 'visa', filename_or_path),
                os.path.join(media_root, 'documents', 'nid', filename_or_path),
                os.path.join(media_root, 'documents', 'work_permit', filename_or_path),
            ])
        
        # Current working directory as fallback
        search_paths.extend([
            os.path.join(os.getcwd(), filename_or_path),
            os.path.join(os.getcwd(), 'photos', filename_or_path),
        ])
    
    # Find existing file
    for path in search_paths:
        if os.path.exists(path):
            return path
    
    # Log all attempted paths for debugging
    return None


def _process_worker_portrait(worker, photo_path):
    """Process worker portrait photo from file path."""
    
    try:
        with open(photo_path, 'rb') as f:
            photo_content = f.read()
        
        # Get filename
        filename = os.path.basename(photo_path)
        
        # Save to worker photo field
        worker.photo.save(filename, ContentFile(photo_content), save=True)
        
    except Exception as e:
        pass


def _process_worker_portrait_from_bytes(worker, image_bytes, filename):
    """Process worker portrait photo from bytes (embedded image)."""
    
    try:
        # Save to worker photo field
        worker.photo.save(filename, ContentFile(image_bytes), save=True)
        
    except Exception as e:
        pass


def _process_document_with_sinosecu(worker, photo_path, document_type, created_by):
    """Process document photo using SinoSecu API for faster processing."""
    
    try:
        # Import SinoSecu service
        from sinosecu.services import sinosecu_service
        
        # Process document with actual OCR
        scan = sinosecu_service.create_scan_with_ocr(
            user=created_by,
            image_path=photo_path,
            document_type=document_type,
            worker_id=worker.id
        )
        
        if scan.status == 'completed' and hasattr(scan, 'result'):
            # Create document record from OCR results
            _create_document_from_ocr_scan(worker, scan, document_type, photo_path, created_by)
            
            # Update worker info if OCR provided better data
            _update_worker_from_ocr(worker, scan.result)
            
            
        elif scan.status == 'failed':
            # Still create basic document record for file storage
            _create_basic_document_from_scan(worker, scan, document_type, photo_path, created_by)
        
    except Exception as e:
        pass


def _process_document_from_bytes_with_sinosecu(worker, image_bytes, document_type, created_by, mrz_text, mrz_lines):
    """Process document from embedded image bytes using SinoSecu API."""
    
    print("Starting with new extract data...", mrz_text)

    try:
        # Import SinoSecu service
        from sinosecu.services import sinosecu_service
        import tempfile
        
        # Save bytes to temporary file for OCR processing
        with tempfile.NamedTemporaryFile(delete=False, suffix='.jpg') as tmp_file:
            tmp_file.write(image_bytes)
            tmp_file.flush()
            temp_path = tmp_file.name
        
        try:
            # Process with OCR
            scan = sinosecu_service.create_scan_with_ocr(
                user=created_by,
                image_path=temp_path,
                document_type=document_type,
                worker_id=worker.id
            )
            
            if scan.status == 'completed' and hasattr(scan, 'result'):
                # Create document record from OCR results

                # document_date = mrz_extract_image(tmp_file.name)

                _create_document_from_embedded_image(worker, scan, document_type, image_bytes, tmp_file.name,created_by, mrz_text, mrz_lines)
                
                # Update worker info if OCR provided better data
                _update_worker_from_ocr(worker, scan.result)
                
                
            elif scan.status == 'failed':
                # Still create basic document record for file storage
                _create_basic_document_from_embedded(worker, scan, document_type, image_bytes, created_by)
        
        finally:
            # Clean up temp file
            if os.path.exists(temp_path):
                os.unlink(temp_path)
            
    except Exception as e:
        pass


def _create_document_from_embedded_image(worker, scan, document_type, image_bytes, photo_name, created_by, mrz_text, mrz_lines):
    """Create document record from embedded image with OCR/MRZ results."""
    # Call the new function with empty MRZ text for backward compatibility

    return _create_document_from_embedded_image_with_mrz(worker, scan, document_type, image_bytes, photo_name, created_by, mrz_text, mrz_lines)

def _create_document_from_embedded_image_with_mrz(worker, scan, document_type, image_bytes, photo_name, created_by, raw_mrz_text, mrz_lines):
    """Create document record from embedded image with OCR/MRZ results and raw MRZ text."""
    from django.core.files.base import ContentFile
    
    scan_result = scan.result if hasattr(scan, 'result') else None
    filename = f"embedded_{worker.id}_{document_type}_{scan.scan_id}.jpg"
    
    # Parse dates safely
    def parse_date(date_field):
        if date_field:
            return date_field
        return timezone.now().date()
    
    if scan_result:
        pass
    
    # Try to use scan_result data if available
    if scan_result and scan_result.document_number:
        try:
            # Log scan_result data for debugging

            document_mrz = parse_mrz_lines(mrz_lines)

            # Use MRZ extracted data properly
            if(document_mrz):
                print("document_mrz==:", document_mrz)
                doc_number = document_mrz["document_number"]
                expiry_date = document_mrz["expiry_date"] if document_mrz["expiry_date"] else timezone.now().date() + timedelta(days=365)
                issue_date = timezone.now().date()
                issuing_authority = document_mrz["issuing_country"]
                
                # Map issuing country codes to standard nationality codes
                nationality_map = {
                    'KHM': 'KH', 'KH': 'KH', 'CAMBODIA': 'KH',
                    'THA': 'TH', 'TH': 'TH', 'THAILAND': 'TH', 
                    'VNM': 'VN', 'VN': 'VN', 'VIETNAM': 'VN', 'NM': 'VN',
                    'PHL': 'PH', 'PH': 'PH', 'PHILIPPINES': 'PH',
                    'IDN': 'ID', 'ID': 'ID', 'INDONESIA': 'ID',
                    'MMR': 'MM', 'MM': 'MM', 'MYANMAR': 'MM',
                    'LAO': 'LA', 'LA': 'LA', 'LAOS': 'LA',
                    'SGP': 'SG', 'SG': 'SG', 'SINGAPORE': 'SG',
                    'MYS': 'MY', 'MY': 'MY', 'MALAYSIA': 'MY',
                    # OCR misreading corrections
                    'CI': 'ID',   # Common OCR error - CI often misread for IDN (Indonesia)
                    'CID': 'ID',  # Common OCR error - CID often misread for IDN (Indonesia)
                    'CIN': 'ID',  # Common OCR error - CIN often misread for IDN
                    'ID1': 'ID',  # Indonesia - N often misread as 1
                    'ID0': 'ID',  # Indonesia - N often misread as 0
                }
                
                # Special handling for Cambodian ID cards
                if document_mrz.get("document_type") == "id_card" and issuing_authority == "KHM":
                    # For Cambodian ID cards:
                    worker.nationality = "KH"  # Always Cambodian nationality
                    # Keep issuing_authority as KHM for the document
                else:
                    # For passports and other documents, use the nationality from MRZ if available
                    mrz_nationality = document_mrz.get("nationality")
                    if mrz_nationality:
                        # Use the nationality from MRZ parsing (already corrected for OCR errors)
                        mapped_nationality = nationality_map.get(mrz_nationality.upper(), mrz_nationality.upper())
                        worker.nationality = mapped_nationality
                    else:
                        # Fallback to mapping issuing authority if no nationality in MRZ
                        mapped_nationality = nationality_map.get(issuing_authority.upper(), issuing_authority.upper())
                        worker.nationality = mapped_nationality
                    # Update issuing_authority to match nationality for consistency
                    issuing_authority = mapped_nationality
                
                worker.save()
            else:
                doc_number = scan_result.document_number.strip()
                expiry_date = scan_result.expiry_date if scan_result.expiry_date else timezone.now().date() + timedelta(days=365)
                issue_date = scan_result.issue_date if scan_result.issue_date else timezone.now().date()
                issuing_authority = scan_result.issuing_country or scan_result.issuing_authority or "MRZ Extracted"
            
            
            # Use provided raw MRZ text if available, otherwise extract from scan_result
            if raw_mrz_text:
                # Use the raw MRZ text passed from the MRZ processing
                mrz_display = raw_mrz_text
            else:
                # Fallback to extracting from scan_result (for backward compatibility)
                mrz_lines = []
                if hasattr(scan_result, 'mrz_line1') and scan_result.mrz_line1:
                    mrz_lines.append(scan_result.mrz_line1)
                if hasattr(scan_result, 'mrz_line2') and scan_result.mrz_line2:
                    mrz_lines.append(scan_result.mrz_line2)
                if hasattr(scan_result, 'mrz_line3') and scan_result.mrz_line3:
                    mrz_lines.append(scan_result.mrz_line3)
                
                mrz_display = "\n".join(mrz_lines) if mrz_lines else "No MRZ lines available"
            
            # Create the note with just the MRZ data (no header)
            if raw_mrz_text and raw_mrz_text.strip():
                # If we have actual MRZ text, use it directly
                note_text = raw_mrz_text.strip()
            elif mrz_lines and any(line.strip() for line in mrz_lines):
                # If we have MRZ lines from scan_result, use those
                note_text = "\n".join(line.strip() for line in mrz_lines if line.strip())
            else:
                # No MRZ data available - leave note empty
                note_text = ""
            
            # Override document type mapping if KHM is detected in MRZ
            if note_text and ('KHM' in note_text.upper() or 'IDKHM' in note_text.upper() or 'LDKHM' in note_text.upper()):
                mapped_doc_type = 'id_card'
            else:
                mapped_doc_type = _map_document_type(document_type)
            
            document = Document.objects.create(
                worker=worker,
                document_type=mapped_doc_type,
                document_number=doc_number,
                issue_date=issue_date,
                expiry_date=expiry_date,
                issuing_authority=issuing_authority,
                notes=note_text,
                created_by=created_by,
            )
        except Exception as e:
            # Fallback to basic document - still try to save raw MRZ if available
            fallback_note = ""
            if raw_mrz_text and raw_mrz_text.strip():
                fallback_note = raw_mrz_text.strip()
            
            # Override document type mapping if KHM is detected in MRZ
            if fallback_note and ('KHM' in fallback_note.upper() or 'IDKHM' in fallback_note.upper() or 'LDKHM' in fallback_note.upper()):
                mapped_doc_type = 'id_card'
            else:
                mapped_doc_type = _map_document_type(document_type)
            
            document = Document.objects.create(
                worker=worker,
                document_type=mapped_doc_type,
                document_number=f"IMPORT_{scan.scan_id}",
                issue_date=timezone.now().date(),
                expiry_date=timezone.now().date() + timedelta(days=365),
                issuing_authority="Pending Verification",
                notes=fallback_note,
                created_by=created_by,
            )
    else:
        # No scan result, create basic document - still save raw MRZ if provided
        basic_note = ""
        if raw_mrz_text and raw_mrz_text.strip():
            basic_note = raw_mrz_text.strip()
        
        # Override document type mapping if KHM is detected in MRZ
        if basic_note and ('KHM' in basic_note.upper() or 'IDKHM' in basic_note.upper() or 'LDKHM' in basic_note.upper()):
            mapped_doc_type = 'id_card'
        else:
            mapped_doc_type = _map_document_type(document_type)
        
        document = Document.objects.create(
            worker=worker,
            document_type=mapped_doc_type,
            document_number=f"IMPORT_{scan.scan_id}",
            issue_date=timezone.now().date(),
            expiry_date=timezone.now().date() + timedelta(days=365),
            issuing_authority="Pending Verification",
            notes=basic_note,
            created_by=created_by,
        )
    
    # Save document file
    document.document_file.save(filename, ContentFile(image_bytes), save=True)
    
    return document


# create document with MRZ data
def create_document_with_mrz(worker, scan, document_type, image_bytes, created_by, document_record):

    from django.core.files.base import ContentFile
    
    filename = f"embedded_{worker.id}_{document_type}_{scan.scan_id}.jpg"
    
    if document_record:
        print("document_record")
        note_text = ""
        doc_number = ""
        issue_date=""
        expiry_date=""
        issuing_authority=""
        document = Document.objects.create(
                worker=worker,
                document_type=_map_document_type(document_type),
                document_number=doc_number,
                issue_date=issue_date,
                expiry_date=expiry_date,
                issuing_authority=issuing_authority,
                notes=note_text,
                created_by=created_by,
            )
        
    else:
        # No scan result, create basic document - still save raw MRZ if provided
        basic_note = ""        
        document = Document.objects.create(
            worker=worker,
            document_type=_map_document_type(document_type),
            document_number=f"IMPORT_{scan.scan_id}",
            issue_date=timezone.now().date(),
            expiry_date=timezone.now().date() + timedelta(days=365),
            issuing_authority="Pending Verification",
            notes=basic_note,
            created_by=created_by,
        )

    # Save document file
    document.document_file.save(filename, ContentFile(image_bytes), save=True)
    
    return document

def _create_basic_document_from_embedded(worker, scan, document_type, image_bytes, created_by):
    """Create basic document record from embedded image when OCR fails."""
    
    filename = f"embedded_{worker.id}_{document_type}_{scan.scan_id}.jpg"
    
    # Create basic document record
    document = Document.objects.create(
        worker=worker,
        document_type=_map_document_type(document_type),
        document_number=f"MANUAL_VERIFY_{scan.scan_id}",
        issue_date=timezone.now().date(),
        expiry_date=timezone.now().date() + timedelta(days=365),
        issuing_authority="Manual Verification Required",
        notes=f"Embedded image from Excel - OCR failed: {scan.error_message}. Manual verification required.",
        created_by=created_by,
    )
    
    # Save document file
    document.document_file.save(filename, ContentFile(image_bytes), save=True)
    
    return document


def _map_document_type(field_type):
    """Map field type to document type."""
    mapping = {
        'id_document': 'passport',  # IDKH/PASSPORT column contains passports, not ID cards
        'work_permit': 'work_permit', 
        'visa': 'visa',
    }
    return mapping.get(field_type, 'other')


def _create_document_from_ocr_scan(worker, scan, document_type, photo_path, created_by):
    """Create document record from OCR scan results."""
    
    filename = os.path.basename(photo_path)
    scan_result = scan.result
    
    # Read photo content
    with open(photo_path, 'rb') as f:
        photo_content = f.read()
    
    # Parse dates safely
    def parse_date(date_field):
        if date_field:
            return date_field
        return timezone.now().date()
    
    # Create document record with OCR extracted data
    document = Document.objects.create(
        worker=worker,
        document_type=_map_document_type(document_type),
        document_number=scan_result.document_number or f"IMPORTED_{scan.scan_id}",
        issue_date=parse_date(scan_result.issue_date),
        expiry_date=parse_date(scan_result.expiry_date),
        issuing_authority=scan_result.issuing_authority or "OCR Extracted",
        notes=f"Imported from Excel via SinoSecu OCR scan {scan.scan_id} (Confidence: {scan_result.ocr_confidence:.2f})",
        created_by=created_by,
    )
    
    # Save document file
    document.document_file.save(filename, ContentFile(photo_content), save=True)
    
    return document


def _create_basic_document_from_scan(worker, scan, document_type, photo_path, created_by):
    """Create basic document record when OCR fails."""
    
    filename = os.path.basename(photo_path)
    
    # Read photo content
    with open(photo_path, 'rb') as f:
        photo_content = f.read()
    
    # Create basic document record
    document = Document.objects.create(
        worker=worker,
        document_type=_map_document_type(document_type),
        document_number=f"MANUAL_VERIFY_{scan.scan_id}",  # Needs manual verification
        issue_date=timezone.now().date(),
        expiry_date=timezone.now().date() + timedelta(days=365),  # Default 1 year
        issuing_authority="Manual Verification Required",
        notes=f"Imported from Excel - OCR failed: {scan.error_message}. Manual verification required.",
        created_by=created_by,
    )
    
    # Save document file
    document.document_file.save(filename, ContentFile(photo_content), save=True)
    
    return document


def _update_worker_from_ocr(worker, scan_result):
    """Update worker information from OCR results if data is missing or empty."""
    
    updated_fields = []
    
    # Update first name if empty
    if not worker.first_name.strip() and scan_result.given_names:
        worker.first_name = scan_result.given_names
        updated_fields.append('first_name')
    
    # Update last name if empty
    if not worker.last_name.strip() and scan_result.surname:
        worker.last_name = scan_result.surname
        updated_fields.append('last_name')
    
    # Update sex if empty
    if not worker.sex and scan_result.gender:
        # Map OCR gender to worker sex choices
        gender_map = {'M': 'M', 'F': 'F', 'MALE': 'M', 'FEMALE': 'F'}
        mapped_gender = gender_map.get(scan_result.gender.upper())
        if mapped_gender:
            worker.sex = mapped_gender
            updated_fields.append('sex')
    
    # Update date of birth if missing
    if not worker.dob and scan_result.date_of_birth:
        worker.dob = scan_result.date_of_birth
        updated_fields.append('dob')
    
    # Update nationality if empty
    if not worker.nationality and scan_result.nationality:
        # Try to map nationality codes
        nationality_map = {
            'KHM': 'KH', 'CAMBODIA': 'KH', 'KH': 'KH',
            'THA': 'TH', 'THAILAND': 'TH', 'TH': 'TH',
            'VNM': 'VN', 'VIETNAM': 'VN', 'VN': 'VN', 'NM': 'VN',
            'PHL': 'PH', 'PHILIPPINES': 'PH', 'PH': 'PH',
            'IDN': 'ID', 'INDONESIA': 'ID', 'ID': 'ID',
            'MMR': 'MM', 'MYANMAR': 'MM', 'MM': 'MM',
            'LAO': 'LA', 'LAOS': 'LA', 'LA': 'LA',
            'SGP': 'SG', 'SINGAPORE': 'SG', 'SG': 'SG',
            'MYS': 'MY', 'MALAYSIA': 'MY', 'MY': 'MY',
            # OCR misreading corrections
            'CI': 'ID',  # Common OCR error - CI often misread for IDN (Indonesia)
            'CIN': 'ID', # Common OCR error - CIN often misread for IDN
            'ID1': 'ID', # Indonesia - N often misread as 1
            'ID0': 'ID', # Indonesia - N often misread as 0
        }
        mapped_nationality = nationality_map.get(scan_result.nationality.upper())
        if mapped_nationality:
            worker.nationality = mapped_nationality
            updated_fields.append('nationality')
    
    # Save if any fields were updated
    if updated_fields:
        worker.save(update_fields=updated_fields)

def upload_excel(request):
    
    return render(request, "zone/excel_upload/upload_excel.html")

from openpyxl import load_workbook
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.files.storage import FileSystemStorage

@csrf_exempt
def excel_to_json(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        fs = FileSystemStorage()
        filename = fs.save(excel_file.name, excel_file)
        file_path = fs.path(filename)

        # Define output directory relative to your Django project
        output_dir = os.path.join(settings.BASE_DIR, 'media', 'worker_import_json')
        os.makedirs(output_dir, exist_ok=True)
        
        try:
            workbook = load_workbook(file_path)
            sheet = workbook.active
            
            data = []
            headers = [cell.value for cell in sheet[1]]  # Assuming first row is header
            
            for row in sheet.iter_rows(min_row=2):  # Skip header
                row_data = {}
                for idx, cell in enumerate(row):
                    row_data[headers[idx]] = cell.value
                data.append(row_data)
            
            # Save to JSON file
            json_filename = os.path.splitext(filename)[0] + '.json'
            json_filepath = os.path.join(output_dir, json_filename)
            
            with open(json_filepath, 'w') as json_file:
                json.dump(data, json_file, indent=4, cls=DateTimeEncoder)
            
            return JsonResponse({
                'status': 'success',
                'message': 'File converted successfully',
                'json_file': json_filename
            })
            
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
        
        finally:
            fs.delete(filename)
            
    return JsonResponse({'status': 'error', 'message': 'Invalid request'})

from json import JSONEncoder
from datetime import datetime, date
class DateTimeEncoder(JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        return super().default(obj)


@login_required
def get_import_progress(request, operation_id):
    """Get real-time progress for import operation."""
    progress_data = ImportProgressTracker.get_progress_by_id(operation_id)
    
    if not progress_data:
        return JsonResponse({'error': 'Progress not found'}, status=404)
    
    return JsonResponse(progress_data)

def extract_images_to_folder(excel_file):
    """Extract embedded images from Excel file to media folder."""
    extracted_images = {}
    
    try:
        # Create temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            excel_file.seek(0)  # Reset file pointer
            for chunk in excel_file.chunks():
                tmp_file.write(chunk)
            tmp_file.flush()
            tmp_path = tmp_file.name
        
        try:
            # Create directory for extracted images
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            images_dir = os.path.join(settings.MEDIA_ROOT, 'worker_import_images', timestamp)
            os.makedirs(images_dir, exist_ok=True)
            
            # Extract images using zipfile
            with zipfile.ZipFile(tmp_path, 'r') as zip_ref:
                media_files = [f for f in zip_ref.namelist() if f.startswith('xl/media/')]
                media_files.sort()  # Sort for consistent ordering
                
                
                # Extract each image
                for idx, media_file in enumerate(media_files):
                    try:
                        # Read image data
                        image_data = zip_ref.read(media_file)
                        
                        # Get file extension
                        file_ext = media_file.split('.')[-1].lower()
                        
                        # Create filename
                        filename = f'image_{idx+1:02d}.{file_ext}'
                        file_path = os.path.join(images_dir, filename)
                        
                        # Write image to file
                        with open(file_path, 'wb') as img_file:
                            img_file.write(image_data)
                        
                        # Store reference (use forward slashes for URLs)
                        relative_path = f'worker_import_images/{timestamp}/{filename}'
                        url = f'/media/{relative_path}'
                        
                        extracted_images[idx] = {
                            'filename': filename,
                            'path': file_path,
                            'url': url,
                            'relative_path': relative_path
                        }
                        
                        
                    except Exception as e:
                        pass
                        
        finally:
            # Clean up temp file
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
                
    except Exception as e:
        pass
    
    return extracted_images


def add_image_references_to_records(records, extracted_images):
    """Add image references to JSON records - using dynamic mapping based on Excel structure."""
    if not extracted_images or not records:
        return
        
    
    # Excel structure: Images are in columns N (14), O (15), P (16), Q (17)
    # which correspond to IDKH/PASSPORT, WORKINGPERMIT, VISA, PHOTO
    image_columns = {
        'IDKH/PASSPORT': 14,  # Column N
        'WORKINGPERMIT': 15,  # Column O  
        'VISA': 16,           # Column P
        'PHOTO': 17           # Column Q
    }
    
    # Calculate images per row (4 image columns)
    images_per_row = 4
    
    # Map images to records dynamically
    for record_idx, record in enumerate(records):
        worker_name = record.get('NAME', f'Worker {record_idx + 1}')
        
        # Calculate which images belong to this row
        # Row 0 gets images 0-3, Row 1 gets images 4-7, etc.
        base_image_idx = record_idx * images_per_row
        
        # Map each image column to the corresponding image
        for column_name, col_position in image_columns.items():
            # Calculate the image index for this cell
            # For row 0: IDKH=0, PERMIT=1, VISA=2, PHOTO=3
            # For row 1: IDKH=4, PERMIT=5, VISA=6, PHOTO=7
            relative_position = col_position - 14  # N is column 14, so N=0, O=1, P=2, Q=3
            img_idx = base_image_idx + relative_position
            
            # Check if this image exists
            if img_idx in extracted_images:
                img_info = extracted_images[img_idx]
                record[column_name] = img_info['url']
                filename = img_info.get('filename', f'image_{img_idx+1:02d}')
            else:
                # No image found for this cell
                record[column_name] = ''
    


@login_required 
def simple_excel_import(request):
    """Simple Excel import page with 2 clear steps."""
    return render(request, 'zone/worker_import/simple_excel_import.html')

@login_required
def step1_extract_excel_to_json(request):
    """Step 1: Extract Excel data for preview - simplified without MRZ/OCR processing."""
    if request.method != 'POST':
        # If someone accesses this API endpoint directly, redirect them to the import page
        if request.method == 'GET':
            return redirect('zone:simple_excel_import')
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    try:
        form = WorkerExcelImportForm(request.POST, request.FILES)
        if not form.is_valid():
            # Return detailed error information
            error_details = {}
            for field, errors in form.errors.items():
                error_details[field] = [str(error) for error in errors]
            return JsonResponse({
                'error': 'Invalid form data',
                'details': error_details,
                'success': False
            }, status=400)
        
        # Create progress tracker for extraction only
        progress_tracker = ImportProgressTracker()
        operation_id = progress_tracker.start(
            total_items=1,
            operation_type="excel_extraction"
        )
        
        # Process Excel file directly without complex image extraction
        excel_file = form.cleaned_data['excel_file']
        
        try:
            # Read Excel directly with pandas and convert to clean JSON
            df = pd.read_excel(excel_file, sheet_name=0)
            
            # Extract embedded images from Excel to folder
            extracted_images = extract_images_to_folder(excel_file)
            
            # Convert DataFrame to list of dictionaries (clean JSON representation)
            records = df.to_dict('records')
            
            # Clean up the records - convert NaN to empty strings and handle data types
            for record in records:
                for key, value in record.items():
                    if pd.isna(value):
                        record[key] = ''
                    elif isinstance(value, pd.Timestamp):
                        record[key] = value.strftime('%Y-%m-%d') if value else ''
                    else:
                        record[key] = str(value) if value is not None else ''
            
            # Add image references to records
            add_image_references_to_records(records, extracted_images)
            
            # Store extracted images in session for step 2
            # Make sure the dictionary is JSON serializable
            extracted_images_for_session = {}
            for k, v in extracted_images.items():
                extracted_images_for_session[str(k)] = {
                    'filename': v.get('filename'),
                    'path': v.get('path'),
                    'url': v.get('url')
                }
            request.session['extracted_images'] = extracted_images_for_session
            
            record_count = len(records)
            original_column_order = df.columns.tolist()
            
            # Store data in session for next step
            request.session['excel_data'] = records
            request.session['excel_columns'] = original_column_order
            
            progress_tracker.complete(success=True, final_message=f"Extracted {record_count} records")
            
            
            # Limit preview to first 50 records for performance
            preview_limit = min(50, len(records))
            
            
            # Count extracted images
            image_count = len(extracted_images)
            
            return JsonResponse({
                'success': True,
                'operation_id': operation_id,
                'record_count': record_count,
                'image_count': image_count,
                'message': f'Extracted {record_count} records and {image_count} images',
                'preview_data': records[:preview_limit],  # Return up to 50 records for preview
                'original_column_order': original_column_order,  # Include column order
                'preview_showing': preview_limit,
                'preview_total': len(records),
                'raw_json': records[:preview_limit]  # Include raw JSON for inspection
            })
            
        except ValueError as e:
            # Handle specific validation errors (like invalid Excel files)
            progress_tracker.complete(success=False, final_message=f"Validation failed: {str(e)}")
            return JsonResponse({'error': f'Validation failed: {str(e)}'}, status=400)
        except Exception as e:
            progress_tracker.complete(success=False, final_message=f"Extraction failed: {str(e)}")
            return JsonResponse({'error': f'Extraction failed: {str(e)}'}, status=500)
                    
    except Exception as e:
        return JsonResponse({'error': f'Extraction failed: {str(e)}'}, status=500)


@login_required
def save_edited_preview_data(request):
    """Save edited preview data back to the extracted JSON file."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    try:
        # Get the updated data from the request
        updated_data_str = request.POST.get('updated_data')
        if not updated_data_str:
            return JsonResponse({'error': 'No data provided'}, status=400)
        
        # Parse the updated data
        updated_data = json.loads(updated_data_str)
        
        # Get the JSON path from session
        json_path = request.session.get('excel_json_path')
        if not json_path or not os.path.exists(json_path):
            return JsonResponse({'error': 'No extracted data file found'}, status=400)
        
        # Load the existing JSON file
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # Update the records with the edited data
        if isinstance(json_data, dict) and 'records' in json_data:
            # New structure with metadata
            json_data['records'] = updated_data
        else:
            # Old structure - direct list of records
            json_data = updated_data
        
        # Save the updated data back to the file
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, ensure_ascii=False, indent=2)
        
        return JsonResponse({
            'success': True,
            'message': 'Changes saved successfully',
            'updated_count': len(updated_data)
        })
        
    except json.JSONDecodeError as e:
        return JsonResponse({'error': f'Invalid JSON data: {str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'error': f'Error saving changes: {str(e)}'}, status=500)

@login_required
def step2_process_mrz_and_import(request):
    """Step 2: Import workers directly from Excel data - simplified without MRZ processing."""
    if request.method != 'POST':
        # If someone accesses this API endpoint directly, redirect them to the import page
        if request.method == 'GET':
            return redirect('zone:simple_excel_import')
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    try:
        import os
        
        # Get Excel data from session
        excel_data = request.session.get('excel_data')
        if not excel_data:
            return JsonResponse({'error': 'No extracted data found. Please extract Excel file first.'}, status=400)
            
        # Get extracted images from session
        extracted_images_session = request.session.get('extracted_images', {})
        extracted_images = {}
        for k, v in extracted_images_session.items():
            try:
                # Keep original string keys - don't convert to integers
                extracted_images[k] = v
            except (ValueError, TypeError) as e:
                pass
        
        
        # Create progress tracker for import
        progress_tracker = ImportProgressTracker()
        
        record_count = len(excel_data)
        operation_id = progress_tracker.start(
            total_items=record_count,
            operation_type="simple_import"
        )
        
        
        # Import workers directly from Excel data
        results = import_workers_from_excel_data(excel_data, request, progress_tracker, extracted_images)

        # Clean up extracted image files after successful import
        if extracted_images:
            cleanup_extracted_image_files(extracted_images)
        
        # Clean up session data
        request.session.pop('excel_data', None)
        request.session.pop('excel_columns', None)
        if 'extracted_images' in request.session:
            request.session.pop('extracted_images', None)
        
        progress_tracker.complete(success=True, final_message=f"Successfully imported {results['processed']} workers")
        
        return JsonResponse({
            'success': True,
            'operation_id': operation_id,
            'total_rows': results['processed'],
            'images_processed': 0,  # No image processing in simplified version
            'mrz_processed': 0,     # No MRZ processing in simplified version
            'message': f"Successfully imported {results['processed']} workers"
        })
        
    except Exception as e:
        return JsonResponse({'error': f'Import failed: {str(e)}'}, status=500)


def import_workers_from_excel_data(excel_data, request, progress_tracker=None, extracted_images=None):
    """Import workers directly from Excel data without MRZ processing."""
    from django.core.files.base import ContentFile
    
    results = {
        'total_rows': 0,
        'processed': 0,
        'skipped': 0,
        'duplicates': [],
        'errors': [],
        'created_workers': [],
    }
    
    try:
        from django.db import connection
        results['total_rows'] = len(excel_data)
        
        if progress_tracker:
            progress_tracker.update(
                stage=ProgressStages.PROCESSING_WORKERS,
                stage_message=f'Processing {len(excel_data)} workers from Excel...'
            )
        
        # Process each record
        for index, record in enumerate(excel_data):
            try:
                # Update progress
                if progress_tracker:
                    worker_name = record.get('NAME', f'Worker {index + 1}')
                    progress_tracker.update(
                        current_item=index + 1,
                        current_worker_name=worker_name,
                        processed=results['processed'],
                        skipped=results['skipped'],
                        failed=len(results['errors'])
                    )
                
                
                # Extract worker data from Excel record
                worker_data = _extract_worker_data_from_excel_record(record)
                
                # Check for duplicates
                duplicate_info = _is_duplicate_worker(worker_data)
                if duplicate_info['is_duplicate']:
                    results['skipped'] += 1
                    duplicate_detail = {
                        'row': index + 1,
                        'worker_name': f"{worker_data['first_name']} {worker_data['last_name']}",
                        'matching_workers': duplicate_info['matching_workers'],
                        'reasons': duplicate_info['match_reasons']
                    }
                    results['duplicates'].append(duplicate_detail)
                    continue
                
                # Create worker with transaction
                with transaction.atomic():
                    worker = _create_worker_from_data(worker_data, request.user)
                    
                    # Process photos and documents
                    if extracted_images:
                        _process_worker_photos_direct(worker, record, extracted_images, index, request.user)
                    else:
                        _process_worker_photos_from_json(worker, record, request)
                        # Create documents without images if no images available
                        _create_worker_documents(worker, worker_data, request.user)
                    
                    
                    results['processed'] += 1
                    results['created_workers'].append({
                        'name': worker.get_full_name(),
                        'worker_id': worker.worker_id,
                    })
                    
            except Exception as e:
                error_msg = f"Record {index + 1}: {str(e)}"
                results['errors'].append(error_msg)
        
        
        if progress_tracker:
            progress_tracker.update(
                stage=ProgressStages.COMPLETED,
                stage_message=f'Import completed: {results["processed"]} workers processed'
            )
    
    except Exception as e:
        results['errors'].append(f"Excel processing error: {str(e)}")
    
    return results



def _extract_worker_data_from_excel_record(record):
    """Extract worker data from Excel record."""
    from datetime import datetime
    
    # Split full name into first and last name
    full_name = str(record.get('NAME', '')).strip()
    name_parts = full_name.split(' ', 1)
    first_name = name_parts[0] if name_parts else ''
    last_name = name_parts[1] if len(name_parts) > 1 else ''
    
    # Parse BUILDING field which contains ZONE-BUILDING-FLOOR (e.g., "3H-B1-F11")
    zone_name = 'Default Zone'
    building_name = ''
    floor_name = ''
    
    building_field = record.get('BUILDING', '').strip()
    if building_field and '-' in building_field:
        parts = building_field.split('-')
        if len(parts) >= 3:
            zone_name = parts[0].strip()  # e.g., "3H"
            building_name = parts[1].strip()  # e.g., "B1"
            floor_name = parts[2].strip()  # e.g., "F11"
        elif len(parts) == 2:
            zone_name = parts[0].strip()
            building_name = parts[1].strip()
        elif len(parts) == 1:
            zone_name = parts[0].strip()
    elif building_field:
        # If no dashes, treat entire field as zone
        zone_name = building_field
    
    # Helper function to parse date ranges (issue_date-expiry_date)
    def parse_date_range(date_range_str):
        """Parse date range like '14/11/2018-14/11/2028' and return issue_date, expiry_date"""
        if not date_range_str or date_range_str == '' or pd.isna(date_range_str):
            return None, None
            
        date_range_str = str(date_range_str).strip()
        if '-' not in date_range_str:
            return None, None
            
        parts = date_range_str.split('-')
        if len(parts) != 2:
            return None, None
            
        issue_str = parts[0].strip()
        expiry_str = parts[1].strip()
        
        issue_date = None
        expiry_date = None
        
        # Parse issue date
        for date_format in ['%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y', '%Y-%m-%d']:
            try:
                issue_date = datetime.strptime(issue_str, date_format).date()
                break
            except ValueError:
                continue
                
        # Parse expiry date
        for date_format in ['%d/%m/%Y', '%Y/%m/%d', '%m/%d/%Y', '%Y-%m-%d']:
            try:
                expiry_date = datetime.strptime(expiry_str, date_format).date()
                break
            except ValueError:
                continue
                
        return issue_date, expiry_date
    
    # Parse dates
    dob = None
    if record.get('DOB'):
        try:
            dob_str = str(record['DOB']).strip()
            if dob_str and dob_str != '':
                # Try different date formats
                for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']:
                    try:
                        dob = datetime.strptime(dob_str, date_format).date()
                        break
                    except ValueError:
                        continue
        except Exception:
            pass
    
    joined_date = None
    if record.get('JOINEDDATE'):
        try:
            joined_str = str(record['JOINEDDATE']).strip()
            if joined_str and joined_str != '':
                for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']:
                    try:
                        joined_date = datetime.strptime(joined_str, date_format).date()
                        break
                    except ValueError:
                        continue
        except Exception:
            pass
    
    resigned_date = None
    if record.get('RESIGNEDDATE'):
        try:
            resigned_str = str(record['RESIGNEDDATE']).strip()
            if resigned_str and resigned_str != '':
                for date_format in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y/%m/%d']:
                    try:
                        resigned_date = datetime.strptime(resigned_str, date_format).date()
                        break
                    except ValueError:
                        continue
        except Exception:
            pass
    
    # Parse document dates (IDKH/PASSPORT_DATE, VISA_DATE, WORKINGPERMIT_DATE)
    idkh_issue_date, idkh_expiry_date = parse_date_range(record.get('IDKH/PASSPORT_DATE'))
    visa_issue_date, visa_expiry_date = parse_date_range(record.get('VISA_DATE'))
    permit_issue_date, permit_expiry_date = parse_date_range(record.get('WORKINGPERMIT_DATE'))
    
    return {
        'first_name': first_name,
        'last_name': last_name,
        'sex': record.get('SEX', '').strip().upper() if record.get('SEX') else '',
        'dob': dob,
        'nationality': record.get('NATIONALITY', '').strip() if record.get('NATIONALITY') else '',
        'phone_number': str(record.get('PHONE_NUMBER', '')).strip() if record.get('PHONE_NUMBER') else '',
        'zone_name': zone_name,
        'building_name': building_name,
        'floor_name': floor_name,
        'position_name': record.get('POSITION', '').strip() if record.get('POSITION') else 'STAFF',  # Default to STAFF
        'joined_date': joined_date,
        'resigned_date': resigned_date,
        'is_active': record.get('ACTIVE/NOACTIVE', '').strip().upper() == 'ACTIVE' if record.get('ACTIVE/NOACTIVE') else True,
        'status': 'active',  # Default status to active
        'document_number': record.get('DOCUMENT NO', '').strip() if record.get('DOCUMENT NO') else '',
        # Document dates
        'idkh_issue_date': idkh_issue_date,
        'idkh_expiry_date': idkh_expiry_date,
        'visa_issue_date': visa_issue_date,
        'visa_expiry_date': visa_expiry_date,
        'permit_issue_date': permit_issue_date,
        'permit_expiry_date': permit_expiry_date,
    }


def _create_worker_documents(worker, worker_data, created_by):
    """Create document records for worker based on extracted dates."""
    
    # Create ID/Passport document if dates are available
    if worker_data.get('idkh_issue_date') and worker_data.get('idkh_expiry_date'):
        try:
            Document.objects.create(
                worker=worker,
                document_type='passport',  # or 'id_card' based on your logic
                document_number=worker_data.get('document_number', ''),
                issue_date=worker_data['idkh_issue_date'],
                expiry_date=worker_data['idkh_expiry_date'],
                issuing_authority='Immigration',
                created_by=created_by
            )
        except Exception as e:
            pass
    
    # Create Visa document if dates are available
    if worker_data.get('visa_issue_date') and worker_data.get('visa_expiry_date'):
        try:
            Document.objects.create(
                worker=worker,
                document_type='visa',
                document_number='',  # No separate visa number in Excel
                issue_date=worker_data['visa_issue_date'],
                expiry_date=worker_data['visa_expiry_date'],
                issuing_authority='Immigration',
                created_by=created_by
            )
        except Exception as e:
            pass
    
    # Create Work Permit document if dates are available
    if worker_data.get('permit_issue_date') and worker_data.get('permit_expiry_date'):
        try:
            Document.objects.create(
                worker=worker,
                document_type='work_permit',
                document_number='',  # No separate permit number in Excel
                issue_date=worker_data['permit_issue_date'],
                expiry_date=worker_data['permit_expiry_date'],
                issuing_authority='Labor Department',
                created_by=created_by
            )
        except Exception as e:
            pass


def preview_excel_json(request):
    """Extract Excel to JSON for preview without importing."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'error': 'AJAX request required'}, status=400)
    
    try:
        form = WorkerExcelImportForm(request.POST, request.FILES)
        if not form.is_valid():
            return JsonResponse({'error': 'Invalid form data', 'details': form.errors}, status=400)
        
        # Create progress tracker for extraction only
        progress_tracker = ImportProgressTracker()
        operation_id = progress_tracker.start(
            total_items=1,
            operation_type="excel_preview"
        )
        
        # Save Excel file temporarily
        excel_file = form.cleaned_data['excel_file']
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            for chunk in excel_file.chunks():
                tmp_file.write(chunk)
            tmp_file.flush()
            tmp_excel_path = tmp_file.name
        
        try:
            # Step 1: Extract to JSON and Images (NO MRZ processing yet)
            json_path, images_dir, record_count, image_count = extract_excel_to_json_and_images(tmp_excel_path, progress_tracker)
            
            # Load the JSON data to return for preview
            with open(json_path, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            # Analyze the data for preview
            analysis = analyze_json_data(json_data)
            
            progress_tracker.complete(success=True, final_message=f"Step 1 complete: {record_count} workers, {image_count} images extracted")
            
            return JsonResponse({
                'success': True,
                'operation_id': operation_id,
                'json_path': json_path,  # Store for later import
                'images_dir': images_dir,  # Store images directory
                'record_count': record_count,
                'image_count': image_count,
                'json_data': json_data,
                'analysis': analysis,
                'step_completed': 1  # Indicate Step 1 is done
            })
            
        finally:
            # Clean up temp Excel file
            if os.path.exists(tmp_excel_path):
                os.unlink(tmp_excel_path)
                
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def test_separated_workflow(request):
    """Test the separated workflow with existing Excel file."""
    excel_path = os.path.join(settings.BASE_DIR, 'OCR', 'worker_eform1.xlsx')
    
    if not os.path.exists(excel_path):
        return JsonResponse({'error': 'Test Excel file not found'}, status=404)
    
    try:
        # Create progress tracker
        progress_tracker = ImportProgressTracker()
        operation_id = progress_tracker.start(
            total_items=3,
            operation_type="test_separated_workflow"
        )
        
        # Step 1: Extract Excel to JSON and Images
        json_path, images_dir, record_count, image_count = extract_excel_to_json_and_images(excel_path, progress_tracker)
        
        # Step 2: Process MRZ from JSON and images
        mrz_json_path, mrz_record_count = process_mrz_from_json(json_path, progress_tracker)
        
        progress_tracker.complete(success=True, final_message=f"Test complete: {mrz_record_count} records processed")
        
        return JsonResponse({
            'success': True,
            'message': 'Separated workflow test completed successfully',
            'step1': {
                'json_path': json_path,
                'images_dir': images_dir,
                'record_count': record_count,
                'image_count': image_count
            },
            'step2': {
                'mrz_json_path': mrz_json_path,
                'mrz_record_count': mrz_record_count
            },
            'operation_id': operation_id
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def workflow_demo(request):
    """Display the workflow demo page."""
    context = {
        'page_title': 'Worker Import Workflow Demo',
    }
    return render(request, 'zone/worker_import/workflow_demo.html', context)

@login_required
def workflow_demo_api(request):
    """API endpoint for workflow demo with real-time updates."""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST method required'}, status=405)
    
    step = request.POST.get('step')
    
    try:
        if step == 'step1':
            # Step 1: Excel Extraction
            excel_path = os.path.join(settings.BASE_DIR, 'OCR', 'worker_eform1.xlsx')
            
            if not os.path.exists(excel_path):
                return JsonResponse({'error': 'Sample Excel file not found'}, status=404)
            
            # Create progress tracker
            progress_tracker = ImportProgressTracker()
            operation_id = progress_tracker.start(
                total_items=3,
                operation_type="workflow_demo_step1"
            )
            
            # Run Step 1
            json_path, images_dir, record_count, image_count = extract_excel_to_json_and_images(excel_path, progress_tracker)
            
            progress_tracker.complete(success=True, final_message=f"Step 1 complete: {record_count} records, {image_count} images")
            
            # Store paths in session for next step
            request.session['demo_json_path'] = json_path
            request.session['demo_images_dir'] = images_dir
            
            return JsonResponse({
                'success': True,
                'step': 1,
                'data': {
                    'json_path': json_path,
                    'images_dir': images_dir,
                    'record_count': record_count,
                    'image_count': image_count,
                    'operation_id': operation_id
                },
                'message': f'Step 1 completed: Extracted {record_count} records with {image_count} images'
            })
            
        elif step == 'step2':
            # Step 2: MRZ Processing
            json_path = request.session.get('demo_json_path')
            
            if not json_path or not os.path.exists(json_path):
                return JsonResponse({'error': 'Step 1 must be completed first'}, status=400)
            
            # Create progress tracker for step 2
            progress_tracker = ImportProgressTracker()
            operation_id = progress_tracker.start(
                total_items=3,
                operation_type="workflow_demo_step2"
            )
            
            # Run Step 2
            mrz_json_path, mrz_record_count = process_mrz_from_json(json_path, progress_tracker)
            
            progress_tracker.complete(success=True, final_message=f"Step 2 complete: {mrz_record_count} records with MRZ")
            
            # Store MRZ path for next step
            request.session['demo_mrz_json_path'] = mrz_json_path
            
            # Get some sample MRZ data for display
            with open(mrz_json_path, 'r', encoding='utf-8') as f:
                mrz_data = json.load(f)
            
            # Extract sample MRZ info
            sample_mrz = []
            for record in mrz_data:
                if record.get('IDKH/PASSPORT') and isinstance(record['IDKH/PASSPORT'], dict):
                    if record['IDKH/PASSPORT'].get('document'):
                        sample_mrz.append({
                            'name': record.get('NAME'),
                            'document': record['IDKH/PASSPORT']['document']
                        })
            
            return JsonResponse({
                'success': True,
                'step': 2,
                'data': {
                    'mrz_json_path': mrz_json_path,
                    'mrz_record_count': mrz_record_count,
                    'sample_mrz': sample_mrz,
                    'operation_id': operation_id
                },
                'message': f'Step 2 completed: Processed MRZ for {mrz_record_count} records'
            })
            
        elif step == 'step3':
            # Step 3: Database Import (simulate only)
            mrz_json_path = request.session.get('demo_mrz_json_path')
            
            if not mrz_json_path or not os.path.exists(mrz_json_path):
                return JsonResponse({'error': 'Step 2 must be completed first'}, status=400)
            
            # For demo purposes, just simulate import without actually creating workers
            with open(mrz_json_path, 'r', encoding='utf-8') as f:
                import_data = json.load(f)
            
            return JsonResponse({
                'success': True,
                'step': 3,
                'data': {
                    'workers_ready': len(import_data),
                    'would_create': len(import_data),
                    'demo_mode': True
                },
                'message': f'Step 3 ready: Would import {len(import_data)} workers (demo mode - not actually importing)'
            })
            
        else:
            return JsonResponse({'error': 'Invalid step'}, status=400)
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def analyze_json_data(json_data):
    """Analyze JSON data to provide preview statistics and warnings."""
    analysis = {
        'total_workers': len(json_data),
        'total_images': 0,
        'workers_with_photos': 0,
        'workers_with_documents': 0,
        'image_types': {'PHOTO': 0, 'IDKH/PASSPORT': 0, 'WORKINGPERMIT': 0, 'VISA': 0},
        'warnings': []
    }
    
    for record in json_data:
        worker_images = []
        
        # Count images per type
        for img_type in ['PHOTO', 'IDKH/PASSPORT', 'WORKINGPERMIT', 'VISA']:
            if record.get(img_type) and isinstance(record[img_type], dict):
                analysis['image_types'][img_type] += 1
                analysis['total_images'] += 1
                worker_images.append(img_type)
        
        # Count workers with photos vs documents
        if 'PHOTO' in worker_images:
            analysis['workers_with_photos'] += 1
        
        if any(doc_type in worker_images for doc_type in ['IDKH/PASSPORT', 'WORKINGPERMIT', 'VISA']):
            analysis['workers_with_documents'] += 1
        
        # Check for potential issues
        if not record.get('NAME'):
            analysis['warnings'].append(f"Worker in row {record.get('NO', '?')} has no name")
        
        if not record.get('ZONE_NAME'):
            analysis['warnings'].append(f"Worker {record.get('NAME', '?')} has no zone")
        
        if not record.get('BUILDINGNAME'):
            analysis['warnings'].append(f"Worker {record.get('NAME', '?')} has no building")
        
        if not worker_images:
            analysis['warnings'].append(f"Worker {record.get('NAME', '?')} has no images")
    
    return analysis


# import worker vip from excel file
@login_required
def worker_import_excel(request):

    if request.method == 'POST':
        
        # Check if this is an import confirmation from JSON preview
        if request.POST.get('import_from_json'):
            json_file_path = request.POST.get('json_file_path')
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            if not json_file_path or not os.path.exists(json_file_path):
                error_msg = "JSON file not found. Please re-extract the Excel file."
                if is_ajax:
                    return JsonResponse({'success': False, 'error': error_msg}, status=400)
                else:
                    messages.error(request, error_msg)
                    return redirect('zone:worker_import_excel')
            
            try:
                # Create progress tracker for the import
                progress_tracker = ImportProgressTracker()
                
                # Analyze JSON to get the total count for progress tracking
                with open(json_file_path, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                
                operation_id = progress_tracker.start(
                    total_items=len(json_data),
                    operation_type="worker_import_from_json"
                )
                
                # Step 2: Process MRZ from JSON and images
                mrz_json_path, mrz_record_count = process_mrz_from_json(json_file_path, progress_tracker)
                
                # Step 3: Import workers from MRZ-processed JSON
                results = process_json_import(mrz_json_path, request, progress_tracker)
                
                progress_tracker.complete(success=True, final_message=f"Successfully imported {results['processed']} workers with MRZ processing")
                
                if is_ajax:
                    return JsonResponse({
                        'success': True,
                        'operation_id': operation_id,
                        'total_rows': results['processed'],
                        'images_processed': results.get('images_processed', 0),
                        'message': f"Successfully imported {results['processed']} workers"
                    })
                else:
                    messages.success(request, f"Successfully imported {results['processed']} workers from JSON")
                    return redirect('zone:worker_list')
                    
            except Exception as e:
                if is_ajax:
                    return JsonResponse({'success': False, 'error': str(e)}, status=500)
                else:
                    messages.error(request, f"Import failed: {str(e)}")
                    return redirect('zone:worker_import_excel')
        
        form = WorkerExcelImportForm(request.POST, request.FILES)
      
        if not form.is_valid():
            pass
        
        if form.is_valid():
            # Check if this is an AJAX request for actual import
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
            
            # ALWAYS validate first, even if not a preview request
            validation_results = check_excel_file(form)
   
            
            # Check if this is a preview request OR if there are validation issues
            if 'preview' in request.POST or (validation_results['invalid_rows'] > 0 or validation_results['duplicates']) and not is_ajax:
                # Get template download URL
                template_url = request.build_absolute_uri('/') + 'worker_eform.xlsx'
                
                context = {
                    'form': form,
                    'template_url': template_url,
                    'validation_results': validation_results,
                    'is_preview': True,
                    'page_title': 'Import Workers - Validation Results',
                }
                return render(request, 'zone/worker_import/worker_import_excel.html', context)
            else:
                # Only process if validation passed and not a preview
                
                if is_ajax:
                    # For AJAX request, use the new two-step process with preview
                    try:
                        # Create progress tracker for extraction
                        progress_tracker = ImportProgressTracker()
                        operation_id = progress_tracker.start(
                            total_items=validation_results.get('total_rows', 0),
                            operation_type="excel_extraction"
                        )
                        
                        # Step 1: Extract Excel to JSON with proper image mapping (for preview)
                        excel_file = form.cleaned_data['excel_file']
                        
                        # Save Excel file temporarily
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                            for chunk in excel_file.chunks():
                                tmp_file.write(chunk)
                            tmp_file.flush()
                            tmp_excel_path = tmp_file.name
                        
                        try:
                            # Extract to JSON for preview
                            json_path, record_count, image_count = extract_excel_to_json(tmp_excel_path, progress_tracker)
                            
                            progress_tracker.complete(success=True, final_message=f"Extracted {record_count} records with {image_count} images")
                            
                            # Return JSON path for preview modal
                            return JsonResponse({
                                'success': True,
                                'operation_id': operation_id,
                                'json_file_path': json_path,
                                'total_rows': record_count,
                                'images_extracted': image_count,
                                'show_preview': True
                            })
                            
                        finally:
                            # Clean up temp Excel file
                            if os.path.exists(tmp_excel_path):
                                os.unlink(tmp_excel_path)
                        
                    except Exception as e:
                        progress_tracker.complete(success=False, final_message=f"Extraction failed: {str(e)}")
                        return JsonResponse({'success': False, 'error': str(e)}, status=500)
                else:
                    # Normal form submission
                    return set_process_excel_import(request, form)
    else:
        form = WorkerExcelImportForm()
    
    # Get template download URL
    template_url = request.build_absolute_uri('/') + 'worker_eform.xlsx'
    
    context = {
        'form': form,
        'template_url': template_url,
        'page_title': 'Import Workers from Excel',
    }

    return render(request, "zone/worker_import/worker_import_excel.html", context)



def check_excel_file(form):
    """Preview and validate Excel file before import."""
    import tempfile
    from openpyxl import load_workbook
    
    excel_file = form.cleaned_data['excel_file']
    
    validation_results = {
        'total_rows': 0,
        'valid_rows': 0,
        'invalid_rows': 0,
        'rows_data': [],
        'errors': [],
        'warnings': [],
        'duplicates': [],
        'missing_fields': [],
    }
    
    tmp_file_name = None
    try:
        # Save Excel file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file_name = tmp_file.name
            for chunk in excel_file.chunks():
                tmp_file.write(chunk)
            tmp_file.flush()
        
        # Read Excel file
        df = pd.read_excel(tmp_file_name, sheet_name='Workers')
        
        # Filter out completely empty rows
        df = df.dropna(how='all')
        validation_results['total_rows'] = len(df)
        
        # Validate each row
        for index, row in df.iterrows():
            row_validation = check_worker_row(row, index + 2)  # +2 for Excel row numbering
            validation_results['rows_data'].append(row_validation)
            
            if row_validation['is_valid']:
                validation_results['valid_rows'] += 1
            else:
                validation_results['invalid_rows'] += 1
                
            # Collect errors and warnings
            validation_results['errors'].extend(row_validation['errors'])
            validation_results['warnings'].extend(row_validation['warnings'])
            
            # Check for duplicates
            if row_validation['is_duplicate']:
                validation_results['duplicates'].append(row_validation)
    
    except Exception as e:
        validation_results['errors'].append(f"File processing error: {str(e)}")
    
    finally:
        # Clean up temp file
        if tmp_file_name and os.path.exists(tmp_file_name):
            try:
                os.unlink(tmp_file_name)
            except OSError as e:
                pass
    
    # Calculate how many workers will actually be imported (valid - duplicates)
    validation_results['will_import'] = validation_results['valid_rows'] - len(validation_results['duplicates'])
    
    return validation_results

def check_worker_row(row, row_number):
    """Validate a single worker row and return validation results."""
    
    validation = {
        'row_number': row_number,
        'is_valid': True,
        'is_duplicate': False,
        'errors': [],
        'warnings': [],
        'data': {},
    }
    
    # Required fields validation
    required_fields = ['NAME', 'SEX', 'NATIONALITY']
    
    for field in required_fields:
        value = row.get(field)
        if pd.isna(value) or str(value).strip() == '':
            validation['errors'].append(f"Row {row_number}: Missing required field '{field}'")
            validation['is_valid'] = False
    
    # Extract and validate worker data
    try:
        worker_data = set_extract_worker_data(row)
        validation['data'] = worker_data
        
        # Validate specific fields
        
        # Name validation
        if worker_data.get('first_name'):
            if len(worker_data['first_name']) < 2:
                validation['warnings'].append(f"Row {row_number}: First name seems too short")
        
        # Sex validation
        sex = str(row.get('SEX', '')).strip().upper()
        if sex and sex not in ['M', 'MALE', 'F', 'FEMALE']:
            validation['errors'].append(f"Row {row_number}: Invalid sex value '{sex}'. Use M/Male or F/Female")
            validation['is_valid'] = False
            
        # Phone number validation
        phone = str(row.get('PHONE_NUMBER', '')).strip()
        if phone and phone != 'nan':
            # Basic phone validation - should be digits with possible + or spaces
            import re
            # Clean phone for validation
            clean_phone = phone.replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('+', '')
            
            if not clean_phone.isdigit():
                validation['errors'].append(f"Row {row_number}: Phone number '{phone}' contains invalid characters")
                validation['is_valid'] = False
            elif len(clean_phone) < 6:
                validation['errors'].append(f"Row {row_number}: Phone number '{phone}' is too short (minimum 6 digits)")
                validation['is_valid'] = False
            elif len(clean_phone) > 15:
                validation['warnings'].append(f"Row {row_number}: Phone number '{phone}' seems too long (more than 15 digits)")
        
        # Nationality validation
        nationality = str(row.get('NATIONALITY', '')).strip()
        if nationality and len(nationality) < 2:
            validation['warnings'].append(f"Row {row_number}: Nationality seems too short")
            
        # Date validation
        dob = row.get('DOB')
        if not pd.isna(dob) and isinstance(dob, str):
            try:
                pd.to_datetime(dob)
            except:
                validation['warnings'].append(f"Row {row_number}: Date of birth format may be invalid")
        
        joined_date = row.get('JOINEDDATE')
        if not pd.isna(joined_date) and isinstance(joined_date, str):
            try:
                pd.to_datetime(joined_date)
            except:
                validation['warnings'].append(f"Row {row_number}: Joined date format may be invalid")
        
        # Check for duplicate workers
        duplicate_info = get_is_duplicate_worker(worker_data)
        if duplicate_info['is_duplicate']:
            validation['is_duplicate'] = True
            validation['duplicate_info'] = duplicate_info
            
            # Determine severity based on match type
            match_type = duplicate_info.get('match_type', 'possible')
            
            if match_type == 'exact':
                # Phone number match - definite duplicate
                for match_reason in duplicate_info['match_reasons']:
                    validation['errors'].append(f"Row {row_number}: ❌ DUPLICATE - {match_reason}")
                    validation['is_valid'] = False
            elif match_type == 'strong':
                # Name + DOB match - very likely duplicate
                for match_reason in duplicate_info['match_reasons']:
                    validation['errors'].append(f"Row {row_number}: ⚠️ LIKELY DUPLICATE - {match_reason}")
                    validation['is_valid'] = False
            else:
                # Possible match - warning only
                for match_reason in duplicate_info['match_reasons']:
                    validation['warnings'].append(f"Row {row_number}: ⚠️ POSSIBLE DUPLICATE - {match_reason}")
            
            # Add detailed info about existing workers
            for existing_worker in duplicate_info['matching_workers']:
                details = [
                    f"Name: {existing_worker['name']}",
                    f"ID: {existing_worker['worker_id']}",
                    f"Zone: {existing_worker['zone']}",
                    f"Status: {existing_worker['status']}"
                ]
                if existing_worker.get('phone') and existing_worker['phone'] != 'N/A':
                    details.append(f"Phone: {existing_worker['phone']}")
                if existing_worker.get('dob') and existing_worker['dob'] != 'N/A':
                    details.append(f"DOB: {existing_worker['dob']}")
                
                validation['warnings'].append(
                    f"Row {row_number}: Existing worker - {' | '.join(details)}"
                )
        
        # Position and Department validation
        position = str(row.get('POSITION', '')).strip()
        if position and position != 'nan':
            # Check if position exists
            if not Position.objects.filter(name__iexact=position).exists():
                validation['warnings'].append(f"Row {row_number}: Position '{position}' doesn't exist and will be created")
        
        department = str(row.get('DEPARTMENT', '')).strip()
        if department and department != 'nan':
            # Check if department exists
            if not Department.objects.filter(name__iexact=department).exists():
                validation['warnings'].append(f"Row {row_number}: Department '{department}' doesn't exist and will be created")
        
        # Zone validation
        zone_name = str(row.get('ZONE_NAME', '')).strip()
        if zone_name and zone_name != 'nan':
            if not Zone.objects.filter(name__iexact=zone_name).exists():
                validation['errors'].append(f"Row {row_number}: Zone '{zone_name}' doesn't exist")
                validation['is_valid'] = False
        
        # Building and Floor validation (handles format like "B14-F1")
        building_floor = str(row.get('BUILDINGNAME', '')).strip()
        if building_floor and building_floor != 'nan':
            # Parse building and floor from format like "B14-F1"
            if '-' in building_floor:
                parts = building_floor.split('-', 1)
                building_name = parts[0].strip()
                floor_name = parts[1].strip() if len(parts) > 1 else ''
            else:
                building_name = building_floor
                floor_name = ''
            
            # Validate building exists
            if building_name:
                if not Building.objects.filter(name__iexact=building_name).exists():
                    validation['errors'].append(f"Row {row_number}: Building '{building_name}' doesn't exist")
                    validation['is_valid'] = False
                elif floor_name:
                    # If building exists, validate floor exists for that building
                    building = Building.objects.filter(name__iexact=building_name).first()
                    if building and not Floor.objects.filter(building=building, name__iexact=floor_name).exists():
                        validation['warnings'].append(f"Row {row_number}: Floor '{floor_name}' doesn't exist in building '{building_name}'")
                        # Not setting is_valid to False for floor as it might be optional
        
    except Exception as e:
        validation['errors'].append(f"Row {row_number}: Data extraction error - {str(e)}")
        validation['is_valid'] = False
    
    return validation


def set_extract_worker_data(row):
    """Extract worker data from Excel row."""
    
    # Split full name into first and last name
    full_name = str(row.get('NAME', '')).strip()
    name_parts = full_name.split(' ', 1)
    first_name = name_parts[0] if name_parts else ''
    last_name = name_parts[1] if len(name_parts) > 1 else ''
    
    # Convert DOB to proper date format
    dob = row.get('DOB')
    if pd.isna(dob):
        dob = None
    elif isinstance(dob, str):
        try:
            dob = pd.to_datetime(dob).date()
        except:
            dob = None
    
    # Convert joined date
    joined_date = row.get('JOINEDDATE')
    if pd.isna(joined_date):
        joined_date = timezone.now().date()
    elif isinstance(joined_date, str):
        try:
            joined_date = pd.to_datetime(joined_date).date()
        except:
            joined_date = timezone.now().date()
    
    # Map status
    status_value = str(row.get('ACTIVE/NOACTIVE', 'ACTIVE')).upper()
    status = 'active' if status_value == 'ACTIVE' else 'inactive'
    
    # Parse building and floor from format like "B14-F1"
    building_floor = str(row.get('BUILDINGNAME', '')).strip()
    building_name = ''
    floor_name = ''
    
    if building_floor and building_floor != 'nan':
        if '-' in building_floor:
            parts = building_floor.split('-', 1)
            building_name = parts[0].strip()
            floor_name = parts[1].strip() if len(parts) > 1 else ''
        else:
            building_name = building_floor
    
    return {
        'first_name': first_name,
        'last_name': last_name,
        'sex': str(row.get('SEX', '')).upper(),
        'dob': dob,
        'nationality': str(row.get('NATIONALITY', '')),
        'phone_number': str(row.get('PHONE_NUMBER', '')),
        'zone_name': str(row.get('ZONE_NAME', '')),
        'building_name': building_name,
        'floor_name': floor_name,
        'position_name': str(row.get('POSITION', '')),
        'joined_date': joined_date,
        'status': status,
        'photo_filename': str(row.get('PHOTO', '')),
        'id_document_filename': str(row.get('IDKH/PASSPORT', '')),
        'work_permit_filename': str(row.get('WORKINGPERMIT', '')),
        'visa_filename': str(row.get('VISA', '')),
    }


def get_is_duplicate_worker(worker_data):
    """Check if worker already exists and return detailed information."""
    duplicate_info = {
        'is_duplicate': False,
        'matching_workers': [],
        'match_reasons': [],
        'match_type': None  # 'exact', 'strong', 'possible'
    }
    
    matched_worker_ids = set()  # Track already matched workers to avoid duplicates
    
    # 1. Check by phone number (high confidence match)
    if worker_data['phone_number'] and worker_data['phone_number'] != 'nan' and worker_data['phone_number'].strip():
        # Clean phone number for comparison
        clean_phone = worker_data['phone_number'].strip().replace(' ', '').replace('-', '').replace('+', '')
        if clean_phone:
            phone_matches = Worker.objects.filter(phone_number__icontains=clean_phone)
            if phone_matches.exists():
                duplicate_info['is_duplicate'] = True
                duplicate_info['match_type'] = 'exact'
                for worker in phone_matches:
                    if worker.id not in matched_worker_ids:
                        matched_worker_ids.add(worker.id)
                        duplicate_info['matching_workers'].append({
                            'id': worker.id,
                            'name': worker.get_full_name(),
                            'worker_id': worker.worker_id,
                            'zone': worker.zone.name if worker.zone else 'N/A',
                            'status': worker.status,
                            'phone': worker.phone_number,
                            'dob': worker.dob.strftime('%Y-%m-%d') if worker.dob else 'N/A'
                        })
                duplicate_info['match_reasons'].append(f"Phone number '{worker_data['phone_number']}' already exists")
    
    # 2. Check by DOB + Name combination (high confidence match)
    if worker_data['dob'] and worker_data['first_name'] and worker_data['last_name']:
        dob_name_matches = Worker.objects.filter(
            dob=worker_data['dob'],
            first_name__iexact=worker_data['first_name'],
            last_name__iexact=worker_data['last_name']
        )
        if dob_name_matches.exists():
            duplicate_info['is_duplicate'] = True
            if not duplicate_info['match_type']:
                duplicate_info['match_type'] = 'strong'
            for worker in dob_name_matches:
                if worker.id not in matched_worker_ids:
                    matched_worker_ids.add(worker.id)
                    duplicate_info['matching_workers'].append({
                        'id': worker.id,
                        'name': worker.get_full_name(),
                        'worker_id': worker.worker_id,
                        'zone': worker.zone.name if worker.zone else 'N/A',
                        'status': worker.status,
                        'phone': worker.phone_number if worker.phone_number else 'N/A',
                        'dob': worker.dob.strftime('%Y-%m-%d') if worker.dob else 'N/A'
                    })
            dob_str = worker_data['dob'].strftime('%Y-%m-%d') if hasattr(worker_data['dob'], 'strftime') else str(worker_data['dob'])
            duplicate_info['match_reasons'].append(
                f"Same name '{worker_data['first_name']} {worker_data['last_name']}' with DOB {dob_str} already exists"
            )
    
    # 3. Check by exact name match only (medium confidence)
    elif worker_data['first_name'] and worker_data['last_name']:
        name_matches = Worker.objects.filter(
            first_name__iexact=worker_data['first_name'],
            last_name__iexact=worker_data['last_name']
        )
        if name_matches.exists():
            # Check if any have matching DOB (if DOB provided)
            strong_match = False
            if worker_data['dob']:
                for worker in name_matches:
                    if worker.dob == worker_data['dob']:
                        strong_match = True
                        break
            
            if strong_match or len(name_matches) == 1:  # Single name match or DOB match
                duplicate_info['is_duplicate'] = True
                if not duplicate_info['match_type']:
                    duplicate_info['match_type'] = 'strong' if strong_match else 'possible'
                for worker in name_matches:
                    if worker.id not in matched_worker_ids:
                        matched_worker_ids.add(worker.id)
                        duplicate_info['matching_workers'].append({
                            'id': worker.id,
                            'name': worker.get_full_name(),
                            'worker_id': worker.worker_id,
                            'zone': worker.zone.name if worker.zone else 'N/A',
                            'status': worker.status,
                            'phone': worker.phone_number if worker.phone_number else 'N/A',
                            'dob': worker.dob.strftime('%Y-%m-%d') if worker.dob else 'N/A'
                        })
                duplicate_info['match_reasons'].append(
                    f"Name '{worker_data['first_name']} {worker_data['last_name']}' already exists"
                )
    
    # 4. Check by DOB + partial name (possible match)
    if worker_data['dob'] and (worker_data['first_name'] or worker_data['last_name']):
        dob_matches = Worker.objects.filter(dob=worker_data['dob'])
        
        for worker in dob_matches:
            # Check for partial name match
            name_match = False
            if worker_data['first_name'] and worker.first_name:
                if worker_data['first_name'].lower() in worker.first_name.lower() or \
                   worker.first_name.lower() in worker_data['first_name'].lower():
                    name_match = True
            if worker_data['last_name'] and worker.last_name:
                if worker_data['last_name'].lower() in worker.last_name.lower() or \
                   worker.last_name.lower() in worker_data['last_name'].lower():
                    name_match = True
            
            if name_match and worker.id not in matched_worker_ids:
                duplicate_info['is_duplicate'] = True
                if not duplicate_info['match_type']:
                    duplicate_info['match_type'] = 'possible'
                matched_worker_ids.add(worker.id)
                duplicate_info['matching_workers'].append({
                    'id': worker.id,
                    'name': worker.get_full_name(),
                    'worker_id': worker.worker_id,
                    'zone': worker.zone.name if worker.zone else 'N/A',
                    'status': worker.status,
                    'phone': worker.phone_number if worker.phone_number else 'N/A',
                    'dob': worker.dob.strftime('%Y-%m-%d') if worker.dob else 'N/A'
                })
                if f"Similar worker with same DOB" not in str(duplicate_info['match_reasons']):
                    dob_str = worker_data['dob'].strftime('%Y-%m-%d') if hasattr(worker_data['dob'], 'strftime') else str(worker_data['dob'])
                    duplicate_info['match_reasons'].append(f"Similar worker with same DOB {dob_str} found")
    
    return duplicate_info

def process_json_import(json_path, request, progress_tracker=None):
    """Import workers from JSON data with proper image mapping and OCR processing."""
    import json
    from django.core.files.base import ContentFile
    from sinosecu.models import PassportScan, ScanImage, ScanResult
    
    results = {
        'total_rows': 0,
        'processed': 0,
        'skipped': 0,
        'duplicates': [],
        'errors': [],
        'created_workers': [],
    }
    
    try:
        # Load JSON data
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        # Handle new JSON structure with metadata
        if isinstance(json_data, dict) and 'records' in json_data:
            # New structure with metadata
            records = json_data['records']
            metadata = json_data.get('metadata', {})
        else:
            # Backward compatibility with old structure
            records = json_data
            metadata = {}
        
        results['total_rows'] = len(records)
        
        if progress_tracker:
            progress_tracker.update(
                stage=ProgressStages.PROCESSING_WORKERS,
                stage_message=f'Processing {len(records)} workers from JSON...'
            )
        
        # Process each record
        for index, record in enumerate(records):
            try:
                # Update progress
                if progress_tracker:
                    worker_name = record.get('NAME', f'Worker {index + 1}')
                    progress_tracker.update(
                        current_item=index + 1,
                        current_worker_name=worker_name,
                        processed=results['processed'],
                        skipped=results['skipped'],
                        failed=len(results['errors'])
                    )
                
                # Extract worker data from JSON record
                worker_data = _extract_worker_data_from_json(record)
                
                # Check for duplicates
                duplicate_info = _is_duplicate_worker(worker_data)
                if duplicate_info['is_duplicate']:
                    results['skipped'] += 1
                    duplicate_detail = {
                        'row': index + 1,
                        'worker_name': f"{worker_data['first_name']} {worker_data['last_name']}",
                        'matching_workers': duplicate_info['matching_workers'],
                        'reasons': duplicate_info['match_reasons']
                    }
                    results['duplicates'].append(duplicate_detail)
                    continue
                
                # Create worker with transaction
                with transaction.atomic():
                    worker = _create_worker_from_data(worker_data, request.user)
                    
                    # Process photos and documents from JSON
                    for col in ['PHOTO', 'IDKH/PASSPORT', 'WORKINGPERMIT', 'VISA']:
                        pass
                    
                    # Process photos and documents
                    if extracted_images:
                        _process_worker_photos_direct(worker, record, extracted_images, index, request.user)
                    else:
                        _process_worker_photos_from_json(worker, record, request)
                    
                    
                    results['processed'] += 1
                    results['created_workers'].append({
                        'name': worker.get_full_name(),
                        'worker_id': worker.worker_id,
                    })
                    
            except Exception as e:
                error_msg = f"Record {index + 1}: {str(e)}"
                results['errors'].append(error_msg)
        
        
        if progress_tracker:
            progress_tracker.update(
                stage=ProgressStages.COMPLETED,
                stage_message=f'Import completed: {results["processed"]} workers processed'
            )
    
    except Exception as e:
        results['errors'].append(f"JSON processing error: {str(e)}")
    
    return results

def _extract_worker_data_from_json(record):
    """Extract worker data from JSON record."""
    from datetime import datetime
    
    # Split full name into first and last name
    full_name = str(record.get('NAME', '')).strip()
    name_parts = full_name.split(' ', 1)
    first_name = name_parts[0] if name_parts else ''
    last_name = name_parts[1] if len(name_parts) > 1 else ''
    
    # Parse dates
    dob = None
    if record.get('DOB'):
        try:
            if isinstance(record['DOB'], str):
                dob = datetime.strptime(record['DOB'], '%Y-%m-%d').date()
            else:
                dob = record['DOB']
        except:
            dob = None
    
    joined_date = None
    if record.get('JOINEDDATE'):
        try:
            if isinstance(record['JOINEDDATE'], str):
                joined_date = datetime.strptime(record['JOINEDDATE'], '%Y-%m-%d').date()
            else:
                joined_date = record['JOINEDDATE']
        except:
            joined_date = timezone.now().date()
    else:
        joined_date = timezone.now().date()
    
    # Handle both old and new column structures for location data
    # New structure: ZONE, BUILDING, FLOOR
    # Old structure: ZONE_NAME, BUILDINGNAME
    zone_name = ''
    building_name = ''
    floor_name = ''
    
    if 'ZONE' in record and 'BUILDING' in record and 'FLOOR' in record:
        # New structure
        zone_name = str(record.get('ZONE', '')).strip()
        building_name = str(record.get('BUILDING', '')).strip()
        floor_name = str(record.get('FLOOR', '')).strip()
    else:
        # Old structure fallback
        zone_name = str(record.get('ZONE_NAME', '')).strip()
        building_floor = str(record.get('BUILDINGNAME', '')).strip()
        
        if building_floor and building_floor != 'None':
            if '-' in building_floor:
                parts = building_floor.split('-', 1)
                building_name = parts[0].strip()
                floor_name = parts[1].strip() if len(parts) > 1 else ''
            else:
                building_name = building_floor
    
    # Map status
    status_value = str(record.get('ACTIVE/NOACTIVE', 'ACTIVE')).upper()
    status = 'active' if status_value == 'ACTIVE' else 'inactive'
    
    return {
        'first_name': first_name,
        'last_name': last_name,
        'sex': str(record.get('SEX', '')).upper(),
        'dob': dob,
        'nationality': str(record.get('NATIONALITY', '')),
        'phone_number': str(record.get('PHONE_NUMBER', '')),
        'zone_name': zone_name,
        'building_name': building_name,
        'floor_name': floor_name,
        'position_name': str(record.get('POSITION', '')),
        'joined_date': joined_date,
        'status': status,
    }

def _process_worker_photos_from_json(worker, record, request):
    """Process worker photos and documents from JSON record data with URLs."""
    from django.core.files.base import ContentFile
    from django.conf import settings
    import os
    
    
    # Map column names to document types and field types
    photo_columns = {
        'PHOTO': {'field_type': 'photo', 'doc_type': None},
        'IDKH/PASSPORT': {'field_type': 'document', 'doc_type': 'passport'},
        'WORKINGPERMIT': {'field_type': 'document', 'doc_type': 'work_permit'},
        'VISA': {'field_type': 'document', 'doc_type': 'visa'},
    }
    
    for column_name, config in photo_columns.items():
        image_url = record.get(column_name)
        
        # Check if this column has an image URL
        if image_url and isinstance(image_url, str) and image_url.startswith('/media/'):
            try:
                # Convert URL to file path
                image_path = os.path.join(settings.MEDIA_ROOT, image_url.replace('/media/', ''))
                
                if os.path.exists(image_path):
                    
                    # Read image data
                    with open(image_path, 'rb') as f:
                        image_data = f.read()
                    
                    filename = os.path.basename(image_path)
                    
                    if config['field_type'] == 'photo':
                        # Save worker photo
                        worker.photo.save(filename, ContentFile(image_data), save=True)
                        
                    elif config['field_type'] == 'document':
                        # Create document record with proper metadata
                        doc_type = config['doc_type']
                        _create_document_with_metadata(worker, image_data, filename, doc_type, record, request.user)
                        
                else:
                    pass
                        
            except Exception as e:
                pass
        
        else:
            pass


def cleanup_extracted_image_files(extracted_images):
    """Clean up extracted image files and folders after successful import."""
    import shutil
    
    try:
        # Get unique folder paths from extracted images
        folders_to_cleanup = set()
        for image_info in extracted_images.values():
            if 'path' in image_info and image_info['path']:
                folder_path = os.path.dirname(image_info['path'])
                folders_to_cleanup.add(folder_path)
        
        # Remove each unique folder
        for folder_path in folders_to_cleanup:
            if os.path.exists(folder_path):
                try:
                    shutil.rmtree(folder_path)
                except Exception as e:
                    pass
        
        
    except Exception as e:
        pass


def _create_document_with_metadata(worker, image_data, filename, doc_type, record, created_by):
    """Create or update document record with metadata extracted from Excel."""
    from zone.models import Document
    from django.core.files.base import ContentFile
    
    try:
        # Get document metadata based on document type and Excel columns
        document_number = ''
        issue_date = None
        expiry_date = None
        
        
        if doc_type == 'passport':
            document_number = str(record.get('DOCUMENT NO', '')).strip() if record.get('DOCUMENT NO') else ''
            # Parse IDKH/PASSPORT_DATE (format: "14/11/2018-14/11/2028")
            date_range = str(record.get('IDKH/PASSPORT_DATE', '')).strip() if record.get('IDKH/PASSPORT_DATE') else ''
            if date_range and '-' in date_range:
                date_parts = date_range.split('-')
                if len(date_parts) == 2:
                    try:
                        from datetime import datetime
                        issue_date = datetime.strptime(date_parts[0].strip(), '%d/%m/%Y').date()
                        expiry_date = datetime.strptime(date_parts[1].strip(), '%d/%m/%Y').date()
                    except ValueError as e:
                        pass
        
        elif doc_type == 'work_permit':
            # Extract work permit document number
            # Try various possible column names for work permit number
            for col_name in ['WORKINGPERMIT NO', 'WORKINGPERMIT_NO', 'WORKPERMIT_NO', 'WORK_PERMIT_NO', 'WORKPERMITNO']:
                if record.get(col_name):
                    document_number = str(record.get(col_name, '')).strip()
                    break
            
            # Parse WORKINGPERMIT_DATE
            date_range = str(record.get('WORKINGPERMIT_DATE', '')).strip() if record.get('WORKINGPERMIT_DATE') else ''
            if date_range and '-' in date_range:
                date_parts = date_range.split('-')
                if len(date_parts) == 2:
                    try:
                        from datetime import datetime
                        issue_date = datetime.strptime(date_parts[0].strip(), '%d/%m/%Y').date()
                        expiry_date = datetime.strptime(date_parts[1].strip(), '%d/%m/%Y').date()
                    except ValueError as e:
                        pass
        
        elif doc_type == 'visa':
            # Extract visa document number  
            # Try various possible column names for visa number
            for col_name in ['VISA NO', 'VISA_NO', 'VISANO', 'VISA_NUMBER']:
                if record.get(col_name):
                    document_number = str(record.get(col_name, '')).strip()
                    break
            
            # Parse VISA_DATE
            date_range = str(record.get('VISA_DATE', '')).strip() if record.get('VISA_DATE') else ''
            if date_range and '-' in date_range:
                date_parts = date_range.split('-')
                if len(date_parts) == 2:
                    try:
                        from datetime import datetime
                        issue_date = datetime.strptime(date_parts[0].strip(), '%d/%m/%Y').date()
                        expiry_date = datetime.strptime(date_parts[1].strip(), '%d/%m/%Y').date()
                    except ValueError as e:
                        pass
        
        # Fallback: Try to extract document number from any field if not found yet
        if not document_number or not document_number.strip():
            import re
            # Pattern for document numbers (letters + numbers, 6+ chars)  
            # Exclude common non-document values
            doc_pattern = r'^[A-Z0-9]{6,}$'
            excluded_values = {'VIETNAM', 'CAMBODIA', 'THAILAND', 'LAOS', 'MYANMAR', 'SINGAPORE', 'MALAYSIA', 'PHILIPPINES', 'INDONESIA'}
            
            for key, value in record.items():
                if value and isinstance(value, str) and len(value.strip()) > 0:
                    clean_value = value.strip().upper()
                    if re.match(doc_pattern, clean_value) and clean_value not in excluded_values:
                        document_number = value.strip()
                        break
        
        # Try to find existing document to update, or create new one
        existing_doc = None
        if document_number:
            # If we have a document number, try to find existing doc by number
            existing_doc = worker.documents.filter(
                document_type=doc_type, 
                document_number=document_number
            ).first()
        else:
            # If no document number, try to find existing doc by type without file
            existing_doc = worker.documents.filter(
                document_type=doc_type,
                document_file__in=['', None]
            ).first()
        
        if existing_doc:
            # Update existing document
            document = existing_doc
            # Only update document number if we have a new one
            if document_number and document_number.strip():
                document.document_number = document_number
            # Only update dates if we have valid ones
            if issue_date:
                document.issue_date = issue_date
            if expiry_date:
                document.expiry_date = expiry_date
            document.issuing_authority = 'Immigration' if doc_type in ['passport', 'visa'] else 'Labor Department'
            document.save()
        else:
            # Only create new document if we have proper dates
            if not issue_date or not expiry_date:
                return  # Don't create document without proper dates
            
            document = Document.objects.create(
                worker=worker,
                document_type=doc_type,
                document_number=document_number,
                issue_date=issue_date,
                expiry_date=expiry_date,
                issuing_authority='Immigration' if doc_type in ['passport', 'visa'] else 'Labor Department',
                created_by=created_by
            )
        
        # Save document image
        document.document_file.save(filename, ContentFile(image_data), save=True)
        
        
    except Exception as e:
        pass


def _process_worker_photos_direct(worker, record, extracted_images, worker_index, created_by):
    """Process worker photos and documents directly from extracted images using Excel mapping."""
    from django.core.files.base import ContentFile
    import os
    
    
    # Check which image columns have references in the record
    image_columns = ['PHOTO', 'IDKH/PASSPORT', 'WORKINGPERMIT', 'VISA']
    
    # Look for image references that were added to the record during preview
    images_to_process = []
    for column_name in image_columns:
        if column_name in record and record[column_name]:
            # This column has an image reference
            images_to_process.append(column_name)
    
    # Map document types
    doc_type_mapping = {
        'PHOTO': None,
        'IDKH/PASSPORT': 'passport',
        'WORKINGPERMIT': 'work_permit',
        'VISA': 'visa'
    }
    
    processed_images = 0
    
    # Process images based on what's in the record
    for column_name in images_to_process:
        # Extract the image URL/path from the record
        image_ref = record[column_name]
        
        # Find the corresponding image in extracted_images
        # The image_ref might be a URL like '/media/worker_import_images/...' 
        # We need to find the matching image in extracted_images
        image_found = False
        
        for img_key, img_info in extracted_images.items():
            if img_info.get('url') == image_ref or img_info.get('filename') in image_ref:
                try:
                    
                    # Get image file path
                    image_path = img_info.get('path')
                    if not image_path:
                        continue
                    
                    if not os.path.exists(image_path):
                        # Try alternative path if it's a Windows path issue
                        if '\\' in image_path:
                            image_path = image_path.replace('\\', '/')
                            if not os.path.exists(image_path):
                                continue
                        else:
                            continue
                    
                    # Read image data
                    with open(image_path, 'rb') as f:
                        image_data = f.read()
                    
                    filename = img_info.get('filename', f'image_{img_key}')
                    
                    if column_name == 'PHOTO':
                        # Save worker photo
                        try:
                            worker.photo.save(filename, ContentFile(image_data), save=True)
                            processed_images += 1
                        except Exception as photo_error:
                            pass
                        
                    else:
                        # Create document
                        doc_type = doc_type_mapping.get(column_name)
                        if doc_type:
                            _create_document_with_metadata(worker, image_data, filename, doc_type, record, created_by)
                            processed_images += 1
                    
                    image_found = True
                    break  # Found and processed the image, move to next column
                        
                except Exception as e:
                    pass
        
        if not image_found:
            pass
        
    
    # Also create documents for dates without images (like VISA with dates but no image)
    # Check for visa dates without image
    if record.get('VISA_DATE') and not record.get('VISA'):
        date_range = str(record.get('VISA_DATE', '')).strip()
        if date_range and '-' in date_range:
            date_parts = date_range.split('-')
            if len(date_parts) == 2:
                try:
                    from datetime import datetime
                    visa_issue = datetime.strptime(date_parts[0].strip(), '%d/%m/%Y').date()
                    visa_expiry = datetime.strptime(date_parts[1].strip(), '%d/%m/%Y').date()
                    
                    # Check if visa document already exists
                    existing_visa = worker.documents.filter(document_type='visa').first()
                    if not existing_visa:
                        Document.objects.create(
                            worker=worker,
                            document_type='visa',
                            document_number='',
                            issue_date=visa_issue,
                            expiry_date=visa_expiry,
                            issuing_authority='Immigration',
                            created_by=created_by
                        )
                except Exception as e:
                    pass
    
    # If no images were processed at all, create all documents based on dates
    if processed_images == 0:
        _create_worker_documents(worker, _extract_worker_data_from_excel_record(record), created_by)
    


def _process_document_from_mrz_data(worker, image_data, image_info, document_type, created_by):
    """Process document using already extracted MRZ data instead of re-running OCR."""
    try:
        from django.core.files.base import ContentFile
        from sinosecu.models import PassportScan, ScanImage, ScanResult
        import json
        
        # Get the MRZ data
        mrz_data = image_info.get('document', {})
        mrz_text = image_info.get('mrz', '')  # This is the raw MRZ text
        

        # document_data = parse_mrz_lines(mrz_data.get("raw_mrz_lines"))
        # print("document_data==:", document_data)
     
        # Create a PassportScan record
        # Generate a unique scan_id
        import uuid
        scan_id = f"mrz_import_{uuid.uuid4().hex[:8]}_{worker.id}"
        
        # Check for KHM patterns in MRZ text to force ID Card classification
        mrz_text = image_info.get('mrz', '')
        raw_mrz_lines = mrz_data.get('raw_mrz_lines', [])
        
        
        # Force ID Card if we detect KHM patterns anywhere in the MRZ
        has_khm_pattern = False
        
        # Check mrz_text first
        if mrz_text and ('KHM' in mrz_text.upper() or 'IDKHM' in mrz_text.upper() or 'LDKHM' in mrz_text.upper()):
            has_khm_pattern = True
        
        # Check raw_mrz_lines
        elif raw_mrz_lines and any('KHM' in str(line).upper() or 'IDKHM' in str(line).upper() or 'LDKHM' in str(line).upper() for line in raw_mrz_lines):
            has_khm_pattern = True
        
        # Additional fallback: check all available data in image_info for KHM patterns
        else:
            # Check if there's MRZ data anywhere in the image_info structure
            image_info_str = str(image_info)
            if 'KHM' in image_info_str.upper() or 'IDKHM' in image_info_str.upper() or 'LDKHM' in image_info_str.upper():
                has_khm_pattern = True
        
        if has_khm_pattern:
            final_doc_type = 'id_card'
        else:
            # Use document type from MRZ parsing if available, otherwise use the parameter
            detected_doc_type = mrz_data.get('document_type', document_type)
            if detected_doc_type == 'id_card':
                final_doc_type = 'id_card'
            else:
                final_doc_type = 'passport'  # Default to passport for unknown types
            
        scan = PassportScan.objects.create(
            user=created_by,
            scan_id=scan_id,
            document_type=final_doc_type,
            status='completed'
        )
        
        
        # Save the image
        image_filename = f"scan_{scan.id}_{final_doc_type}.jpg"

        
        scan_image = ScanImage.objects.create(
            scan=scan,
            image=ContentFile(image_data, name=image_filename),
            is_primary=True
        )
        
        # Parse dates properly
        from datetime import datetime as dt
        
        def parse_date(date_str):
            """Parse date string to date object or None"""
            if not date_str or date_str == 'N/A':
                return None
            try:
                # Try common date formats
                for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%Y/%m/%d', '%d/%m/%Y']:
                    try:
                        return dt.strptime(date_str, fmt).date()
                    except ValueError:
                        continue
                return None
            except:
                return None
        
        # Extract MRZ lines from raw text or raw_mrz_lines
        mrz_lines = []
        if mrz_text:
            # If we have raw MRZ text, split it into lines
            mrz_lines = [line.strip() for line in mrz_text.split('\n') if line.strip()]
        elif mrz_data.get('raw_mrz_lines'):
            # Otherwise use the raw_mrz_lines if available
            mrz_lines = mrz_data.get('raw_mrz_lines', [])
        
        # Create ScanResult with MRZ data (only use fields that exist in the model)
        scan_result = ScanResult.objects.create(
            scan=scan,
            ocr_confidence=95.0,  # High confidence since MRZ was successfully parsed
            document_type=mrz_data.get('document_type', document_type),
            document_number=mrz_data.get('document_number', ''),
            issuing_country=mrz_data.get('country_code', ''),
            nationality=mrz_data.get('nationality', ''),
            expiry_date=parse_date(mrz_data.get('expiry_date')),
            date_of_birth=parse_date(mrz_data.get('date_of_birth')),
            gender=mrz_data.get('sex', ''),  # Use 'gender' field name, not 'sex'
            surname=mrz_data.get('surname', ''),
            given_names=mrz_data.get('given_names', ''),
            # Store MRZ lines if available
            mrz_line1=mrz_lines[0] if len(mrz_lines) > 0 else '',
            mrz_line2=mrz_lines[1] if len(mrz_lines) > 1 else ''
        )
        
        scan.result = scan_result
        scan.save()
        
        # new extract
  
        # Pass the raw MRZ text to document creation
        print("create_document_from_embedded_image_with_mrz mrz_text=:",mrz_text)
        print("create_document_from_embedded_image_with_mrz mrz_lines=:",mrz_lines)
       
        _create_document_from_embedded_image_with_mrz(worker, scan, document_type, image_data, image_info.get('filename', 'unknown.jpg'), created_by, mrz_text, mrz_lines)
        
        # Update worker info if MRZ provided better data
        _update_worker_from_ocr(worker, scan_result)
        
        
    except Exception as e:
        # Fallback to original OCR method
                # Get the MRZ data
        mrz_data = image_info.get('document', {})
        mrz_text = image_info.get('mrz', '')  # This is the raw MRZ text
        mrz_lines = mrz_data.get("raw_mrz_lines")
        _process_document_from_bytes_with_sinosecu(worker, image_data, document_type, created_by, mrz_text, mrz_lines)

def set_process_excel_import(request, form, ajax_mode=False, progress_tracker=None):
    """Process the Excel file import with optional AJAX mode for progress tracking."""
    import tempfile
    import json
    from django.core.files.base import ContentFile
    from sinosecu.models import PassportScan, ScanImage, ScanResult
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
    from PIL import Image as PILImage
    import io
    
    excel_file = form.cleaned_data['excel_file']
    process_photos = form.cleaned_data.get('process_photos', True)
    skip_duplicates = form.cleaned_data.get('skip_duplicates', True)
    
    results = {
        'total_rows': 0,
        'processed': 0,
        'skipped': 0,
        'duplicates': [],  # Track duplicate details separately
        'errors': [],
        'created_workers': [],
    }
    
    tmp_file_name = None
    try:
        # Update progress: Starting
        if progress_tracker:
            progress_tracker.update(
                stage=ProgressStages.STARTING,
                stage_message=STAGE_MESSAGES[ProgressStages.STARTING]
            )
        
        # Save Excel file temporarily
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            tmp_file_name = tmp_file.name
            for chunk in excel_file.chunks():
                tmp_file.write(chunk)
            tmp_file.flush()
        
        # Update progress: Reading file
        if progress_tracker:
            progress_tracker.update(
                stage=ProgressStages.READING_FILE,
                stage_message=STAGE_MESSAGES[ProgressStages.READING_FILE]
            )
        
        # Extract embedded images from Excel
        embedded_images = _extract_embedded_images_from_excel(tmp_file_name)
        
        # Read Excel file
        df = pd.read_excel(tmp_file_name, sheet_name='Workers')
        
        # Filter out empty rows
        df = df.dropna(subset=['NAME'])
        results['total_rows'] = len(df)
        
        
        # Update progress: Processing workers
        if progress_tracker:
            progress_tracker.update(
                stage=ProgressStages.PROCESSING_WORKERS,
                stage_message=STAGE_MESSAGES[ProgressStages.PROCESSING_WORKERS]
            )
        
        # Process each row
        for index, row in df.iterrows():
            try:
                # Update progress for current worker
                if progress_tracker:
                    worker_name = str(row.get('NAME', f'Worker at row {index + 1}')).strip()
                    progress_tracker.update(
                        current_item=index + 1,
                        current_worker_name=worker_name,
                        processed=results['processed'],
                        skipped=results['skipped'],
                        failed=len(results['errors'])
                    )
                
                worker_data = _extract_worker_data(row)
                
                # Check for duplicates
                duplicate_info = _is_duplicate_worker(worker_data)
                if skip_duplicates and duplicate_info['is_duplicate']:
                    results['skipped'] += 1
                    # Store duplicate details separately for clearer messaging
                    duplicate_detail = {
                        'row': index + 2,
                        'worker_name': f"{worker_data['first_name']} {worker_data['last_name']}",
                        'matching_workers': duplicate_info['matching_workers'],
                        'reasons': duplicate_info['match_reasons']
                    }
                    results['duplicates'].append(duplicate_detail)
                    continue
                
                # Create worker with transaction
                with transaction.atomic():
                    worker = _create_worker_from_data(worker_data, request.user)
                    
                    # Process photos if enabled
                    if process_photos:
                        # Pass embedded images along with row data
                        _process_worker_photos(worker, row, request, embedded_images, index)
                    else:
                        pass
                    
                    results['processed'] += 1
                    results['created_workers'].append({
                        'name': worker.get_full_name(),
                        'worker_id': worker.worker_id,
                    })
                    
            except Exception as e:
                error_msg = f"Row {index + 2}: {str(e)}"
                results['errors'].append(error_msg)
    
    except Exception as e:
        results['errors'].append(f"File processing error: {str(e)}")
    finally:
        # Clean up temp file
        if tmp_file_name and os.path.exists(tmp_file_name):
            try:
                os.unlink(tmp_file_name)
            except OSError as e:
                pass
    
    # Show results with detailed messages
    if results['processed'] > 0:
        messages.success(request, f"✓ Successfully imported {results['processed']} new worker(s).")
    
    # Show detailed duplicate information
    if results['duplicates']:
        duplicate_msg = f"⚠ Skipped {results['skipped']} duplicate worker(s):"
        messages.warning(request, duplicate_msg)
        
        # Add details for each duplicate (limit to first 5 for readability)
        for dup in results['duplicates'][:5]:
            existing_info = []
            for worker in dup['matching_workers']:
                existing_info.append(f"{worker['name']} (ID: {worker['worker_id']}, Zone: {worker['zone']})")
            
            detail_msg = f"• Row {dup['row']}: {dup['worker_name']} matches existing: {', '.join(existing_info)}"
            messages.info(request, detail_msg)
        
        if len(results['duplicates']) > 5:
            messages.info(request, f"... and {len(results['duplicates']) - 5} more duplicate(s)")
    
    # Show errors if any
    if results['errors']:
        error_summary = f"✗ Encountered {len(results['errors'])} error(s):"
        messages.error(request, error_summary)
        
        # Show first 3 errors
        for error in results['errors'][:3]:
            messages.error(request, f"• {error}")
        
        if len(results['errors']) > 3:
            messages.error(request, f"... and {len(results['errors']) - 3} more error(s)")
    
    # Summary message
    total_attempted = results['total_rows']
    if total_attempted > 0:
        success_rate = (results['processed'] / total_attempted) * 100
        summary = f"Import Summary: {results['processed']}/{total_attempted} workers imported ({success_rate:.0f}% success rate)"
        if results['skipped'] > 0:
            summary += f", {results['skipped']} duplicates skipped"
        messages.info(request, summary)
    
    # Store import results in session for display on worker list page
    request.session['import_results'] = {
        'total_rows': results['total_rows'],
        'processed': results['processed'],
        'skipped': results['skipped'],
        'duplicates': [
            {
                'row': dup['row'],
                'worker_name': dup['worker_name'],
                'matches': ', '.join([f"{w['name']} (ID: {w['worker_id']})" for w in dup['matching_workers'][:2]])
            }
            for dup in results['duplicates'][:10]  # Store first 10 for display
        ],
        'errors': results['errors'][:5],  # Store first 5 errors
        'timestamp': timezone.now().isoformat()
    }
    
    # Return results dictionary for AJAX mode, otherwise redirect
    if ajax_mode:
        return results
    else:
        return redirect('zone:worker_list')

def mrz_extract_image(img_data):
   
    try:
        # Fix PIL compatibility issue
        import PIL
        if not hasattr(PIL.Image, 'ANTIALIAS'):
            PIL.Image.ANTIALIAS = PIL.Image.LANCZOS
        
        # Initialize with English and other languages that might be in passports
        reader = easyocr.Reader(['en'], gpu=False)  # Set gpu=True if you have CUDA
        print("EasyOCR initialized successfully")
    except Exception as e:
        print(f"Warning: Could not initialize EasyOCR: {e}")
        reader = None

    if reader:
        try:
            print(f"    Using EasyOCR for text extraction...")
            
            # EasyOCR can read directly from file path
            result = reader.readtext(img_data, detail=1)
            
            # Combine all text from EasyOCR
            easyocr_text = '\n'.join([text[1] for text in result])
            
            # Also get bounding boxes and confidence scores
            ocr_details = []
            for bbox, text, conf in result:
                # Convert numpy arrays to lists for JSON serialization
                bbox_list = [[float(x), float(y)] for x, y in bbox] if bbox is not None else None
                ocr_details.append({
                    "text": text,
                    "confidence": float(conf),
                    "bbox": bbox_list
                })
            
            # Extract MRZ lines from EasyOCR results
            all_lines = easyocr_text.strip().split('\n')
            mrz_lines = []
            
            # First, try to find clear MRZ patterns
            for line in all_lines:
                line = line.strip()
                # Check for MRZ patterns
                if (('<' in line and len(line) > 20) or
                    ('P<' in line.upper()) or
                    ('ID<' in line.upper()) or
                    (len(line) > 35 and line.count('<') > 3) or
                    (len(line) > 30 and any(c in line for c in ['<', '«']) and any(c.isdigit() for c in line))):
                    mrz_lines.append(line)
            
            # If no MRZ found, look more aggressively in OCR details
            if not mrz_lines:
                # Look for lines that might be MRZ but poorly OCR'd
                for detail in ocr_details:
                    text = detail['text'].strip()
                    # Look for patterns typical of MRZ
                    if (len(text) > 25 and 
                        (('<' in text) or 
                            ('P<' in text.upper()) or
                            (text.count(' ') < 3 and any(c.isdigit() for c in text) and any(c.isupper() for c in text)) or
                            (any(char in text for char in ['<', '«', 'T', 'A']) and len([c for c in text if c.isdigit()]) > 5))):
                        mrz_lines.append(text)
            
            # If still no MRZ found, try enhanced preprocessing with Tesseract
            if not mrz_lines:
                print(f"    No MRZ found with EasyOCR, trying enhanced preprocessing...")
                try:
                    # Load image and preprocess
                    img_pil = Image.open(img_data)
                    processed_img = preprocess_image_for_mrz(img_pil)
                    processed_pil = Image.fromarray(processed_img)
                    
                    # Try OCR with different settings
                    configs = [
                        '--oem 3 --psm 6',
                        '--oem 3 --psm 7',
                        '-c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789< --oem 3 --psm 7'
                    ]
                    
                    for config in configs:
                        text = pytesseract.image_to_string(processed_pil, config=config)
                        lines = [line.strip() for line in text.split('\n') if line.strip() and len(line.strip()) > 20]
                        
                        for line in lines:
                            # Check for MRZ patterns
                            if (('<' in line and len(line) > 25) or
                                ('P<' in line.upper()) or
                                (len(line) > 35 and any(c.isdigit() for c in line) and any(c.isupper() for c in line))):
                                mrz_lines.append(line)
                        
                        if mrz_lines:
                            print(f"    Found {len(mrz_lines)} MRZ lines with enhanced preprocessing")
                            break
                            
                except Exception as e:
                    print(f"    Enhanced preprocessing failed: {str(e)[:100]}")
            
            # Clean and format MRZ output
            if mrz_lines:
                # Clean each MRZ line to remove invalid characters
                cleaned_lines = []
                for line in mrz_lines:
                    # Keep only valid MRZ characters: A-Z, 0-9, <
                    cleaned_line = ''.join(c for c in line.upper() if c.isalnum() or c == '<')
                    if len(cleaned_line) > 20:  # Valid MRZ lines are typically long
                        cleaned_lines.append(cleaned_line)
                
                if cleaned_lines:
                    passport_mrz = '\n'.join(cleaned_lines)
                    print(f"    Found {len(cleaned_lines)} cleaned MRZ lines with EasyOCR")
                else:
                    passport_mrz = None
                    print(f"    No valid MRZ lines after cleaning")
            else:
                passport_mrz = None
                print(f"    No MRZ lines detected, but extracted {len(result)} text regions")
            
        except Exception as e:
            # Fallback to Tesseract OCR
            print(f"    EasyOCR failed: {str(e)[:100]}")
            print(f"    Falling back to Tesseract OCR...")
            img = Image.open(img_data)
            
            # Try OCR with English language
            custom_config = r'--oem 3 --psm 6'
            text = pytesseract.image_to_string(img, lang='eng', config=custom_config)
            
            # Extract MRZ lines
            all_lines = text.strip().split('\n')
            mrz_lines = []
           
           
            
            # Clean and format MRZ output
            if mrz_lines:
                # Clean each MRZ line to remove invalid characters
                cleaned_lines = []
                for line in mrz_lines:
                    # Keep only valid MRZ characters: A-Z, 0-9, <
                    cleaned_line = ''.join(c for c in line.upper() if c.isalnum() or c == '<')
                    if len(cleaned_line) > 20:  # Valid MRZ lines are typically long
                        cleaned_lines.append(cleaned_line)
                
                if cleaned_lines:
                    passport_mrz = '\n'.join(cleaned_lines)
                    print(f"    Found {len(cleaned_lines)} cleaned MRZ lines")
                else:
                    passport_mrz = None
                    print(f"    No valid MRZ lines after cleaning")
            else:
                passport_mrz = None
    else:
        # No FastMRZ, use basic OCR
        img = Image.open(img_data)
        custom_config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(img, lang='eng', config=custom_config)
        
        # Extract MRZ lines
        all_lines = text.strip().split('\n')
        mrz_lines = [line.strip() for line in all_lines if '<' in line and len(line.strip()) > 20]
        
        # Clean and format MRZ output
        if mrz_lines:
            # Clean each MRZ line to remove invalid characters
            cleaned_lines = []
            for line in mrz_lines:
                # Keep only valid MRZ characters: A-Z, 0-9, <
                cleaned_line = ''.join(c for c in line.upper() if c.isalnum() or c == '<')
                if len(cleaned_line) > 20:  # Valid MRZ lines are typically long
                    cleaned_lines.append(cleaned_line)
            
            if cleaned_lines:
                passport_mrz = '\n'.join(cleaned_lines)
                print(f"    Found {len(cleaned_lines)} cleaned MRZ lines")
            else:
                passport_mrz = None
                print(f"    No valid MRZ lines after cleaning")
        else:
            passport_mrz = None
        print("passport_mrz :", passport_mrz)   

    # Initialize document_mrz for all paths
    # document_mrz = {}
    
    # Process MRZ if lines were found (check both EasyOCR and basic OCR paths)
    # if 'mrz_lines' in locals() and mrz_lines:
    #     doucType = detect_mrz_type(mrz_lines)
                        
    #     if doucType == "KHID":
    #         document_mrz = parse_khid_mrz(mrz_lines)
    #         print("KHID:", parse_khid_mrz(mrz_lines))
    #     elif doucType == "PASSPORT":
    #         document_mrz = parse_passport_mrz(mrz_lines)
    #         print("PASSPORT: ", parse_passport_mrz(mrz_lines))

    # print("    OCR processing complete!")
    # document_context={}
    # if document_mrz: 
    #     document_context = {
    #         "document_type": document_mrz["document_type"],
    #         "document_number": document_mrz["document_number"],
    #         "issuing_country": document_mrz["issuing_country"],
    #         "expiry_date": document_mrz["expiry_date"],
    #     }
    
    return {
        "mrz_text":passport_mrz,
        "mrz_lines":mrz_lines
        }
        

def preprocess_image_for_mrz(image):
    """Enhanced image preprocessing specifically for MRZ extraction"""
    # Convert PIL Image to OpenCV format (if needed)
    if isinstance(image, Image.Image):  # PIL Image
        image = np.array(image)
        image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)  # PIL uses RGB; OpenCV uses BGR
    
    # Convert to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # Increase contrast with CLAHE
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
    enhanced = clahe.apply(gray)
    
    # Apply threshold
    _, thresh = cv2.threshold(enhanced, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # Scale up the image for better OCR
    scale_factor = 3
    height, width = thresh.shape
    resized = cv2.resize(thresh, (width * scale_factor, height * scale_factor), interpolation=cv2.INTER_CUBIC)
    
    return resized



