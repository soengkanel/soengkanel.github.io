#!/usr/bin/env python3
"""
NextHR Project Planning Generator
Creates comprehensive project plan with modules, tasks, PIC, and estimates
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_project_plan():
    wb = Workbook()
    ws = wb.active
    ws.title = "NextHR Project Plan"

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    module_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
    task_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    completed_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    inprogress_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    pending_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    module_font = Font(bold=True, size=11)
    task_font = Font(size=10)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Project data structure
    project_data = [
        {
            "module": "1. Core Infrastructure",
            "status": "Completed",
            "progress": "100%",
            "tasks": [
                {"task": "Multi-tenancy setup with PostgreSQL schemas", "pic": "Kanel", "est": 5, "status": "Completed"},
                {"task": "Database models and migrations", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Tenant routing and middleware", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Environment configuration (.env setup)", "pic": "Kanel", "est": 1, "status": "Completed"},
                {"task": "Static files and media handling", "pic": "Kanel", "est": 2, "status": "Completed"},
            ]
        },
        {
            "module": "2. User Management & Authentication",
            "status": "Completed",
            "progress": "95%",
            "tasks": [
                {"task": "Custom user model and authentication", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Role-based access control (RBAC)", "pic": "Kanel", "est": 5, "status": "Completed"},
                {"task": "Permission system implementation", "pic": "Kanel", "est": 4, "status": "Completed"},
                {"task": "User groups and hierarchy", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Login/logout functionality", "pic": "Chhay", "est": 2, "status": "Completed"},
                {"task": "Password reset and management", "pic": "Chhay", "est": 2, "status": "Completed"},
            ]
        },
        {
            "module": "3. Company & Organization Setup",
            "status": "Completed",
            "progress": "90%",
            "tasks": [
                {"task": "Company profile management", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Department structure", "pic": "Kanel", "est": 2, "status": "Completed"},
                {"task": "Position/Job title management", "pic": "Kanel", "est": 2, "status": "Completed"},
                {"task": "Zone/Building/Floor hierarchy", "pic": "Kanel", "est": 4, "status": "Completed"},
                {"task": "Master data scripts", "pic": "Kanel", "est": 2, "status": "Completed"},
            ]
        },
        {
            "module": "4. HR & Employee Management",
            "status": "In Progress",
            "progress": "70%",
            "tasks": [
                {"task": "Employee profile (personal info, documents)", "pic": "Chhay", "est": 5, "status": "Completed"},
                {"task": "Employee photo upload and management", "pic": "Chhay", "est": 2, "status": "Completed"},
                {"task": "Employment history tracking", "pic": "Chhay", "est": 3, "status": "Completed"},
                {"task": "Contract management", "pic": "Chhay", "est": 4, "status": "In Progress"},
                {"task": "Probation tracking", "pic": "Chhay", "est": 3, "status": "In Progress"},
                {"task": "Employee search and filtering", "pic": "Bonna", "est": 3, "status": "Completed"},
                {"task": "Employee status management (active/inactive)", "pic": "Chhay", "est": 2, "status": "Completed"},
                {"task": "Employee autocomplete component", "pic": "Bonna", "est": 2, "status": "Completed"},
            ]
        },
        {
            "module": "5. ID Card Management",
            "status": "Completed",
            "progress": "95%",
            "tasks": [
                {"task": "ID card request workflow", "pic": "Chhay", "est": 4, "status": "Completed"},
                {"task": "Card printing and generation", "pic": "Chhay", "est": 3, "status": "Completed"},
                {"task": "Card lifecycle (issue, renew, replace, return)", "pic": "Chhay", "est": 5, "status": "Completed"},
                {"task": "QR code generation", "pic": "Chhay", "est": 2, "status": "Completed"},
                {"task": "Card history and audit trail", "pic": "Chhay", "est": 2, "status": "Completed"},
            ]
        },
        {
            "module": "6. Leave Management",
            "status": "Completed",
            "progress": "90%",
            "tasks": [
                {"task": "Leave type configuration", "pic": "Kanel", "est": 2, "status": "Completed"},
                {"task": "Leave balance tracking", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Leave request and approval workflow", "pic": "Chhay", "est": 5, "status": "Completed"},
                {"task": "Leave calendar view", "pic": "Bonna", "est": 3, "status": "Completed"},
                {"task": "Leave reports and analytics", "pic": "Chhay", "est": 3, "status": "Completed"},
            ]
        },
        {
            "module": "7. Attendance & Timecard",
            "status": "In Progress",
            "progress": "50%",
            "tasks": [
                {"task": "Attendance marking (manual)", "pic": "Chhay", "est": 3, "status": "Completed"},
                {"task": "Fingerprint integration (back office)", "pic": "Kanel", "est": 5, "status": "In Progress"},
                {"task": "Timecard Excel import template", "pic": "Kanel", "est": 3, "status": "Pending"},
                {"task": "Timecard verification UI", "pic": "Chhay", "est": 4, "status": "Pending"},
                {"task": "Cut-off date logic (20th monthly)", "pic": "Kanel", "est": 2, "status": "Pending"},
                {"task": "Attendance reports", "pic": "Chhay", "est": 3, "status": "In Progress"},
                {"task": "Manual timecard fallback", "pic": "Chhay", "est": 2, "status": "Pending"},
            ]
        },
        {
            "module": "8. Overtime Management",
            "status": "In Progress",
            "progress": "55%",
            "tasks": [
                {"task": "OT types (Normal ×1, Holiday ×1.5)", "pic": "Kanel", "est": 2, "status": "Completed"},
                {"task": "OT claim form", "pic": "Chhay", "est": 3, "status": "Completed"},
                {"task": "OT Excel import template", "pic": "Kanel", "est": 3, "status": "Pending"},
                {"task": "Multi-project OT tracking", "pic": "Kanel", "est": 4, "status": "In Progress"},
                {"task": "HR verification workflow", "pic": "Chhay", "est": 4, "status": "Pending"},
                {"task": "OT approval workflow (Manager)", "pic": "Chhay", "est": 3, "status": "Pending"},
                {"task": "OT adjustment interface", "pic": "Chhay", "est": 2, "status": "Pending"},
            ]
        },
        {
            "module": "9. Payroll System",
            "status": "In Progress",
            "progress": "60%",
            "tasks": [
                {"task": "Salary structure and components", "pic": "Kanel", "est": 4, "status": "Completed"},
                {"task": "Payroll period setup", "pic": "Kanel", "est": 2, "status": "Completed"},
                {"task": "Salary calculation engine", "pic": "Kanel", "est": 5, "status": "In Progress"},
                {"task": "OT integration with payroll", "pic": "Kanel", "est": 4, "status": "Pending"},
                {"task": "Deductions and allowances", "pic": "Kanel", "est": 3, "status": "In Progress"},
                {"task": "Payslip generation", "pic": "Chhay", "est": 4, "status": "In Progress"},
                {"task": "Payroll summary reports", "pic": "Chhay", "est": 3, "status": "Pending"},
                {"task": "Bulk payslip generation", "pic": "Kanel", "est": 2, "status": "Pending"},
            ]
        },
        {
            "module": "10. Project Management",
            "status": "In Progress",
            "progress": "65%",
            "tasks": [
                {"task": "Project creation and setup", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Project team allocation", "pic": "Chhay", "est": 4, "status": "In Progress"},
                {"task": "Project timeline and milestones", "pic": "Chhay", "est": 3, "status": "In Progress"},
                {"task": "Project status tracking", "pic": "Chhay", "est": 2, "status": "Completed"},
                {"task": "Multi-project employee assignment", "pic": "Kanel", "est": 3, "status": "In Progress"},
            ]
        },
        {
            "module": "11. E-Forms System",
            "status": "In Progress",
            "progress": "40%",
            "tasks": [
                {"task": "Form builder infrastructure", "pic": "Kanel", "est": 5, "status": "Completed"},
                {"task": "Worker registration form", "pic": "Chhay", "est": 4, "status": "Completed"},
                {"task": "Document upload forms", "pic": "Chhay", "est": 3, "status": "In Progress"},
                {"task": "Form validation and MRZ scanning", "pic": "Kanel", "est": 4, "status": "In Progress"},
                {"task": "6 additional form pages", "pic": "Chhay", "est": 12, "status": "Pending"},
                {"task": "Form submission workflow", "pic": "Chhay", "est": 3, "status": "Pending"},
            ]
        },
        {
            "module": "12. Billing & Payments",
            "status": "Completed",
            "progress": "85%",
            "tasks": [
                {"task": "Service charge configuration", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Billing cycle management", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Invoice generation", "pic": "Chhay", "est": 4, "status": "Completed"},
                {"task": "Payment recording", "pic": "Chhay", "est": 2, "status": "Completed"},
                {"task": "Payment reports", "pic": "Chhay", "est": 2, "status": "Completed"},
            ]
        },
        {
            "module": "13. Document Tracking",
            "status": "Completed",
            "progress": "80%",
            "tasks": [
                {"task": "Document type management", "pic": "Kanel", "est": 2, "status": "Completed"},
                {"task": "Document upload and storage", "pic": "Chhay", "est": 3, "status": "Completed"},
                {"task": "Document expiry tracking", "pic": "Chhay", "est": 2, "status": "Completed"},
                {"task": "Document verification workflow", "pic": "Chhay", "est": 3, "status": "Completed"},
            ]
        },
        {
            "module": "14. Audit & Logging",
            "status": "Completed",
            "progress": "90%",
            "tasks": [
                {"task": "Audit trail infrastructure", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Change tracking for critical data", "pic": "Kanel", "est": 4, "status": "Completed"},
                {"task": "User activity logging", "pic": "Kanel", "est": 2, "status": "Completed"},
                {"task": "Audit reports and search", "pic": "Chhay", "est": 2, "status": "Completed"},
            ]
        },
        {
            "module": "15. Dashboard & Analytics",
            "status": "In Progress",
            "progress": "35%",
            "tasks": [
                {"task": "Main dashboard layout", "pic": "Bonna", "est": 3, "status": "Completed"},
                {"task": "Employee statistics widgets", "pic": "Bonna", "est": 3, "status": "In Progress"},
                {"task": "Attendance charts", "pic": "Bonna", "est": 3, "status": "Pending"},
                {"task": "Leave statistics", "pic": "Bonna", "est": 2, "status": "Pending"},
                {"task": "Payroll summary dashboard", "pic": "Bonna", "est": 3, "status": "Pending"},
                {"task": "Custom report builder", "pic": "Kanel", "est": 5, "status": "Pending"},
            ]
        },
        {
            "module": "16. REST API",
            "status": "In Progress",
            "progress": "45%",
            "tasks": [
                {"task": "API authentication (JWT)", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Employee endpoints", "pic": "Kanel", "est": 3, "status": "Completed"},
                {"task": "Attendance endpoints", "pic": "Kanel", "est": 3, "status": "In Progress"},
                {"task": "Leave endpoints", "pic": "Kanel", "est": 2, "status": "In Progress"},
                {"task": "Payroll endpoints", "pic": "Kanel", "est": 3, "status": "Pending"},
                {"task": "API documentation (Swagger)", "pic": "Kanel", "est": 2, "status": "Pending"},
            ]
        },
        {
            "module": "17. Recruitment Module",
            "status": "Pending",
            "progress": "0%",
            "tasks": [
                {"task": "Job posting management", "pic": "TBD", "est": 4, "status": "Pending"},
                {"task": "Applicant tracking", "pic": "TBD", "est": 5, "status": "Pending"},
                {"task": "Interview scheduling", "pic": "TBD", "est": 3, "status": "Pending"},
                {"task": "Candidate evaluation", "pic": "TBD", "est": 4, "status": "Pending"},
                {"task": "Offer management", "pic": "TBD", "est": 3, "status": "Pending"},
            ]
        },
        {
            "module": "18. Training & Development",
            "status": "Pending",
            "progress": "0%",
            "tasks": [
                {"task": "Training course catalog", "pic": "TBD", "est": 3, "status": "Pending"},
                {"task": "Training enrollment", "pic": "TBD", "est": 3, "status": "Pending"},
                {"task": "Training attendance tracking", "pic": "TBD", "est": 2, "status": "Pending"},
                {"task": "Certification management", "pic": "TBD", "est": 3, "status": "Pending"},
                {"task": "Training reports", "pic": "TBD", "est": 2, "status": "Pending"},
            ]
        },
        {
            "module": "19. Performance Management",
            "status": "Pending",
            "progress": "0%",
            "tasks": [
                {"task": "Performance review templates", "pic": "TBD", "est": 4, "status": "Pending"},
                {"task": "Goal setting and tracking", "pic": "TBD", "est": 5, "status": "Pending"},
                {"task": "360-degree feedback", "pic": "TBD", "est": 5, "status": "Pending"},
                {"task": "Performance rating system", "pic": "TBD", "est": 3, "status": "Pending"},
                {"task": "Performance reports", "pic": "TBD", "est": 3, "status": "Pending"},
            ]
        },
        {
            "module": "20. Testing & QA",
            "status": "In Progress",
            "progress": "30%",
            "tasks": [
                {"task": "Unit test setup", "pic": "Kanel", "est": 2, "status": "Completed"},
                {"task": "Critical module testing", "pic": "All", "est": 10, "status": "In Progress"},
                {"task": "Integration testing", "pic": "All", "est": 8, "status": "Pending"},
                {"task": "User acceptance testing (UAT)", "pic": "All", "est": 10, "status": "Pending"},
                {"task": "Bug fixing", "pic": "All", "est": 15, "status": "In Progress"},
            ]
        },
        {
            "module": "21. Deployment & DevOps",
            "status": "In Progress",
            "progress": "40%",
            "tasks": [
                {"task": "Production server setup", "pic": "Kanel", "est": 3, "status": "Pending"},
                {"task": "Database backup strategy", "pic": "Kanel", "est": 2, "status": "Pending"},
                {"task": "SSL certificate configuration", "pic": "Kanel", "est": 1, "status": "Pending"},
                {"task": "Deployment scripts", "pic": "Kanel", "est": 2, "status": "Completed"},
                {"task": "Monitoring and logging setup", "pic": "Kanel", "est": 2, "status": "Pending"},
            ]
        },
    ]

    # Headers
    headers = ["No", "Module / Task", "PIC", "Est (Days)", "Status", "Progress", "Notes"]
    ws.append(headers)

    # Style header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Add data
    row_num = 2
    task_counter = 1

    for module_data in project_data:
        # Module row
        module_row = [
            "",
            module_data["module"],
            "",
            "",
            module_data["status"],
            module_data["progress"],
            ""
        ]
        ws.append(module_row)

        # Style module row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = module_font
            cell.fill = module_fill
            cell.border = thin_border
            if col_num in [5, 6]:  # Status and Progress columns
                cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num += 1

        # Task rows
        for task_data in module_data["tasks"]:
            task_row = [
                task_counter,
                f"  • {task_data['task']}",
                task_data["pic"],
                task_data["est"],
                task_data["status"],
                "",
                ""
            ]
            ws.append(task_row)

            # Style task row
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = task_font
                cell.fill = task_fill
                cell.border = thin_border

                # Center alignment for specific columns
                if col_num in [1, 3, 4, 5]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Status-based coloring
                if col_num == 5:
                    if task_data["status"] == "Completed":
                        cell.fill = completed_fill
                    elif task_data["status"] == "In Progress":
                        cell.fill = inprogress_fill
                    elif task_data["status"] == "Pending":
                        cell.fill = pending_fill

            row_num += 1
            task_counter += 1

    # Set column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 30

    # Add summary sheet
    summary_ws = wb.create_sheet(title="Project Summary")

    # Summary data
    summary_data = [
        ["NextHR Project Summary", ""],
        ["", ""],
        ["Project Name", "NextHR - Multi-tenant HR Management System"],
        ["Customer", "MEC (Manufacturing/Construction Company)"],
        ["Project Start Date", "September 2025"],
        ["Target Launch", "November 2025"],
        ["Overall Progress", "65%"],
        ["", ""],
        ["Team Members", ""],
        ["Backend/API Lead", "SOENG Kanel"],
        ["Fullstack Developer", "CHUNG Kuch Chhay"],
        ["Frontend Developer", "ORM Bonna"],
        ["", ""],
        ["Module Statistics", ""],
        ["Total Modules", "21"],
        ["Completed Modules", "6"],
        ["In Progress Modules", "10"],
        ["Pending Modules", "5"],
        ["", ""],
        ["Critical Path (November 2025)", ""],
        ["Priority 1", "Payroll System"],
        ["Priority 2", "Attendance & Timecard"],
        ["Priority 3", "Overtime Management"],
        ["", ""],
        ["Tech Stack", ""],
        ["Framework", "Django 5.2.1"],
        ["Database", "PostgreSQL (Multi-tenant)"],
        ["Frontend", "Bootstrap 5 + Tailwind CSS"],
        ["API", "Django REST Framework"],
    ]

    for row_data in summary_data:
        summary_ws.append(row_data)

    # Style summary sheet
    summary_ws['A1'].font = Font(bold=True, size=14)
    summary_ws['A1'].fill = header_fill
    summary_ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")

    for row in [3, 9, 14, 20, 25]:
        summary_ws[f'A{row}'].font = Font(bold=True, size=11)
        summary_ws[f'A{row}'].fill = module_fill

    summary_ws.column_dimensions['A'].width = 25
    summary_ws.column_dimensions['B'].width = 40

    # Save workbook
    filename = f"NextHR_Project_Plan_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"[OK] Project plan created successfully: {filename}")
    print(f"Total tasks: {task_counter - 1}")
    print(f"Total modules: {len(project_data)}")

    return filename

if __name__ == "__main__":
    create_project_plan()
