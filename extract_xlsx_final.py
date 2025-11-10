import zipfile
import os
import json
import openpyxl
import shutil
from datetime import datetime
import xml.etree.ElementTree as ET
import re

def get_exact_image_mapping(file_path):
    """
    Get the exact cell-to-image mapping from Excel metadata
    """
    mapping = {}
    
    with zipfile.ZipFile(file_path, 'r') as zip_ref:
        # Get image relationships
        image_rel_map = {}  # rId -> image filename
        
        if 'xl/richData/_rels/richValueRel.xml.rels' in zip_ref.namelist():
            content = zip_ref.read('xl/richData/_rels/richValueRel.xml.rels')
            root = ET.fromstring(content)
            
            for elem in root:
                if 'Relationship' in elem.tag:
                    rel_id = elem.get('Id')
                    target = elem.get('Target')
                    if '../media/' in target:
                        image_name = target.replace('../media/', '')
                        image_rel_map[rel_id] = image_name
        
        # Get cell-to-vm mapping from worksheet
        cell_vm_map = {}  # cell -> vm_id
        vm_cell_map = {}  # vm_id -> cell
        
        if 'xl/worksheets/sheet1.xml' in zip_ref.namelist():
            content = zip_ref.read('xl/worksheets/sheet1.xml')
            content_str = content.decode('utf-8')
            
            # Find all cells with vm attribute
            cell_vm_pattern = r'<c r="([^"]+)"[^>]*vm="(\d+)"'
            matches = re.findall(cell_vm_pattern, content_str)
            
            for cell_ref, vm_id in matches:
                vm_id = int(vm_id)
                cell_vm_map[cell_ref] = vm_id
                vm_cell_map[vm_id] = cell_ref
        
        # Get the ordering of rIds from richValueRel
        rel_order = []
        if 'xl/richData/richValueRel.xml' in zip_ref.namelist():
            content = zip_ref.read('xl/richData/richValueRel.xml')
            root = ET.fromstring(content)
            
            # Find all rel elements in order
            for elem in root.iter():
                if 'rel' in elem.tag.lower() and elem.tag != root.tag:
                    for attr_name, attr_value in elem.attrib.items():
                        if 'id' in attr_name.lower() and attr_value.startswith('rId'):
                            rel_order.append(attr_value)
                            break
        
        # Map vm_id to rId (assuming they're in order)
        for vm_id in sorted(vm_cell_map.keys()):
            idx = vm_id - 1  # vm_id is 1-based
            if idx < len(rel_order):
                rel_id = rel_order[idx]
                if rel_id in image_rel_map:
                    cell_ref = vm_cell_map[vm_id]
                    image_name = image_rel_map[rel_id]
                    
                    # Parse cell reference (e.g., B2 -> row=2, col=B)
                    col_letter = ''.join(c for c in cell_ref if c.isalpha())
                    row_num = int(''.join(c for c in cell_ref if c.isdigit()))
                    col_num = openpyxl.utils.column_index_from_string(col_letter)
                    
                    matrix_pos = f"{row_num}x{col_num}"
                    mapping[matrix_pos] = {
                        'image': image_name,
                        'cell': cell_ref,
                        'row': row_num,
                        'col': col_num,
                        'col_letter': col_letter
                    }
    
    return mapping

def extract_excel_with_exact_mapping(file_path, output_dir='extracted_final'):
    """
    Extract Excel data and images with exact cell mapping
    """
    
    # Clean and create output directories
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    
    os.makedirs(output_dir)
    images_dir = os.path.join(output_dir, 'images')
    os.makedirs(images_dir)
    
    # Get the exact mapping
    image_mapping = get_exact_image_mapping(file_path)
    
    print("=" * 70)
    print("EXTRACTING WITH EXACT CELL-IMAGE MAPPING")
    print("=" * 70)
    print("\nImage Mapping Found:")
    for matrix_pos, info in sorted(image_mapping.items()):
        print(f"  {info['cell']} ({matrix_pos}) -> {info['image']}")
    
    # Extract images with correct names
    with zipfile.ZipFile(file_path, 'r') as zip_ref:
        for matrix_pos, info in image_mapping.items():
            source_path = f"xl/media/{info['image']}"
            
            if source_path in zip_ref.namelist():
                image_data = zip_ref.read(source_path)
                
                # Get extension
                _, ext = os.path.splitext(info['image'])
                
                # Create filename with matrix position
                new_filename = f"{matrix_pos}{ext}"
                output_path = os.path.join(images_dir, new_filename)
                
                # Write file
                with open(output_path, 'wb') as f:
                    f.write(image_data)
                
                print(f"  Extracted: {new_filename} (from {info['image']})")
    
    # Load workbook and extract all data
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    all_data = {}
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_data = {
            'data': [],
            'images': [],
            'merged_cells': []
        }
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Extract cell data
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                cell_ref = f"{col_letter}{row_idx}"
                matrix_ref = f"{row_idx}x{col_idx}"
                
                cell_info = {
                    'row': row_idx,
                    'column': col_idx,
                    'column_letter': col_letter,
                    'value': cell.value,
                    'cell_reference': cell_ref,
                    'matrix_reference': matrix_ref,
                    'has_image': False
                }
                
                # Convert datetime to string
                if isinstance(cell_info['value'], datetime):
                    cell_info['value'] = cell_info['value'].strftime('%Y-%m-%d %H:%M:%S')
                
                # Check if this cell has an image
                if matrix_ref in image_mapping:
                    cell_info['has_image'] = True
                    cell_info['image'] = {
                        'path': os.path.join(images_dir, f"{matrix_ref}{os.path.splitext(image_mapping[matrix_ref]['image'])[1]}"),
                        'filename': f"{matrix_ref}{os.path.splitext(image_mapping[matrix_ref]['image'])[1]}",
                        'original': image_mapping[matrix_ref]['image'],
                        'matrix_position': matrix_ref
                    }
                
                row_data.append(cell_info)
            
            sheet_data['data'].append(row_data)
        
        # Extract merged cells
        for merged_range in sheet.merged_cells.ranges:
            sheet_data['merged_cells'].append({
                'range': str(merged_range),
                'min_row': merged_range.min_row,
                'max_row': merged_range.max_row,
                'min_col': merged_range.min_col,
                'max_col': merged_range.max_col
            })
        
        # Collect all images for this sheet
        for row in sheet_data['data']:
            for cell in row:
                if cell.get('has_image'):
                    sheet_data['images'].append({
                        'matrix_position': cell['matrix_reference'],
                        'cell': cell['cell_reference'],
                        'row': cell['row'],
                        'column': cell['column'],
                        'column_letter': cell['column_letter'],
                        'image_path': cell['image']['path'],
                        'filename': cell['image']['filename'],
                        'original_image': cell['image']['original']
                    })
        
        all_data[sheet_name] = sheet_data
    
    # Save complete data to JSON
    json_path = os.path.join(output_dir, 'excel_data.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, indent=2, ensure_ascii=False, default=str)
    
    # Create simplified JSON for Workers sheet
    create_workers_json(all_data, output_dir)
    
    # Create summary report
    create_detailed_report(all_data, image_mapping, output_dir)
    
    print(f"\n[OK] Data extracted to: {json_path}")
    print(f"[OK] Images extracted to: {images_dir}")
    print(f"[OK] Total images extracted: {len(image_mapping)}")
    
    return all_data

def create_workers_json(data, output_dir):
    """
    Create a simplified JSON for workers with all their associated images
    """
    workers_path = os.path.join(output_dir, 'workers_with_images.json')
    
    workers = []
    
    for sheet_name, sheet_data in data.items():
        if 'worker' in sheet_name.lower():
            # Process each row (skip header)
            for row_idx, row in enumerate(sheet_data['data'][1:], start=2):
                worker = {
                    'row': row_idx,
                    'data': {},
                    'images': {}
                }
                
                # Process each cell
                for cell in row:
                    col_letter = cell['column_letter']
                    value = cell['value']
                    
                    # Store cell data
                    worker['data'][col_letter] = value
                    
                    # Store image info if present
                    if cell.get('has_image'):
                        worker['images'][col_letter] = {
                            'matrix_position': cell['matrix_reference'],
                            'path': cell['image']['path'],
                            'filename': cell['image']['filename'],
                            'original': cell['image']['original']
                        }
                    
                    # Map specific columns
                    if cell['column'] == 1:  # A - NO
                        worker['no'] = value
                    elif cell['column'] == 3:  # C - NAME
                        worker['name'] = value
                    elif cell['column'] == 4:  # D - SEX
                        worker['sex'] = value
                    elif cell['column'] == 5:  # E - DOB
                        worker['dob'] = value
                    elif cell['column'] == 6:  # F - NATIONALITY
                        worker['nationality'] = value
                
                # Only add if worker has data
                if worker.get('no'):
                    workers.append(worker)
    
    with open(workers_path, 'w', encoding='utf-8') as f:
        json.dump(workers, f, indent=2, ensure_ascii=False)
    
    print(f"[OK] Workers JSON created: {workers_path}")
    return workers

def create_detailed_report(data, image_mapping, output_dir):
    """
    Create a detailed report with exact image positions
    """
    report_path = os.path.join(output_dir, 'extraction_report.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write("=" * 70 + "\n")
        f.write("EXCEL EXTRACTION REPORT - EXACT IMAGE MAPPING\n")
        f.write("=" * 70 + "\n\n")
        
        f.write("IMAGE LOCATIONS:\n")
        f.write("-" * 40 + "\n")
        
        # Group images by row
        images_by_row = {}
        for matrix_pos, info in image_mapping.items():
            row = info['row']
            if row not in images_by_row:
                images_by_row[row] = []
            images_by_row[row].append(info)
        
        for row in sorted(images_by_row.keys()):
            f.write(f"\nRow {row}:\n")
            for img_info in sorted(images_by_row[row], key=lambda x: x['col']):
                f.write(f"  Column {img_info['col_letter']} ({img_info['col']}): {img_info['image']}\n")
                f.write(f"    Matrix: {img_info['row']}x{img_info['col']}\n")
        
        f.write("\n" + "=" * 70 + "\n")
        f.write("SHEET SUMMARY:\n")
        f.write("-" * 40 + "\n")
        
        for sheet_name, sheet_data in data.items():
            f.write(f"\n{sheet_name}:\n")
            f.write(f"  Total rows: {len(sheet_data['data'])}\n")
            f.write(f"  Total columns: {len(sheet_data['data'][0]) if sheet_data['data'] else 0}\n")
            f.write(f"  Images in sheet: {len(sheet_data['images'])}\n")
            
            if sheet_data['images']:
                f.write("  Image cells:\n")
                for img in sheet_data['images']:
                    f.write(f"    {img['cell']} ({img['matrix_position']}): {img['filename']}\n")
    
    print(f"[OK] Report created: {report_path}")

def main():
    excel_file = 'worker_eform.xlsx'
    output_directory = 'extracted_final'
    
    if not os.path.exists(excel_file):
        print(f"Error: File '{excel_file}' not found!")
        return
    
    # Extract with exact mapping
    extracted_data = extract_excel_with_exact_mapping(excel_file, output_directory)
    
    # Display summary
    print("\n" + "=" * 70)
    print("EXTRACTION COMPLETE")
    print("=" * 70)
    
    # Read and display workers data
    workers_json_path = os.path.join(output_directory, 'workers_with_images.json')
    if os.path.exists(workers_json_path):
        with open(workers_json_path, 'r', encoding='utf-8') as f:
            workers = json.load(f)
            
            print(f"\nWorkers found: {len(workers)}")
            
            for worker in workers[:2]:  # Show first 2
                print(f"\n• Worker #{worker.get('no', 'N/A')} (Row {worker.get('row', 'N/A')}):")
                print(f"  Name: {worker.get('name', 'N/A')}")
                
                if worker.get('images'):
                    print(f"  Images in this row:")
                    for col, img_info in worker['images'].items():
                        print(f"    Column {col}: {img_info['filename']} (original: {img_info['original']})")
                
                print(f"  Sex: {worker.get('sex', 'N/A')}")
                print(f"  DOB: {worker.get('dob', 'N/A')}")

if __name__ == "__main__":
    main()